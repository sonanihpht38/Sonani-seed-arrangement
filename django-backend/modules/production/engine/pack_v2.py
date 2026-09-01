r"""
General multi-plate packer for CVD-diamond blocks (rectangles, 90 deg rotation).

Algorithm (data-agnostic, nothing tuned to a specific sheet):
  1. Read every row, keep blocks whose thickness HEIGHT is in [T_LO, T_HI].
  2. Treat each block as its true LENGTH_1 x WIDTH rectangle.
  3. Sort ALL blocks largest-area first (global reorder -> not sheet sequence).
  4. Bottom-left best-fit placement: for each block try BOTH orientations
     (0 and 90 deg) over the whole usable area, place it at the lowest-then-
     leftmost feasible spot with zero gap. Whichever orientation lands lower wins.
  5. Blocks that don't fit the current plate roll to a new plate. Repeat until
     every block is placed.

Two usable-area definitions are generated as separate outputs:
  * square : inscribed square of the Ø87 circle (side = 87/sqrt2) -- "ignore curve"
  * circle : the full Ø87 circular area

Outputs per mode:
  images_<mode>\plate_01.png ...      (one layout per plate)
  arrangement_<mode>.xlsx             (Summary sheet + one Plate_NN sheet each)

Usage:
  .\venv\Scripts\python.exe pack_v2.py BLOCK.xlsx
"""
import os, math, openpyxl
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Circle, Rectangle as MplRect, Polygon as MplPoly
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from shapely.geometry import MultiPoint

XLSX = "BLOCK.xlsx"                       # default; overridden by CLI in __main__
PLATE_D, USABLE_D = 90.0, 87.0
R = USABLE_D / 2.0                       # 43.5 mm
S2 = (USABLE_D / math.sqrt(2)) / 2.0     # half side of inscribed square (~30.76 mm)
T_LO, T_HI = 0.67, 0.73
GRID = 0.5                               # placement scan step (mm)

def fnum(x):
    try: return float(x)
    except (TypeError, ValueError): return None

def load_blocks(path, shapes=("square", "rectangle"), square_tol=0.05):
    """Load seeds, keep those within the thickness window (priority gate), then
    keep only the allowed shape(s). A seed is 'square' when |L-W| <= square_tol
    * max(L,W) (default 5%); otherwise 'rectangle'. `shapes` is a collection that
    may contain 'square', 'rectangle', or both."""
    shapes = set(shapes)
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb.active
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        _b, stock, _pcs, cts, L, W, H = r
        L, W, H, cts = fnum(L), fnum(W), fnum(H), fnum(cts)
        if None in (L, W, H): continue
        if not (T_LO <= H <= T_HI):                 # thickness first
            continue
        sh = "square" if abs(L - W) <= square_tol * max(L, W) else "rectangle"
        if sh not in shapes:                        # shape filter (secondary)
            continue
        out.append({"stock": stock, "cts": cts or 0.0, "L": L, "W": W, "H": H, "shape": sh})
    return out

# ------------------------------------------------------------------ geometry grid
def make_grid(mode):
    if mode == "square":
        g = np.arange(-S2, S2 + 1e-9, GRID)
    else:
        g = np.arange(-R, R + 1e-9, GRID)
    X, Y = np.meshgrid(g, g)             # X[i,j]=g[j], Y[i,j]=g[i]
    return g, X, Y

def best_bottom_left(w, h, occupied, grid, mode):
    """Lowest-then-leftmost feasible bottom-left (x,y) for a w x h rect, or None."""
    g, X, Y = grid
    if mode == "square":
        ok = (X + w <= S2) & (Y + h <= S2)          # grid already >= -S2
    else:
        cx = np.maximum(np.abs(X), np.abs(X + w))    # farthest corner from center
        cy = np.maximum(np.abs(Y), np.abs(Y + h))
        ok = (cx * cx + cy * cy) <= R * R
    for ox, oy, ow, oh in occupied:                  # remove overlaps (touching allowed)
        ok &= ~((X < ox + ow) & (X + w > ox) & (Y < oy + oh) & (Y + h > oy))
    iy, ix = np.where(ok)
    if iy.size == 0:
        return None
    Yv, Xv = g[iy], g[ix]
    k = np.lexsort((Xv, Yv))[0]                       # primary Y, secondary X
    return float(g[ix[k]]), float(g[iy[k]])

# ------------------------------------------------------------------ multi-plate pack
# CLEARANCE = minimum gap to force between blocks (mm). 0 = packed as tight as
# possible. Set e.g. 0.5 if the growth process needs a uniform seed spacing.
CLEARANCE = 0.0

# Different placement orders suit different size mixes; we try all and keep the
# densest plate (= least gap). This is what actively minimises the gap.
ORDERINGS = [
    lambda d: -(d["L"] * d["W"]),        # largest area first
    lambda d: -max(d["L"], d["W"]),      # longest side first
    lambda d: -(d["L"] + d["W"]),        # largest perimeter first
    lambda d: -min(d["L"], d["W"]),      # largest short-side first
]

def fill_one_plate(ordered, grid, mode):
    occupied, placed, leftover = [], [], []
    c = CLEARANCE
    for blk in ordered:
        best = None
        for w, h, ang in ((blk["L"], blk["W"], 0), (blk["W"], blk["L"], 90)):
            pos = best_bottom_left(w + c, h + c, occupied, grid, mode)
            if pos and (best is None or (pos[1], pos[0]) < (best[0][1], best[0][0])):
                best = (pos, w, h, ang)
        if best:
            (x, y), w, h, ang = best
            occupied.append((x, y, w + c, h + c))
            placed.append({**blk, "x": x, "y": y, "w": w, "h": h, "angle": ang})
        else:
            leftover.append(blk)
    return placed, leftover

INSCRIBED = USABLE_D / math.sqrt(2)          # 61.52 mm square fits inside Ø87

def _split_free(F, U):
    fx, fy, fw, fh = F; ux, uy, uw, uh = U
    if ux >= fx + fw or ux + uw <= fx or uy >= fy + fh or uy + uh <= fy:
        return [F]                            # no overlap
    out = []
    if ux > fx:                 out.append((fx, fy, ux - fx, fh))
    if ux + uw < fx + fw:       out.append((ux + uw, fy, fx + fw - ux - uw, fh))
    if uy > fy:                 out.append((fx, fy, fw, uy - fy))
    if uy + uh < fy + fh:       out.append((fx, uy + uh, fw, fy + fh - uy - uh))
    return out

def _prune(free):
    def inside(a, b):
        return (b[0] <= a[0] + 1e-6 and b[1] <= a[1] + 1e-6 and
                a[0] + a[2] <= b[0] + b[2] + 1e-6 and a[1] + a[3] <= b[1] + b[3] + 1e-6)
    keep = []
    for i, a in enumerate(free):
        if a[2] <= 1e-6 or a[3] <= 1e-6:
            continue
        if not any(j != i and inside(a, b) for j, b in enumerate(free)):
            keep.append(a)
    return keep

def maxrects_one_bin(rects, side):
    """MaxRects Best-Short-Side-Fit with 90 deg rotation into a side x side bin."""
    free = [(0.0, 0.0, side, side)]; placed = []; leftover = []
    for blk in rects:
        best = None                                # (short, long, x, y, w, h, ang)
        for w, h, ang in ((blk["L"], blk["W"], 0), (blk["W"], blk["L"], 90)):
            for fx, fy, fw, fh in free:
                if w <= fw + 1e-9 and h <= fh + 1e-9:
                    s, l = sorted((fw - w, fh - h))
                    if best is None or (s, l) < (best[0], best[1]):
                        best = (s, l, fx, fy, w, h, ang)
        if best is None:
            leftover.append(blk); continue
        _, _, x, y, w, h, ang = best
        placed.append({**blk, "x": x, "y": y, "w": w, "h": h, "angle": ang})
        nf = []
        for f in free:
            nf += _split_free(f, (x, y, w, h))
        free = _prune(nf)
    return placed, leftover

def pack_grid(blocks):
    """Tight rectangle packing into the inscribed square, centered on the plate."""
    remaining = sorted(blocks, key=lambda d: -(d["L"] * d["W"]))
    plates, unplaceable = [], []
    while remaining:
        placed, leftover = maxrects_one_bin(remaining, INSCRIBED)
        if not placed:
            unplaceable = leftover; break
        for p in placed:                            # center the square pack on the plate
            p["x"] -= INSCRIBED / 2; p["y"] -= INSCRIBED / 2
        plates.append(placed)
        area = sum(p["w"] * p["h"] for p in placed)
        print(f"  [grid] plate {len(plates):02d}: {len(placed):2d} blocks, "
              f"{inner_gap(placed):.1f}% inter-seed gap, "
              f"{100 * area / (math.pi * R * R):.1f}% plate-fill, {len(leftover)} left")
        remaining = leftover
    return plates, unplaceable

def pack_rows(blocks):
    """Side-by-side row layout: blocks normalized landscape, sorted by height so
    each row is uniform, laid edge-to-edge L->R, rows stacked touching, centered.
    Gives the cleanest aligned-grid look with minimum gap between neighbours."""
    items = []
    for b in blocks:
        if b["W"] > b["L"]:
            w, h, ang = b["W"], b["L"], 90
        else:
            w, h, ang = b["L"], b["W"], 0
        items.append({**b, "w": w, "h": h, "angle": ang})
    items.sort(key=lambda d: -d["h"])                  # tall rows first -> uniform rows
    budget = INSCRIBED
    rows, row, rw, rh = [], [], 0.0, 0.0
    for it in items:
        if row and rw + it["w"] > budget + 1e-9:
            rows.append((row, rh)); row, rw, rh = [], 0.0, 0.0
        row.append(it); rw += it["w"]; rh = max(rh, it["h"])
    if row: rows.append((row, rh))
    groups, cur, cur_h = [], [], 0.0                   # pack rows into plates by height
    for r, h in rows:
        if cur and cur_h + h > budget + 1e-9:
            groups.append(cur); cur, cur_h = [], 0.0
        cur.append((r, h)); cur_h += h
    if cur: groups.append(cur)
    plates = []
    for grp in groups:
        total_h = sum(h for _, h in grp)
        placed, y = [], -total_h / 2
        for r, h in grp:
            total_w = sum(it["w"] for it in r)
            x = -total_w / 2                           # center each row
            for it in r:
                placed.append({**it, "x": x, "y": y}); x += it["w"]
            y += h
        plates.append(placed)
        area = sum(p["w"] * p["h"] for p in placed)
        print(f"  [rows] plate {len(plates):02d}: {len(placed):2d} blocks, "
              f"{inner_gap(placed):.1f}% inter-seed gap, "
              f"{100 * area / (math.pi * R * R):.1f}% plate-fill")
    return plates, []

def layout_grid(take, C):
    """Arrange `take` blocks row-major into C columns as an aligned lattice
    (shared column x-lines and row y-lines), each block centered in its cell.
    Returns placed list, or None if the grid's corners fall outside Ø87."""
    nrows = math.ceil(len(take) / C)
    grid = [take[i * C:(i + 1) * C] for i in range(nrows)]
    col_w = [0.0] * C; row_h = [0.0] * nrows
    for i, row in enumerate(grid):
        for j, b in enumerate(row):
            col_w[j] = max(col_w[j], b["w"]); row_h[i] = max(row_h[i], b["h"])
    total_w, total_h = sum(col_w), sum(row_h)
    if math.hypot(total_w / 2, total_h / 2) > R:          # grid corner outside circle
        return None
    xleft, acc = [], -total_w / 2
    for j in range(C): xleft.append(acc); acc += col_w[j]
    ytop, acc = [], total_h / 2
    for i in range(nrows): ytop.append(acc); acc -= row_h[i]
    placed = []
    for i, row in enumerate(grid):
        for j, b in enumerate(row):
            cx = xleft[j] + col_w[j] / 2; cy = ytop[i] - row_h[i] / 2
            placed.append({**b, "x": cx - b["w"] / 2, "y": cy - b["h"] / 2})
    return placed

def pack_lattice(blocks):
    """Aligned-lattice arrangement (rows AND columns line up), like grown tiles."""
    items = []
    for b in blocks:
        if b["W"] > b["L"]:
            w, h, ang = b["W"], b["L"], 90
        else:
            w, h, ang = b["L"], b["W"], 0
        items.append({**b, "w": w, "h": h, "angle": ang})
    items.sort(key=lambda d: (-d["h"], -d["w"]))          # uniform rows -> straight seams
    queue = items[:]
    plates, unplaceable = [], []
    while queue:
        best = None
        for C in range(1, 9):
            if C > len(queue): break
            layout = None
            for n in range(C, len(queue) + 1):            # grow until it stops fitting
                lay = layout_grid(queue[:n], C)
                if lay is None: break
                layout = lay
            if layout and (best is None or len(layout) > len(best)):
                best = layout
        if not best:
            unplaceable = queue; break
        plates.append(best)
        area = sum(p["w"] * p["h"] for p in best)
        print(f"  [lattice] plate {len(plates):02d}: {len(best):2d} blocks, "
              f"{inner_gap(best):.1f}% inter-seed gap, "
              f"{100 * area / (math.pi * R * R):.1f}% plate-fill")
        queue = queue[len(best):]
    return plates, unplaceable

def _mixed_landscape(blocks):
    """Normalize every seed to landscape (w >= h, 90deg flag) and sort tall-first
    -- the queue order the mixed/edgefill row packer consumes."""
    items = []
    for b in blocks:
        if b["W"] > b["L"]:
            w, h, ang = b["W"], b["L"], 90
        else:
            w, h, ang = b["L"], b["W"], 0
        items.append({**b, "w": w, "h": h, "angle": ang})
    items.sort(key=lambda d: -d["h"])
    return items

# SEED-SIZE OPTIMISATION — biggest seeds in the middle of the plate, smallest
# toward the rim.
#
# The VERTICAL half of that rule was already here and is untouched:
# _mixed_landscape sorts the queue tall-first and rows are laid centre-out, so
# the tallest seeds settle into the middle rows and the short ones end up in the
# shallow rows against the rim.
#
# What was missing is the HORIZONTAL half. Within a row, seeds were laid in
# whatever order the queue happened to hand them over, so a row could read
# 4-13-6-12-5 across the plate. _centre_out_row re-orders each row into an
# "organ pipe": widest in the middle, tapering to the narrowest at both chord
# ends.
#
# This is a REORDER and nothing else. The row keeps the same seeds, the same
# height (rhh = max h, unchanged by permutation) and the same total width, and
# it is still centred on x = 0 — so it occupies the identical x-interval and
# every chord/overlap property the packer relies on is preserved by
# construction. No seed is added, dropped, resized or rotated.
CENTRE_OUT_ROWS = True


def _centre_out_row(row):
    """Order one row widest-in-the-middle, narrowest at the two ends.

    Sort widest first, then deal alternately to the right and left of centre and
    read the row back out left-to-right. The two widest seeds end up adjacent in
    the middle and each successive pair steps outward, which is the arrangement
    the shop floor asks for: the big stones where the plate is fullest and the
    small ones where it curves away.

    Ties are broken by the seed's own height so the result is deterministic —
    `generate_final` re-packs a saved run to redraw it, and two runs of the same
    pool must lay out identically or the finalized image would not match the
    arrangement it claims to show.
    """
    ordered = sorted(row, key=lambda b: (-b["w"], -b["h"], str(b.get("stock", ""))))
    left, right = [], []
    for i, b in enumerate(ordered):
        (right if i % 2 == 0 else left).append(b)
    return left[::-1] + right


def _mixed_one_plate(queue):
    """Build ONE center-out mixed plate, popping the seeds it uses off `queue`
    (mutated in place). Each row spans the FULL circle chord at its height and is
    laid out with a CLEARANCE gap between neighbours (the user's "distance between
    seeds", in mm); rows are stacked with the same gap between them. CLEARANCE = 0
    reproduces the old edge-to-edge behaviour exactly.
    Returns the placed list, or None if not even a center row fits."""
    c = CLEARANCE                                    # distance between seeds (mm)

    def chord(y_edge):
        v = R * R - y_edge * y_edge
        return math.sqrt(v) if v > 0 else 0.0

    def take_row(budget):
        # each seat after the first also reserves a `c` gap before it
        row, w = [], 0.0
        while queue:
            need = queue[0]["w"] + (c if row else 0.0)
            if w + need <= budget + 1e-9:
                it = queue.pop(0); row.append(it); w += need
            else:
                break
        return row

    if not queue:
        return None
    h_c = queue[0]["h"]
    center = take_row(2 * chord(h_c / 2))
    if not center:
        return None
    rh = max(b["h"] for b in center)
    rows = [(center, -rh / 2, rh)]
    y_top, y_bot = rh / 2, -rh / 2
    while queue:
        progressed = False
        if queue:                                   # add a row above (leave a `c` gap first)
            outer = y_top + c + queue[0]["h"]
            r = take_row(2 * chord(outer)) if outer <= R else []
            if r:
                rhh = max(b["h"] for b in r)
                rows.append((r, y_top + c, rhh)); y_top += c + rhh; progressed = True
        if queue:                                   # add a row below (leave a `c` gap first)
            outer = y_bot - c - queue[0]["h"]
            r = take_row(2 * chord(abs(outer))) if outer >= -R else []
            if r:
                rhh = max(b["h"] for b in r)
                y_bot -= c + rhh; rows.append((r, y_bot, rhh)); progressed = True
        if not progressed:
            break
    placed = []
    for r, yb, rhh in rows:
        if CENTRE_OUT_ROWS:
            r = _centre_out_row(r)               # widest seeds to the middle
        tw = sum(b["w"] for b in r) + c * (len(r) - 1); x = -tw / 2   # center row; `c` gap between seeds
        for b in r:
            placed.append({**b, "x": x, "y": yb}); x += b["w"] + c
    return placed

def pack_mixed(blocks):
    """Mixed-band row packing built center-out: each row uses the FULL circle
    width at its height (chord), so wide middle rows hold ~5 large blocks and
    edge rows hold more small ones. Tall blocks settle in the center, short ones
    to the edges. Maximises blocks/plate while keeping straight horizontal seams."""
    queue = _mixed_landscape(blocks)
    plates, unplaceable = [], []
    while queue:
        placed = _mixed_one_plate(queue)
        if not placed:
            unplaceable = queue; break
        plates.append(placed)
        area = sum(p["w"] * p["h"] for p in placed)
        print(f"  [mixed] plate {len(plates):02d}: {len(placed):2d} blocks, "
              f"spans Ø{span_mm(placed):.1f}/90 mm, "
              f"{inner_gap(placed):.1f}% inter-seed gap, "
              f"{100 * area / (math.pi * R * R):.1f}% plate-fill")
    return plates, unplaceable

def seed_corners(p):
    """4 (x,y) corners of a placed seed. Rotation-aware: a gap-fill seed carries
    `theta` (degrees, any angle) + center `cx,cy`, so its w x h footprint is
    rotated about its center; an axis-aligned seed (upright pack, angle already
    baked into w/h) uses its bottom-left `x,y`."""
    if p.get("theta") is not None:
        a = math.radians(p["theta"])
        ca, sa = math.cos(a), math.sin(a)
        hw, hh = p["w"] / 2.0, p["h"] / 2.0
        cx, cy = p["cx"], p["cy"]
        return [(cx + dx * ca - dy * sa, cy + dx * sa + dy * ca)
                for dx, dy in ((-hw, -hh), (hw, -hh), (hw, hh), (-hw, hh))]
    x, y, w, h = p["x"], p["y"], p["w"], p["h"]
    return [(x, y), (x + w, y), (x + w, y + h), (x, y + h)]

def inner_gap(placed):
    """% of the seed cluster's footprint (convex hull) that is empty space
    BETWEEN blocks -- the real inter-seed gap, excluding the round-plate margin."""
    if len(placed) < 3:
        return 0.0
    pts = []
    for p in placed:
        pts += seed_corners(p)
    hull = MultiPoint(pts).convex_hull.area
    blk = sum(p["w"] * p["h"] for p in placed)
    return 100 * (1 - blk / hull) if hull else 0.0

def span_mm(placed):
    """Diameter occupied by the arrangement (2 x farthest block corner from the
    centered cluster's center) -- how much of the Ø90 plate the seeds take up."""
    m = 0.0
    for p in placed:
        for cx, cy in seed_corners(p):
            m = max(m, math.hypot(cx, cy))
    return 2 * m

def square_fill(placed):
    """% of the inscribed-square usable area filled by seeds (the 'ignore-curve'
    plate space). Approaches 99.99% when the seed mix includes small fillers."""
    return 100 * sum(p["w"] * p["h"] for p in placed) / (INSCRIBED * INSCRIBED)

def pack_maxfill(blocks):
    """MAX-FILL scenario: zero-gap MaxRects best-fit (90 deg rotation) into the
    inscribed square (curve ignored). Large seeds placed first, then progressively
    smaller ones fill the gaps -> packs toward ~100% when small seeds exist.
    Fully data-agnostic: works for any square/rectangular seeds >= 2x2 mm."""
    remaining = sorted(blocks, key=lambda d: -(d["L"] * d["W"]))   # consider all, big first
    plates, unplaceable = [], []
    while remaining:
        placed, leftover = maxrects_one_bin(remaining, INSCRIBED)
        if not placed:
            unplaceable = leftover; break
        for p in placed:
            p["x"] -= INSCRIBED / 2; p["y"] -= INSCRIBED / 2      # center on plate
        plates.append(placed)
        print(f"  [maxfill] plate {len(plates):02d}: {len(placed):3d} blocks, "
              f"{square_fill(placed):.2f}% square-fill, "
              f"spans Ø{span_mm(placed):.1f}/90 mm, {len(leftover)} left")
        remaining = leftover
    return plates, unplaceable

# ----------------------------------------------------- rotated-rectangle geometry
def _rot_corners(cx, cy, w, h, deg):
    """4 corners of a w x h rect centered at (cx,cy) rotated by `deg` (any angle)."""
    a = math.radians(deg); ca, sa = math.cos(a), math.sin(a)
    return [(cx + dx * ca - dy * sa, cy + dx * sa + dy * ca)
            for dx, dy in ((-w / 2, -h / 2), (w / 2, -h / 2), (w / 2, h / 2), (-w / 2, h / 2))]

def _rects_overlap(A, B):
    """Separating-Axis test for two convex rectangles (corner lists). Touching
    edges (zero overlap) count as NOT overlapping, so seeds may sit flush."""
    for poly in (A, B):
        for i in range(len(poly)):
            x1, y1 = poly[i]; x2, y2 = poly[(i + 1) % len(poly)]
            nx, ny = -(y2 - y1), (x2 - x1)               # edge normal = separating axis
            amin = amax = None
            for x, y in A:
                p = x * nx + y * ny
                amin = p if amin is None else min(amin, p)
                amax = p if amax is None else max(amax, p)
            bmin = bmax = None
            for x, y in B:
                p = x * nx + y * ny
                bmin = p if bmin is None else min(bmin, p)
                bmax = p if bmax is None else max(bmax, p)
            if amax <= bmin + 1e-9 or bmax <= amin + 1e-9:
                return False                              # gap on this axis -> no overlap
    return True

def _inside_circle(corners, rad):
    return all(x * x + y * y <= rad * rad + 1e-6 for x, y in corners)

GAPSTEP = 2.0                                            # gap-fill position scan step (mm)
GAP_ANGLES = (0, 90, 45, 135, 30, 60)                    # upright first, then rotate to fit
GAP_FAIL_LIMIT = 6                                       # stop a plate after N seeds in a row don't fit

def _gap_fill(placed, leftover, rad):
    """Second pass: try to drop leftover seeds into the residual gaps of an already
    upright-packed plate, KEEPING THEM UPRIGHT (0/90) when possible and only rotating
    to an arbitrary angle if that's the only way to fit. Seeds are tried smallest-first
    (small seeds fit small gaps); once GAP_FAIL_LIMIT small seeds in a row don't fit,
    the gaps are exhausted and every remaining seed rolls to the next plate.
    Returns the still-leftover list (order preserved largest-first like the input)."""
    occ = [seed_corners(p) for p in placed]
    aabb = [(p["x"], p["y"], p["w"], p["h"]) for p in placed if p.get("theta") is None]
    g = [round(-rad + k * GAPSTEP, 3) for k in range(int(2 * rad / GAPSTEP) + 1)]
    # candidate centers inside the circle AND not already covered by an upright seed
    # (precomputed once -> failing seeds only re-scan the actual gaps, not the whole disk)
    free = [(cx, cy) for cy in g for cx in g
            if cx * cx + cy * cy <= rad * rad
            and not any(ox <= cx <= ox + ow and oy <= cy <= oy + oh
                        for ox, oy, ow, oh in aabb)]
    ordered = sorted(leftover, key=lambda d: d["L"] * d["W"])   # small-first into small gaps
    rest, fails, i = [], 0, 0
    while i < len(ordered):
        seed = ordered[i]; i += 1
        L, W = seed["L"], seed["W"]
        spot = None
        for deg in GAP_ANGLES:
            for cx, cy in free:
                corners = _rot_corners(cx, cy, L, W, deg)
                if not _inside_circle(corners, rad):
                    continue
                if any(_rects_overlap(corners, o) for o in occ):
                    continue
                spot = (cx, cy, deg, corners); break
            if spot: break
        if spot:
            cx, cy, deg, corners = spot
            placed.append({**seed, "cx": cx, "cy": cy, "w": L, "h": W,
                           "theta": float(deg), "angle": 0,
                           "x": cx - L / 2, "y": cy - W / 2})
            occ.append(corners); fails = 0
        else:
            rest.append(seed); fails += 1
            if fails >= GAP_FAIL_LIMIT:           # gaps exhausted: roll the rest over untried
                rest.extend(ordered[i:]); break
    rest.sort(key=lambda d: -(d["L"] * d["W"]))   # restore largest-first for the next plate
    return rest

def pack_edgefill(blocks):
    """Same ZERO-GAP 'mixed' row packing (tight edge-to-edge rows reaching the
    circle edges) as the base, PLUS a gap-fill pass: leftover seeds are rotated to
    ANY angle ONLY to drop into the residual gaps the rows leave (the top/bottom
    caps and row-end triangles near the curved edge), pushing coverage higher.
    Rotation is the exception; the tight upright rows keep zero gap between seeds."""
    queue = _mixed_landscape(blocks)
    plates, unplaceable = [], []
    while queue:
        placed = _mixed_one_plate(queue)                  # PASS A: zero-gap mixed rows
        if not placed:
            unplaceable = queue; break
        n_rows = len(placed)
        queue[:] = _gap_fill(placed, queue, R)            # PASS B: rotate leftovers into gaps
        queue.sort(key=lambda d: -d["h"])                 # restore tall-first order for next plate
        plates.append(placed)
        area = sum(p["w"] * p["h"] for p in placed)
        n_rot = sum(1 for p in placed if p.get("theta"))
        print(f"  [edgefill] plate {len(plates):02d}: {len(placed):2d} seeds "
              f"({n_rows} rows + {n_rot} rotated gap-fill), "
              f"spans Ø{span_mm(placed):.1f}/{PLATE_D:g} mm, "
              f"{inner_gap(placed):.1f}% inter-seed gap, "
              f"{100 * area / (math.pi * R * R):.1f}% plate-fill, {len(queue)} left")
    return plates, unplaceable

def pack(blocks, mode):
    grid = make_grid(mode)
    remaining = blocks
    plates, unplaceable = [], []
    while remaining:
        best = None                                   # (placed_area, placed, leftover)
        for keyf in ORDERINGS:
            placed, leftover = fill_one_plate(sorted(remaining, key=keyf), grid, mode)
            area = sum(p["w"] * p["h"] for p in placed)
            if best is None or area > best[0]:
                best = (area, placed, leftover)
        area, placed, leftover = best
        if not placed:                                # nothing fits an empty plate
            unplaceable = leftover
            break
        # center the packed cluster on the plate (rigid shift; gaps unchanged)
        xs = [p["x"] for p in placed] + [p["x"] + p["w"] for p in placed]
        ys = [p["y"] for p in placed] + [p["y"] + p["h"] for p in placed]
        dx, dy = -(min(xs) + max(xs)) / 2, -(min(ys) + max(ys)) / 2
        for p in placed:
            p["x"] += dx; p["y"] += dy
        plates.append(placed)
        util = 100 * area / (math.pi * R * R)
        print(f"  [{mode}] plate {len(plates):02d}: {len(placed):2d} blocks, "
              f"{inner_gap(placed):.1f}% inter-seed gap, {util:.1f}% plate-fill, "
              f"{len(leftover)} left")
        remaining = leftover
    return plates, unplaceable

# ------------------------------------------------------------------ square-core + sides
def pack_squarecore(blocks):
    """Per plate: PASS A fills the inscribed square (the 'square first' core),
    then PASS B fills the four circular side segments with remaining blocks so
    the full plate radius is used. Largest-area-first, 90 deg rotation, zero gap."""
    gsq, gci = make_grid("square"), make_grid("circle")
    remaining = sorted(blocks, key=lambda d: -(d["L"] * d["W"]))
    plates, unplaceable = [], []
    while remaining:
        occupied, placed = [], []
        def try_place(blk, grid, mode):
            best = None
            for w, h, ang in ((blk["L"], blk["W"], 0), (blk["W"], blk["L"], 90)):
                pos = best_bottom_left(w, h, occupied, grid, mode)
                if pos and (best is None or (pos[1], pos[0]) < (best[0][1], best[0][0])):
                    best = (pos, w, h, ang)
            if best:
                (x, y), w, h, ang = best
                occupied.append((x, y, w, h))
                placed.append({**blk, "x": x, "y": y, "w": w, "h": h, "angle": ang,
                               "zone": mode})
                return True
            return False
        still    = [b for b in remaining if not try_place(b, gsq, "square")]   # core
        n_core   = len(placed)
        leftover = [b for b in still if not try_place(b, gci, "circle")]       # sides
        if not placed:
            unplaceable = leftover; break
        if len(placed) == n_core:        # nothing reached the sides -> center the core
            xs = [p["x"] for p in placed] + [p["x"] + p["w"] for p in placed]
            ys = [p["y"] for p in placed] + [p["y"] + p["h"] for p in placed]
            dx, dy = -(min(xs) + max(xs)) / 2, -(min(ys) + max(ys)) / 2
            for p in placed: p["x"] += dx; p["y"] += dy
        plates.append(placed)
        print(f"  [squarecore] plate {len(plates):02d}: {len(placed):2d} blocks "
              f"({n_core} core + {len(placed) - n_core} sides), {len(leftover)} left")
        remaining = leftover
    return plates, unplaceable

# ------------------------------------------------------------------ render
def render(placed, plate_no, mode, path):
    n = len(placed); area = sum(p["w"] * p["h"] for p in placed)
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.add_patch(Circle((0, 0), PLATE_D / 2, fc="#e9e9ec", ec="#888", lw=1.5, zorder=0))
    ax.add_patch(Circle((0, 0), R, fc="none", ec="#c0392b", lw=1.4, ls="--", zorder=1))
    if mode in ("square", "squarecore", "grid", "rows", "lattice"):
        ax.add_patch(MplRect((-S2, -S2), 2 * S2, 2 * S2, fc="none",
                             ec="#2471a3", lw=1.2, ls=":", zorder=1))
    span = span_mm(placed)                                  # occupied-diameter ring
    ax.add_patch(Circle((0, 0), span / 2, fc="none", ec="#1f8a3b", lw=1.6, zorder=1))
    ax.text(0, span / 2 + 1.5, f"Ø{span:.1f} mm occupied", ha="center", va="bottom",
            fontsize=8, color="#1f8a3b", zorder=4)
    cmap = plt.cm.viridis
    for i, p in enumerate(placed):
        frac = i / max(1, n - 1)
        if p.get("theta") is not None:                     # gap-fill seed: rotated polygon
            ax.add_patch(MplPoly(seed_corners(p), closed=True, fc=cmap(frac),
                                 ec="white", lw=0.6, alpha=0.92, zorder=2))
            tx, ty = p["cx"], p["cy"]
        else:
            ax.add_patch(MplRect((p["x"], p["y"]), p["w"], p["h"],
                         fc=cmap(frac), ec="white", lw=0.6, alpha=0.92, zorder=2))
            tx, ty = p["x"] + p["w"] / 2, p["y"] + p["h"] / 2
        rot_note = (f"\n↺ {p['theta']:.0f}°" if p.get("theta") is not None
                    else ("\n↺ 90°" if p["angle"] else ""))
        detail = (f"{p['stock']}\n"
                  f"{p['L']:.1f}×{p['W']:.1f} mm\n"
                  f"H {p['H']:.2f} mm\n"
                  f"{p['cts']:.2f} ct" + rot_note)
        fs = min(6.0, max(3.0, min(p["w"], p["h"]) * 0.32))   # scale text to block size
        txt_color = "white" if frac < 0.6 else "#101010"      # readable on light viridis
        rotation = p["theta"] if p.get("theta") is not None else 0
        ax.text(tx, ty, detail, ha="center", va="center", fontsize=fs, color=txt_color,
                linespacing=1.25, rotation=rotation, rotation_mode="anchor", zorder=3)
    lim = PLATE_D / 2 + 4
    ax.set_xlim(-lim, lim); ax.set_ylim(-lim, lim); ax.set_aspect("equal"); ax.axis("off")
    util = 100 * area / (math.pi * R * R)
    sq = (f"{square_fill(placed):.1f}% square-fill  ·  "
          if mode in ("maxfill", "grid", "square", "squarecore", "lattice") else "")
    ax.set_title(f"[{mode}]  Plate {plate_no:02d}  ·  {n} blocks  ·  "
                 f"spans Ø{span_mm(placed):.1f} of {PLATE_D:g} mm  ·  {sq}"
                 f"{inner_gap(placed):.1f}% inter-seed gap  ·  {util:.1f}% plate-fill  ·  "
                 f"{sum(p['cts'] for p in placed):.2f} ct\n"
                 f"Ø{PLATE_D:g} plate · Ø{2 * R:g} usable (dashed) · rect L×W, ↺ = rotated 90°",
                 fontsize=9)
    plt.savefig(path, dpi=150, bbox_inches="tight"); plt.close(fig)

# ------------------------------------------------------------------ seed-map image
def render_seed_map_img(plates, mode, path):
    """Render the seed -> plate mapping as a table image (one shaded block per
    plate, listing its stock numbers)."""
    import textwrap
    entries = []
    for i, plate in enumerate(plates, 1):
        stocks = ", ".join(str(p["stock"]) for p in plate)
        entries.append((i, len(plate), textwrap.wrap(stocks, width=95) or [""]))
    total = sum(len(e[2]) for e in entries)
    fig, ax = plt.subplots(figsize=(12, 1.0 + 0.34 * total + 0.16 * len(entries)))
    ax.axis("off"); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.text(0.5, 0.995, f"Seed → Plate map  ·  {mode}  ·  "
            f"{sum(e[1] for e in entries)} seeds on {len(entries)} plates",
            ha="center", va="top", fontsize=13, weight="bold")
    step = 0.94 / max(1, total + len(entries))
    y = 0.95
    for idx, (i, cnt, lines) in enumerate(entries):
        h = step * len(lines)
        ax.add_patch(plt.Rectangle((0.01, y - h), 0.98, h, zorder=0,
                     fc=("#eef3f8" if idx % 2 == 0 else "#ffffff"), ec="#d5dbdb", lw=0.5))
        ax.text(0.025, y - h / 2, f"Plate {i:02d}\n({cnt} seeds)", ha="left", va="center",
                fontsize=8.5, weight="bold", color="#1f4e78")
        for j, ln in enumerate(lines):
            ax.text(0.135, y - step * (j + 0.5), ln, ha="left", va="center",
                    fontsize=8, family="monospace")
        y -= h
    plt.savefig(path, dpi=150, bbox_inches="tight"); plt.close(fig)

# ------------------------------------------------------------------ excel
HFILL = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF")
def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c); cell.fill = HFILL; cell.font = HFONT
        cell.alignment = Alignment(horizontal="center")

def write_excel(plates, unplaceable, mode, path, excluded=()):
    wb = Workbook()
    sm = wb.active; sm.title = "Summary"
    sm.append(["PLATE", "BLOCKS", "TOTAL_CTS", "AVG_CTS", "SPAN_DIA_mm_of_90",
               "PLATE_FILL_%", "INTER_SEED_GAP_%", "ROTATED", "AVG_THICK"])
    for i, placed in enumerate(plates, 1):
        area = sum(p["w"] * p["h"] for p in placed)
        tcts = sum(p["cts"] for p in placed)
        util = 100 * area / (math.pi * R * R)
        sm.append([i, len(placed), round(tcts, 2), round(tcts / len(placed), 3),
                   round(span_mm(placed), 1), round(util, 1), round(inner_gap(placed), 1),
                   sum(1 for p in placed if p["angle"] or p.get("theta")),
                   round(sum(p["H"] for p in placed) / len(placed), 3)])
    sm.append([])
    nall = sum(len(p) for p in plates)
    tall = sum(p["cts"] for pl in plates for p in pl)
    sm.append(["TOTAL", nall, round(tall, 2), round(tall / nall, 3) if nall else "", "", "",
               "", sum(1 for pl in plates for p in pl if p["angle"] or p.get("theta")), ""])
    style_header(sm, 9)
    for col, wdt in zip("ABCDEFGHI", [8, 9, 11, 9, 18, 13, 18, 9, 11]):
        sm.column_dimensions[col].width = wdt

    for i, placed in enumerate(plates, 1):
        ws = wb.create_sheet(f"Plate_{i:02d}")
        ws.append(["STOCK_NO", "CTS", "LENGTH_1", "WIDTH", "HEIGHT",
                   "PLACED_W", "PLACED_H", "ANGLE", "CENTER_X", "CENTER_Y", "ZONE"])
        for p in placed:
            zone = "CORE" if p.get("zone") == "square" else "SIDE"
            ang = round(p["theta"], 1) if p.get("theta") is not None else p["angle"]
            cx = p["cx"] if p.get("cx") is not None else p["x"] + p["w"] / 2
            cy = p["cy"] if p.get("cy") is not None else p["y"] + p["h"] / 2
            ws.append([p["stock"], round(p["cts"], 2), p["L"], p["W"], p["H"],
                       round(p["w"], 2), round(p["h"], 2), ang,
                       round(cx, 2), round(cy, 2), zone])
        ws.append([])
        avg = sum(p["cts"] for p in placed) / len(placed)
        arow = ws.max_row + 1
        ws.append(["AVERAGE CTS", round(avg, 3)])
        ws.cell(arow, 1).font = Font(bold=True); ws.cell(arow, 2).font = Font(bold=True)
        style_header(ws, 11)
        for col, wdt in zip("ABCDEFGHIJK", [12, 7, 9, 8, 8, 9, 9, 7, 10, 10, 6]):
            ws.column_dimensions[col].width = wdt

    if unplaceable or excluded:
        ws = wb.create_sheet("Unused")
        ws.append(["STOCK_NO", "LENGTH_1", "WIDTH", "HEIGHT", "REASON"])
        for p in unplaceable:
            ws.append([p["stock"], p["L"], p["W"], p["H"], "too big for usable area"])
        for p in excluded:
            ws.append([p["stock"], p["L"], p["W"], p["H"], "user-excluded (unused)"])
        style_header(ws, 5)
        for col, wdt in zip("ABCDE", [13, 9, 8, 8, 22]):
            ws.column_dimensions[col].width = wdt
    safe_save(wb, path)

# ------------------------------------------------------------------ seed -> plate map
def write_seed_map(plates, mode, path, excluded=()):
    """Mirror the provided datasheet and add ELIGIBLE + PLATE columns, so every
    seed shows which plate it was placed on. Plus a reverse 'By_Plate' sheet.
    User-excluded seeds are tagged PLATE='EXCLUDED'."""
    excl_set = {b["stock"] for b in excluded}
    src = openpyxl.load_workbook(XLSX, data_only=True).active
    header = [c.value for c in src[1]]
    stock_plate = {}
    for i, plate in enumerate(plates, 1):
        for p in plate:
            stock_plate[p["stock"]] = i
    wb = Workbook()
    ws = wb.active; ws.title = "Datasheet"
    ws.append(list(header) + ["ELIGIBLE", "PLATE"])
    placed_n = rejected_n = 0
    for row in src.iter_rows(min_row=2, values_only=True):
        H, stock = fnum(row[6]), row[1]
        if H is None or not (T_LO <= H <= T_HI):
            estr, plate = "No (thickness)", "-"; rejected_n += 1
        elif stock in excl_set:
            estr, plate = "Yes", "EXCLUDED"
        else:
            pl = stock_plate.get(stock)
            estr, plate = "Yes", (pl if pl else "not selected")
            if pl: placed_n += 1
        ws.append(list(row) + [estr, plate])
    style_header(ws, len(header) + 2)
    for col, wdt in zip("ABCDEFGHI", [9, 13, 6, 7, 9, 8, 8, 14, 7]):
        ws.column_dimensions[col].width = wdt
    bp = wb.create_sheet("By_Plate")
    bp.append(["PLATE", "SEED_COUNT", "STOCK_NOS"])
    for i, plate in enumerate(plates, 1):
        bp.append([i, len(plate), ", ".join(str(p["stock"]) for p in plate)])
    style_header(bp, 3)
    for col, wdt in zip("ABC", [8, 11, 140]):
        bp.column_dimensions[col].width = wdt
    saved = safe_save(wb, path)
    print(f"  -> seed map: {placed_n} placed across {len(plates)} plates, "
          f"{rejected_n} rejected on thickness  ({saved or path})")

# ------------------------------------------------------------------ main
DISPATCH = {"squarecore": pack_squarecore, "grid": pack_grid, "rows": pack_rows,
            "lattice": pack_lattice, "mixed": pack_mixed, "maxfill": pack_maxfill,
            "edgefill": pack_edgefill}

def safe_save(wb, path):
    """Save a workbook; if it's locked (open in Excel), fall back to *_NEW.xlsx
    and warn, so one open file never aborts the whole run."""
    try:
        wb.save(path); return path
    except PermissionError:
        alt = path[:-5] + "_NEW.xlsx"
        try:
            wb.save(alt)
            print(f"  ! '{path}' is open in Excel -> wrote '{alt}' instead (close it to refresh)")
            return alt
        except PermissionError:
            print(f"  ! could not save '{path}' (open in Excel?) -- skipped")
            return None

def replace_excluded_inplace(plates, excluded):
    """Remove excluded seeds and fill each freed slot in place with the best-fitting
    available seed pulled from the tail plates -- every other seed stays put.
    Returns (filled, empty)."""
    excl = {b["stock"] for b in excluded}
    freed = []                                   # (plate_idx, x, y, w, h) of each removed seed
    for pi in range(len(plates)):
        keep = []
        for p in plates[pi]:
            (freed.append((pi, p["x"], p["y"], p["w"], p["h"]))
             if p["stock"] in excl else keep.append(p))
        plates[pi] = keep
    freed.sort(key=lambda s: -(s[3] * s[4]))     # fill the biggest holes first

    def fit(seed, w, h):                         # can seed sit inside a w x h slot?
        L, W = seed["L"], seed["W"]
        if L <= w + 1e-9 and W <= h + 1e-9: return (L, W, 0)
        if W <= w + 1e-9 and L <= h + 1e-9: return (W, L, 90)
        return None

    filled = empty = 0
    for pi, x, y, w, h in freed:
        best = None                              # largest-area seed from the tail that fits
        for spi in range(len(plates) - 1, -1, -1):
            if spi == pi:
                continue
            for seed in plates[spi]:
                f = fit(seed, w, h)
                if f and (best is None or seed["L"] * seed["W"] > best[2]["L"] * best[2]["W"]):
                    best = (spi, f, seed)
            if best:
                break                            # take from the last plate that has a fit
        if best:
            spi, (rw, rh, ang), seed = best
            plates[spi].remove(seed)
            ns = dict(seed); ns["x"], ns["y"], ns["w"], ns["h"], ns["angle"] = x, y, rw, rh, ang
            plates[pi].append(ns)
            filled += 1
        else:
            empty += 1
    plates[:] = [pl for pl in plates if pl]      # drop any plate emptied by the moves
    return filled, empty

def run(modes, blocks, list_seeds=False, excluded=()):
    for mode in modes:
        try:
            print(f"--- mode: {mode} ---")
            img_dir = f"images_{mode}"
            os.makedirs(img_dir, exist_ok=True)
            for f in os.listdir(img_dir):
                if f.lower().endswith(".png"):
                    try: os.remove(os.path.join(img_dir, f))
                    except PermissionError: pass
            plates, unplaceable = DISPATCH[mode](blocks) if mode in DISPATCH else pack(blocks, mode)
            if excluded:                                # in-place: keep others, fill freed slots
                nf, ne = replace_excluded_inplace(plates, excluded)
                print(f"  replaced {len(excluded)} excluded -> {nf} freed slots filled "
                      f"from spare seeds, {ne} left empty")
            for i, placed in enumerate(plates, 1):
                render(placed, i, mode, os.path.join(img_dir, f"plate_{i:02d}.png"))
            write_excel(plates, unplaceable, mode, f"arrangement_{mode}.xlsx", excluded=excluded)
            write_seed_map(plates, mode, f"seed_plate_map_{mode}.xlsx", excluded=excluded)
            render_seed_map_img(plates, mode, f"seed_plate_map_{mode}.png")
            tot = sum(len(p) for p in plates)
            rot = sum(1 for pl in plates for p in pl if p["angle"] or p.get("theta"))
            print(f"  -> {tot} placed, {len(plates)} plates, {rot} rotated, "
                  f"{len(unplaceable)} unplaceable, {len(excluded)} excluded(unused)  "
                  f"(images_{mode}\\, arrangement_{mode}.xlsx)")
            if list_seeds:                              # print which seed -> which plate
                print(f"  seed -> plate ({mode}):")
                for i, plate in enumerate(plates, 1):
                    print(f"    Plate {i:02d} ({len(plate):2d} seeds): "
                          + ", ".join(str(p["stock"]) for p in plate))
                if excluded:
                    print("    EXCLUDED (unused): "
                          + ", ".join(str(b["stock"]) for b in excluded))
                if unplaceable:
                    print("    UNPLACED: " + ", ".join(str(p["stock"]) for p in unplaceable))
        except Exception as e:
            print(f"  ! mode '{mode}' failed: {type(e).__name__}: {e} -- continuing")

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(
        description="Arrange CVD diamond seeds on a round plate (thickness-gated).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    ap.add_argument("file", nargs="?", default="BLOCK.xlsx", help="input .xlsx datasheet")
    ap.add_argument("--mode", nargs="+", default=["mixed"],
                    help="one or more (space- or comma-separated), or 'all': "
                         "mixed lattice grid maxfill rows squarecore circle square edgefill")
    ap.add_argument("--t-lo", dest="t_lo", type=float, default=0.67, help="min thickness mm")
    ap.add_argument("--t-hi", dest="t_hi", type=float, default=0.73, help="max thickness mm")
    ap.add_argument("--plate", type=float, default=90.0, help="total plate diameter mm")
    ap.add_argument("--margin", type=float, default=3.0,
                    help="required margin mm (usable diameter = plate - margin)")
    ap.add_argument("--usable", type=float, default=None,
                    help="usable diameter mm directly (overrides --margin)")
    ap.add_argument("--clearance", type=float, default=0.0, help="forced gap between seeds mm")
    ap.add_argument("--grid", type=float, default=0.5, help="placement scan step mm")
    ap.add_argument("--shape", nargs="+", default=["all"],
                    help="seed shape(s) to arrange: 'square', 'rectangle', both, or 'all' "
                         "(square = L≈W within tolerance)")
    ap.add_argument("--square-tol", dest="square_tol", type=float, default=0.05,
                    help="relative L/W tolerance to count a seed as square (0.05 = 5%%)")
    ap.add_argument("--list", action="store_true", dest="list_seeds",
                    help="print the seed -> plate mapping to the console")
    ap.add_argument("--exclude", nargs="+", default=[],
                    help="stock number(s) to mark unused and keep off the plates "
                         "(space/comma separated)")
    ap.add_argument("--exclude-file", dest="exclude_file", default=None,
                    help="text file of stock numbers to exclude (one per line, # = comment)")
    a = ap.parse_args()

    # apply CLI settings to the module globals the algorithms read
    XLSX = a.file
    T_LO, T_HI = a.t_lo, a.t_hi
    PLATE_D = a.plate
    USABLE_D = a.usable if a.usable is not None else (a.plate - a.margin)
    if USABLE_D <= 0:
        ap.error(f"usable diameter must be > 0 (plate {a.plate} - margin {a.margin})")
    R = USABLE_D / 2.0
    S2 = (USABLE_D / math.sqrt(2)) / 2.0
    INSCRIBED = USABLE_D / math.sqrt(2)
    GRID = a.grid
    CLEARANCE = a.clearance

    ALL = ("mixed", "lattice", "grid", "maxfill", "rows", "squarecore", "circle", "square",
           "edgefill")
    valid = set(ALL)
    modes = [m.strip() for tok in a.mode for m in tok.split(",") if m.strip()]
    if "all" in modes:
        modes = list(ALL)
    bad = [m for m in modes if m not in valid]
    if bad:
        ap.error(f"unknown mode(s): {bad}. choose from {sorted(valid)} (or 'all')")
    modes = tuple(modes)

    shape_in = [s.strip().lower() for tok in a.shape for s in tok.split(",") if s.strip()]
    bad_sh = [s for s in shape_in if s not in ("all", "square", "rectangle")]
    if bad_sh:
        ap.error(f"unknown shape(s): {bad_sh}. choose from square, rectangle, all")
    allowed = {"square", "rectangle"} if "all" in shape_in else set(shape_in)
    shape_lbl = "+".join(sorted(allowed))

    print(f"file={XLSX}  thickness={T_LO}-{T_HI} mm  shape={shape_lbl}  "
          f"plate=Ø{PLATE_D} mm  usable=Ø{USABLE_D} mm  margin={PLATE_D - USABLE_D:.2f} mm  "
          f"clearance={CLEARANCE} mm  grid={GRID} mm  modes={modes}")
    blocks = load_blocks(XLSX, shapes=allowed, square_tol=a.square_tol)
    nsq = sum(1 for b in blocks if b["shape"] == "square")
    print(f"eligible blocks (thickness {T_LO}-{T_HI} mm, shape={shape_lbl}): "
          f"{len(blocks)}  ({nsq} square, {len(blocks) - nsq} rectangle)")
    if not blocks:
        ap.error("no seeds match the thickness + shape filter")

    # user exclusions -> marked unused, kept off the plates
    excl_set = {s.strip() for tok in a.exclude for s in tok.split(",") if s.strip()}
    if a.exclude_file:
        with open(a.exclude_file) as fh:
            excl_set |= {ln.strip() for ln in fh if ln.strip() and not ln.startswith("#")}
    excluded = [b for b in blocks if b["stock"] in excl_set]
    miss = excl_set - {b["stock"] for b in blocks}
    if miss:
        print(f"  ! --exclude: not in eligible set (ignored): {sorted(miss)}")
    if excluded:
        print(f"excluded by user (unused): {len(excluded)} -> "
              + ", ".join(b["stock"] for b in excluded))
    # pass the full set; excluded seeds are packed then replaced in place from spares
    run(modes, blocks, list_seeds=a.list_seeds, excluded=excluded)
