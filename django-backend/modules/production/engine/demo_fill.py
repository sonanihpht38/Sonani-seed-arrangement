r"""
DEMO max-fill (saved separately).

1. Take the REAL eligible seeds (thickness + shape gated) from the datasheet.
2. Arrange them into inscribed-square plates with MaxRects best-fit + 90 deg
   rotation (checks all blocks, not in sheet order).
3. Fill every remaining gap with DEMO filler seeds (>= 2x2 mm, square/rectangular)
   using an EXACT rectilinear decomposition of the empty space (cut at the real
   seeds' own edges -> no grid-resolution loss). This drives plate-fill toward 100%.
4. Save to a SEPARATE place: images_demofill\  +  arrangement_demofill.xlsx.

Demo fillers are synthetic (NOT from the datasheet) and clearly marked.

Usage:  .\venv\Scripts\python.exe demo_fill.py [FILE] [--plate 90 --margin 3
                                       --t-lo 0.67 --t-hi 0.73 --shape all --min-seed 2]
"""
import os, glob, math, argparse
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patheffects as pe
from matplotlib.patches import Circle, Rectangle as Rect, Patch, Polygon as MplPoly
from matplotlib.collections import PatchCollection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pack_v2 as P
import parallel_pack as _parallel


def _largest_rect(free, n, mind):
    """Largest all-free axis-aligned rectangle in the n×n grid `free` with BOTH sides
    >= mind cells. Returns (i0, j0, wc, hc) or None. Histogram + stack, O(n^2)."""
    best_area = 0
    best = None
    height = [0] * n
    for j in range(n):
        row = free[j]
        for i in range(n):
            height[i] = height[i] + 1 if row[i] else 0
        stack = []                                        # (start_index, bar_height)
        i = 0
        while i <= n:
            h = height[i] if i < n else 0
            start = i
            while stack and stack[-1][1] >= h:
                idx, hh = stack.pop()
                width = i - idx
                if hh >= mind and width >= mind and hh * width > best_area:
                    best_area = hh * width
                    best = (idx, j - hh + 1, width, hh)
                start = idx
            stack.append((start, h))
            i += 1
    return best


def hybrid_fill(placed, R, min_size=2.0, step=1.0):
    """Gap fill, LARGEST-FIRST: build the free-cell grid inside the usable circle
    (cells fully inside and not covered by a real seed), then repeatedly place the
    single biggest empty rectangle as a BIG filler seed until none larger than
    min_size×min_size remains; finally drop fixed min_size square DUMMY seeds into
    whatever small/curved cells are left. Covering big areas with big seeds first
    sharply cuts the dummy count. Returns (big_fillers, small_dummies):
    big = (x, y, w, h) list, small = (x, y) lower-left corners (min_size square)."""
    n = int(round(2 * R / step))
    R2 = R * R
    mind = max(1, int(round(min_size / step)))
    free = [[False] * n for _ in range(n)]
    for j in range(n):                                    # cells fully inside the circle
        y0 = -R + j * step; y1 = y0 + step; my = max(y0 * y0, y1 * y1)
        rowf = free[j]
        for i in range(n):
            x0 = -R + i * step; x1 = x0 + step
            if max(x0 * x0, x1 * x1) + my <= R2 + 1e-9:
                rowf[i] = True
    for p in placed:                                      # subtract real-seed coverage
        i0 = max(0, int((p["x"] + R) / step) - 1); i1 = min(n, int((p["x"] + p["w"] + R) / step) + 2)
        j0 = max(0, int((p["y"] + R) / step) - 1); j1 = min(n, int((p["y"] + p["h"] + R) / step) + 2)
        for j in range(j0, j1):
            cy0 = -R + j * step; cy1 = cy0 + step
            if cy1 <= p["y"] or cy0 >= p["y"] + p["h"]:
                continue
            rowf = free[j]
            for i in range(i0, i1):
                cx0 = -R + i * step
                if cx0 + step > p["x"] and cx0 < p["x"] + p["w"]:
                    rowf[i] = False
    big = []                                              # PASS 1: largest rectangle first
    while True:
        r = _largest_rect(free, n, mind)
        if r is None:
            break
        i0, j0, wc, hc = r
        if wc <= mind and hc <= mind:
            break                                         # nothing bigger than a dummy left
        for jj in range(j0, j0 + hc):
            rowf = free[jj]
            for ii in range(i0, i0 + wc):
                rowf[ii] = False
        big.append((-R + i0 * step, -R + j0 * step, wc * step, hc * step))
    small = []                                            # PASS 2: 2×2 dummies into the rest
    d = mind
    for j in range(n - d + 1):
        for i in range(n - d + 1):
            ok = True
            for dj in range(d):
                rowf = free[j + dj]
                for di in range(d):
                    if not rowf[i + di]:
                        ok = False; break
                if not ok:
                    break
            if ok:
                small.append((-R + i * step, -R + j * step))
                for dj in range(d):
                    rowf = free[j + dj]
                    for di in range(d):
                        rowf[i + di] = False
    return big, small


def guillotine_tri_fill(real, R, min_size, sub_h=2.0, cap_h=2.0, v_inset=1.5, n_cham=1, max_bands=2):
    """MACHINE-CUT dummies shaped like the client's GREEN reference: a RECTANGLE whose curved-
    boundary corner is cut by ONE single straight CHAMFER (a clean convex pentagon), so EVERY
    corner is >= 90 deg (no acute corners, no fine staircase, no multi-segment faux-curve).
    Trick: at the WIDE end of a side gap the outer edge starts with a short VERTICAL segment
    (`v_inset` mm) so it meets the horizontal cap at 90 deg; then ONE straight diagonal
    (n_cham=1) runs to the narrow end. Applied symmetrically to the left/right side gaps.
    The TOP/BOTTOM CAPS are kept as plain axis-aligned RECTANGLES (client: caps must stay
    rectangular/square, not irregular). Returns [(ring, area, (lx,ly))]."""
    from shapely.geometry import Polygon as ShPoly

    def cx(y):
        return math.sqrt(max(0.0, R * R - y * y))

    def outer_chain(y0, y1):
        """Outer boundary points (x>0 side) from y0 up to y1 that meet horizontal caps at
        >= 90 deg: a vertical inset at any WIDE end, then chamfers following the circle."""
        eps = 0.3
        pts = []
        # bottom end
        if cx(y0) > cx(min(y0 + eps, y1)):                # bottom is the wide end -> vertical inset
            yv = min(y0 + v_inset, (y0 + y1) / 2.0)
            pts += [(cx(yv), y0), (cx(yv), yv)]
            lo = yv
        else:                                             # bottom narrow -> reach the circle
            pts.append((cx(y0), y0))
            lo = y0
        # top end
        if cx(y1) > cx(max(y1 - eps, y0)):                # top is the wide end -> vertical inset
            yv2 = max(y1 - v_inset, (y0 + y1) / 2.0)
            top = [(cx(yv2), yv2), (cx(yv2), y1)]
            hi = yv2
        else:
            top = [(cx(y1), y1)]
            hi = y1
        if hi - lo > 0.5:                                 # chamfers along the curve
            for i in range(1, n_cham):
                t = lo + (hi - lo) * i / n_cham
                pts.append((cx(t), t))
        return pts + top

    out = []

    def emit(verts):
        g = ShPoly(verts)
        if not g.is_valid:
            g = g.buffer(0)
        if g.is_empty or g.geom_type != "Polygon" or g.area < min_size * min_size * 0.6:
            return
        ring = [(round(x, 3), round(y, 3)) for x, y in g.exterior.coords]
        rp = g.representative_point()
        out.append((ring, g.area, (round(rp.x, 3), round(rp.y, 3))))

    rows = {}
    for p in real:
        rows.setdefault(round(p["y"], 1), []).append(p)
    tops, bots = [], []
    for yk, r in rows.items():
        y0 = min(p["y"] for p in r); y1 = max(p["y"] + p["h"] for p in r)
        tops.append(y1); bots.append(y0)
        xl = min(p["x"] for p in r); xr = max(p["x"] + p["w"] for p in r)
        # A TAPERED (triangular) gap wastes space under one long diagonal; split it into a FEW
        # stacked single-diagonal pentagons so each hugs the curve closer (client's "divide
        # into two"). Capped at 2-3 pieces (3 only for the sharpest top/bottom corners) to keep
        # the dummy count down; near-uniform gaps stay one piece.
        taper = abs(cx(y0) - cx(y1))
        nb = min(max_bands, 3 if taper >= 12.0 else (2 if taper >= 4.0 else 1))
        while nb > 1 and (y1 - y0) / nb < 4.0:             # no band thinner than ~4mm
            nb -= 1
        for j in range(nb):
            ya = y0 + (y1 - y0) * j / nb
            yb = y0 + (y1 - y0) * (j + 1) / nb
            chain = outer_chain(ya, yb)
            xmax = max(x for x, y in chain)
            if xmax - xr >= min_size:                      # right gap (single-diagonal pentagon)
                emit([(xr, ya)] + chain + [(xr, yb)])
            if xl + xmax >= min_size:                      # left gap (mirror in x)
                emit([(xl, yb)] + [(-x, y) for x, y in reversed(chain)] + [(xl, ya)])
    if tops:
        top, bot = max(tops), min(bots)
        # CAPS: TWO chamfered right-trapezoids per cap (left + right half — the client's two
        # mirror "green" orientations). Each has a single diagonal that follows the circle from
        # the centre apex (0, ±R) down to a short VERTICAL at the outer end (so the outer corner
        # is 90 deg, not acute). All corners >= 90 deg.
        yflat = math.sqrt(max(0.0, R * R - min_size * min_size))   # |y| of the tiny central flat top
        for sgn, base in ((1, top), (-1, bot)):
            h = R - abs(base)                             # cap height (row top -> circle apex)
            if h < min_size:
                continue
            yc = sgn * yflat                              # centre-vertical top (just below the apex)
            if abs(yc) - abs(base) < 0.5:
                continue
            xt = cx(yc)                                   # central flat-top half-width (= min_size)
            yv = base + sgn * min(v_inset, h * 0.6)       # short OUTER vertical foot
            xb = cx(yv)                                   # outer x (foot of the short vertical)
            if xb < min_size:
                continue
            # right half: bottom -> outer short-vertical -> single diagonal -> tiny flat top ->
            # centre vertical (a flat top instead of a sharp apex keeps the centre corner 90 deg)
            emit([(0.0, base), (xb, base), (xb, yv), (xt, yc), (0.0, yc)])
            emit([(0.0, base), (0.0, yc), (-xt, yc), (-xb, yv), (-xb, base)])     # left half (mirror)
    return out


def pool_thin_rows(real, R, min_size):
    """For the MACHINE-CUT fill: a WIDE row centered at y~0 leaves a thin sliver on BOTH sides
    (each < min_size) that can't be filled with a valid dummy. Shift such a row so its slack
    POOLS to one side -> the seed edge sits flush against the circle on the other side (no gap)
    and the pooled side becomes ONE fillable >= min_size gap. Only rows whose BOTH centered
    side-gaps are < min_size are touched (narrow rows keep their symmetric gaps). Returns a
    shifted DEEP COPY (the originals are shared with the other fill sections)."""
    import copy
    real = copy.deepcopy(real)

    def cx(y):
        return math.sqrt(max(0.0, R * R - y * y))

    rows = {}
    for p in real:
        rows.setdefault(round(p["y"], 1), []).append(p)
    for r in rows.values():
        y0 = min(p["y"] for p in r); y1 = max(p["y"] + p["h"] for p in r)
        xl = min(p["x"] for p in r); xr = max(p["x"] + p["w"] for p in r)
        cxn = min(cx(y0), cx(y1))                         # narrowest chord across the row
        lg, rg = xl + cxn, cxn - xr                       # left / right centered gaps
        if lg < min_size and rg < min_size and lg + rg >= min_size:
            for p in r:                                   # make left flush, pool slack to the right
                p["x"] -= lg
    return real


def guillo_plate_job(args):
    """Picklable per-plate worker for the MACHINE-CUT section: real seeds (rows) + TRIANGLE
    dummies from `guillotine_tri_fill` — recursive-guillotine (edge-to-edge of each freed
    piece) triangles reaching the circle. `args` = (real, pi, plate_d, R, min_size, path)."""
    real, pi, plate_d, R, min_size, path = args
    P.PLATE_D = plate_d
    # DUMMY floor (1.5mm) is lower than the REAL-seed floor (min_size, 2mm): the extra 0.5mm
    # lets dummies fill the thin edge slivers beside the widest rows so real-seed CORNERS are
    # backed by dummy material instead of sitting next to a ~2mm empty pocket (chip-risk fix).
    # Real seeds are untouched — only filler shrinks. Also reduces pooling: a 1.5-2mm gap is
    # now fillable on BOTH sides (symmetric support) instead of pooled flush to one side.
    dfloor = min(min_size, 1.5)
    real = pool_thin_rows(real, R, dfloor)                # pool only truly sub-1.5mm wide-row slivers
    placed = list(real)
    # Each row SIDE gap is ONE solid dummy whose outer edge follows the circle with a few
    # straight chords (sub_h spacing) -> no thin sub-strip slivers (old D12 beside a seed
    # corner) AND no flat-trapezoid coverage loss. Caps stay finely stripped (cap_h=4).
    for k, (coords, area, rp) in enumerate(guillotine_tri_fill(real, R, dfloor, sub_h=2.0, cap_h=2.0), 1):
        xs = [c[0] for c in coords]; ys = [c[1] for c in coords]
        bx, by = min(xs), min(ys); bw, bh = max(xs) - bx, max(ys) - by
        placed.append({"stock": f"FILL-{pi:02d}-I{k}", "cts": 0.0, "L": round(bw, 1), "W": round(bh, 1),
                       "H": 0.0, "x": bx, "y": by, "w": bw, "h": bh, "angle": 0, "filler": True,
                       "kind": "irregular", "poly": coords, "area": area, "lx": rp[0], "ly": rp[1]})
    fillH = (round(sum(p["H"] for p in real) / len(real), 3) if real
             else round((P.T_LO + P.T_HI) / 2.0, 3))
    for p in placed:
        if p.get("filler"):
            p["H"] = fillH
    covered = sum(p.get("area", p["w"] * p["h"]) for p in placed)
    fill = 100 * covered / (math.pi * R * R)
    m = 0.0
    for p in placed:
        if p.get("kind") == "irregular":
            for cx, cy in p["poly"]:
                m = max(m, math.hypot(cx, cy))
            continue
        for cx, cy in ((p["x"], p["y"]), (p["x"] + p["w"], p["y"]),
                       (p["x"] + p["w"], p["y"] + p["h"]), (p["x"], p["y"] + p["h"])):
            m = max(m, math.hypot(cx, cy))
    span = 2 * m
    render_cross_circle(placed, real, pi, R, fill, path)
    return (pi, placed, round(fill, 4), round(span, 4), 0.0)


def cw_degrees(deg):
    """Stored rotation -> the CLOCKWISE angle the floor should actually turn.

    Every pose in this engine is produced by `shapely.affinity.rotate`, which
    turns COUNTER-clockwise for a positive angle. The shop floor works
    clockwise, so a stored 90 means "turn it 270 degrees clockwise" — printing
    the stored number on the plate would send the operator the wrong way round
    and land a cut corner facing the wrong side.

    Display only. Nothing in the packing reads this.
    """
    try:
        d = float(deg or 0.0)
    except (TypeError, ValueError):
        return 0
    return int(round((-d) % 360.0)) % 360


def display_turn(p):
    """The CLOCKWISE turn to show for one placed seed.

    A CUT seed has four distinct orientations and every one of them matters —
    which way the ground corner points is the whole question — so its angle is
    reported as-is.

    A plain rectangle does not: turning it 180 degrees leaves the identical
    footprint, so 270 and 90 are the same instruction. Reporting 270 there is
    not more precise, it is more confusing — the operator does a three-quarter
    turn to reach a position a quarter turn would have given. Rectangles are
    therefore folded onto 0 or 90.
    """
    cw = cw_degrees(p.get("angle"))
    return cw if p.get("irregular") else cw % 180


def band_caption():
    """One line naming the THICKNESS and SEED-WIDTH bands the run used.

    Read off the engine globals set by engine_runner._apply_globals, so the
    caption always describes the run that produced the image rather than
    whatever the reader assumes. A width band is optional at both ends, so it is
    reported as a range, a bound, or "any".
    """
    lo = getattr(P, "W_LO", None)
    hi = getattr(P, "W_HI", None)
    if lo is None and hi is None:
        width = "any"
    elif lo is not None and hi is not None:
        width = f"{lo:g}–{hi:g} mm"
    elif lo is not None:
        width = f"≥ {lo:g} mm"
    else:
        width = f"≤ {hi:g} mm"
    return (f"thickness {getattr(P, 'T_LO', 0):g}–{getattr(P, 'T_HI', 0):g} mm"
            f"  ·  seed width {width}")


# Smallest seat, in mm2, that a rotation label is drawn inside. Below this the
# text is wider than the stone and overlaps its neighbours, which makes the plate
# harder to read rather than easier — those seeds carry their angle in the seed
# list beside the plate instead, where every seed appears whatever its size.
ANGLE_LABEL_MIN_MM2 = 45.0


def _draw_plate_numbers(ax, items, angles=None):
    """Label each seat with its NUMBER — bold with a dark halo so it stays
    readable in any seat size / fill colour. items = list of
    (label, lx, ly, area, color).

    `angles`, when given, is a parallel list of CLOCKWISE degrees; a seat big
    enough to hold it also gets its turn printed under the number, so the floor
    can read the intended orientation straight off the plate.
    """
    halo = [pe.withStroke(linewidth=2.6, foreground="#0d0d0dee")]
    for i, (label, lx, ly, area, color) in enumerate(items):
        fs = min(14.0, max(8.0, math.sqrt(max(area, 1.0)) * 0.7))
        cw = angles[i] if angles is not None and i < len(angles) else None
        if cw is None or area < ANGLE_LABEL_MIN_MM2:
            ax.text(lx, ly, label, ha="center", va="center", fontsize=fs,
                    color=color, fontweight="bold", zorder=3, path_effects=halo)
            continue
        # Number above, turn below, so neither has to shrink to fit the other.
        ax.text(lx, ly + fs * 0.030, label, ha="center", va="bottom", fontsize=fs,
                color=color, fontweight="bold", zorder=3, path_effects=halo)
        ax.text(lx, ly - fs * 0.030, f"↻{cw}°", ha="center", va="top",
                fontsize=max(5.0, fs * 0.62), color=color, fontweight="bold",
                zorder=3, path_effects=halo)


# SEED LIST GEOMETRY. The list used to be one column however long it was, and a
# column cannot grow past the height of the page: a Ø158 plate carries 138 seeds
# into a 9.6 inch panel, which is a 0.07 inch row pitch under a 5.5 pt font, so
# the rows printed on top of each other and the list was unreadable. Nothing
# about that is specific to Ø158 — it is purely the seed COUNT, so any plate
# reaches it once enough seeds fit, and a bigger plate reaches it sooner.
#
# The row pitch is therefore fixed and the list flows into as many columns as it
# needs, with the panel widened to hold them. Legibility then does not depend on
# the plate diameter, the seed count, or the band.
LEGEND_MIN_ROW_IN = 0.16    # inches — row pitch that keeps ~8 pt text clear
# Width one column needs. A seed row runs about 40 characters — "DOMI001169
# 14.35x10.35  H 0.73  turn 0deg" — and DejaVu Sans averages ~0.6 em per
# character, so at the ~7 pt this pitch yields the text alone wants 2.3 inches,
# plus the swatch and the number field. 2.8 was measured too tight: the trailing
# angle of one column printed over the next column's number.
LEGEND_COL_IN = 3.4
LEGEND_MIN_PANEL_IN = 5.0   # inches — never narrower than the original panel


def _legend_shape(n, panel_h_in):
    """(columns, rows per column) for `n` seed-list rows in a panel that tall."""
    n = max(int(n), 1)
    per_col = max(1, int(panel_h_in / LEGEND_MIN_ROW_IN))
    ncols = max(1, int(math.ceil(n / float(per_col))))
    # Re-balance so the last column is not a stub — 138 over 2 columns is 69
    # each, not 71 and 67.
    return ncols, int(math.ceil(n / float(ncols)))


def _legend_panel_in(n, panel_h_in):
    """How wide the seed-list panel has to be to hold `n` rows legibly."""
    ncols, _ = _legend_shape(n, panel_h_in)
    return max(LEGEND_MIN_PANEL_IN, ncols * LEGEND_COL_IN)


def _draw_legend_list(axl, title, subtitle, entries):
    """Draw the seed list beside the plate. entries = list of
    (swatch_color, swatch_edge, number_text, description_text, text_color).

    Flows into columns rather than growing one column past the page — see
    LEGEND_MIN_ROW_IN. The caller sizes the panel with _legend_panel_in() so the
    columns have somewhere to go.
    """
    axl.axis("off"); axl.set_xlim(0, 1); axl.set_ylim(0, 1)
    axl.text(0.0, 0.995, title, fontsize=12, fontweight="bold", va="top")
    top = 0.955
    if subtitle:
        axl.text(0.0, 0.965, subtitle, fontsize=8.5, va="top", color="#7a3d00")
        # Give the list back one line's worth of room per EXTRA subtitle line,
        # or a two-line subtitle prints straight over the first few seeds.
        top = 0.925 - 0.022 * subtitle.count("\n")
    n = max(len(entries), 1)
    bot = 0.005
    panel_h = float(axl.figure.get_size_inches()[1])
    ncols, per_col = _legend_shape(n, panel_h * (top - bot))
    rh = (top - bot) / per_col
    # Size the text from the pitch it actually has, not from the seed count. The
    # old `190 / n` bottomed out at its floor on any long list and then said
    # nothing about whether the rows had room.
    lfs = min(8.6, max(5.5, rh * panel_h * 72.0 * 0.55))
    colw = 1.0 / ncols
    for i, (color, edge, num, text, tcolor) in enumerate(entries):
        c, r = divmod(i, per_col)
        x0 = c * colw
        y = top - (r + 0.5) * rh
        axl.add_patch(Rect((x0 + 0.005 * colw / 1.0, y - rh * 0.34), 0.045 * colw,
                      rh * 0.68, facecolor=color, edgecolor=edge, lw=1.0,
                      transform=axl.transAxes, clip_on=False))
        # The number field has to clear a THREE digit seat number: a Ø158 plate
        # numbers past 100, and at 0.135 "150." ran into the stock code.
        axl.text(x0 + 0.060 * colw, y, num, fontsize=lfs, va="center", fontweight="bold")
        axl.text(x0 + 0.165 * colw, y, text, fontsize=lfs, va="center", color=tcolor)


def render_enhanced_circle(placed, real, pi, R, fill, path):
    """Render a MAX COVERAGE plate: every seed placed WHOLE at its measured shape.

    The engine never cuts a seed, so nothing here is trimmed. The drawing shows
    the two settings that decide the layout, because a plate is hard to judge
    without them: the MARGIN (the ring kept clear around the plate, drawn shaded
    between the plate edge and the dashed usable circle) and the DISTANCE BETWEEN
    SEEDS (the gap held between neighbours).
    """
    PLATE = P.PLATE_D
    margin = max(0.0, PLATE / 2.0 - R)
    seed_gap = max(0.0, float(getattr(P, "CLEARANCE", 0.0) or 0.0))
    nr = len(placed)
    cmap = plt.cm.viridis
    facecolors = [cmap(i / max(1, nr - 1)) for i in range(nr)]

    # Two panels: the plate (each seed labelled with a NUMBER only, so it stays readable in
    # any seat) and, beside it, a full seed LIST — a colour swatch matching the seat plus the
    # stock, size, thickness, and (for trimmed seeds) how much was cut off.
    # Widen the panel for the seed list rather than squeezing the list into a
    # fixed one — see LEGEND_MIN_ROW_IN. The plate panel keeps its 9.6 inches,
    # so the plate itself renders exactly as before at every diameter.
    _lw = _legend_panel_in(len(placed), 9.6 * 0.92)
    fig = plt.figure(figsize=(9.6 + _lw, 9.6))
    gs = fig.add_gridspec(1, 2, width_ratios=[9.6, _lw], wspace=0.02)
    ax = fig.add_subplot(gs[0, 0])
    axl = fig.add_subplot(gs[0, 1]); axl.axis("off")

    ax.add_patch(Circle((0, 0), PLATE / 2, fc="#e9e9ec", ec="#888", lw=1.5, zorder=0))
    # Shade the MARGIN ring itself, so the clear band reads as a deliberate
    # setting rather than space the packer failed to use.
    if margin > 0.01:
        ax.add_patch(Circle((0, 0), PLATE / 2, fc="#f6dcd7", ec="none", zorder=0.4))
        ax.add_patch(Circle((0, 0), R, fc="#e9e9ec", ec="none", zorder=0.5))
    ax.add_patch(Circle((0, 0), R, fc="none", ec="#c0392b", lw=1.2, ls="--", zorder=1))
    if margin > 0.01:
        # Call the ring out once, with a leader from the plate edge to the
        # usable circle at the 45° diagonal, where seeds rarely reach.
        d = 0.7071
        ax.annotate(
            f"margin {margin:g} mm",
            xy=(-(R + margin * 0.5) * d, (R + margin * 0.5) * d),
            xytext=(-(PLATE / 2 + 1) * d - 12, (PLATE / 2 + 1) * d + 6),
            fontsize=9, color="#a03a2c", ha="center", va="bottom", zorder=6,
            arrowprops=dict(arrowstyle="-", color="#a03a2c", lw=0.9,
                            shrinkA=0, shrinkB=1))
    if placed:                                          # all real seeds as polygons (full or clipped)
        # ZERO-GAP look: paint each whole seed's edge in its OWN fill colour so the thin
        # anti-aliasing seam matplotlib leaves between abutting patches is covered — no
        # white/dark line between seeds. TRIMMED seeds keep an ORANGE outline.
        edgecolors = ["#e67e22" if p.get("clipped") else facecolors[i] for i, p in enumerate(placed)]
        linewidths = [1.4 if p.get("clipped") else 1.0 for p in placed]
        ax.add_collection(PatchCollection([MplPoly(p["poly"], closed=True) for p in placed],
                          facecolor=facecolors, edgecolors=edgecolors, linewidths=linewidths, zorder=2))
    # NUMBER each seat (1..N) — a number fits any seat size — plus the CLOCKWISE
    # turn for seats large enough to carry it. Reference only: the angle is read
    # back from the placement the packer already chose, never used to compute it.
    _cw = [display_turn(p) for p in placed]
    _draw_plate_numbers(ax, [(str(i + 1), p.get("lx"), p.get("ly"),
                              p.get("area", p["w"] * p["h"]), "white") for i, p in enumerate(placed)],
                        angles=_cw)
    lim = PLATE / 2 + 3
    ax.set_xlim(-lim, lim); ax.set_ylim(-lim, lim); ax.set_aspect("equal"); ax.axis("off")
    circle_area = math.pi * R * R
    covered = fill / 100.0 * circle_area
    # State the settings that shaped this plate. "edge cuts ≥90°" used to sit here
    # and is gone: the engine no longer cuts anything, so it described behaviour
    # that no longer exists.
    ax.set_title(
        f"Max Coverage · Plate {pi:02d} · {covered:.0f} of {circle_area:.0f} mm² covered "
        f"({fill:.1f}%)\n"
        f"plate Ø{PLATE:g} · margin {margin:g} mm → usable Ø{2 * R:g} · "
        f"distance between seeds {seed_gap:g} mm\n"
        # Which BANDS produced this plate. Without them two plates of the same
        # inventory are indistinguishable on paper, and the band is now the main
        # control the operator sets.
        f"{band_caption()}",
        fontsize=10.5)

    # ---- seed list on the right ----
    entries = []
    trimmed = []
    for i, p in enumerate(placed):
        cut, tcolor = "", "#111"
        # How much of the STONE was lost — not how much the plate edge took off
        # its seat. A stone smaller than its clipped seat still fits whole, so it
        # is not trimmed at all. See engine_runner.seed_cut for the reasoning.
        rem, pct = _seed_cut(p)
        if rem > 0.005:
            trimmed.append(i)
            cut, tcolor = f"   ✂ cut {rem:.0f} mm² ({pct:.0f}%)", "#c0392b"
        # Size = the seed's REAL stone measurement from the import (rawL × rawW), so
        # each stock ID reports its true size rather than the cut seat it tiles into.
        # Falls back to the seat box for any seed without stored raw dims (e.g. fillers).
        dw = p.get("rawL") if p.get("rawL") is not None else p["w"]
        dh = p.get("rawW") if p.get("rawW") is not None else p["h"]
        # Exact stored measurement (decimal(18,2)) — no round-off, trailing zeros trimmed.
        sdw = f"{float(dw):.2f}".rstrip("0").rstrip(".")
        sdh = f"{float(dh):.2f}".rstrip("0").rstrip(".")
        # Every seed's CLOCKWISE turn, including the small rim stones whose seat
        # is too tight to carry the label on the plate itself.
        entries.append((facecolors[i], "#e67e22" if i in trimmed else "#555", f"{i + 1}.",
                        f"{p['stock']}   {sdw}×{sdh}   H {p['H']:.2f}   ↻{_cw[i]}°{cut}", tcolor))
    # The old subtitle counted trimmed seeds and is now permanently zero, so it
    # said nothing. Report the settings the reader actually needs instead.
    _draw_legend_list(
        axl, f"Seeds on this plate ({nr})",
        ("all placed whole (nothing cut) · "
         + (f"{seed_gap:g} mm between seeds" if seed_gap > 0 else "seeds touching")
         # Say which way ↻ means, once, where the angles are listed. A number
         # with no stated direction is worse than none: half the time it is read
         # the wrong way round.
         + "\n↻ = turn CLOCKWISE from the seed as measured"),
        entries)

    fig.savefig(path, dpi=125, bbox_inches="tight")
    plt.close(fig)


def _seed_cut(p):
    """(cut_mm2, cut_pct) for one placed stone — how much of the STONE was lost.

    Thin wrapper so the plate image and the seed-list table share ONE definition;
    they previously each computed it, and both measured the seat instead of the
    stone. Falls back to "nothing cut" if the Django package is not importable,
    so this module still runs standalone.
    """
    try:
        from modules.production.engine_runner import seed_cut
    except ImportError:
        return 0.0, 0.0
    return seed_cut(p)


def _cut_direction(g):
    """Unit vector from a seed's centre toward the corner ground off it, or None
    for a plain square/rectangle.

    The missing corner is whatever the outline lacks compared with its own
    bounding box, so this works for any pre-cut shape without being told which
    corner was taken.
    """
    from shapely.geometry import box

    bx0, by0, bx1, by1 = g.bounds
    missing = box(bx0, by0, bx1, by1).difference(g)
    if missing.is_empty or missing.area < 0.01:
        return None
    c, m = g.centroid, missing.centroid
    dx, dy = m.x - c.x, m.y - c.y
    n = math.hypot(dx, dy)
    return (dx / n, dy / n) if n > 1e-9 else None


def _cut_on_rim(g, R):
    """Is this placed cut/trimmed stone actually ON the rim, facing OUT?

    ONE test, used everywhere a cut stone can be seated, so no path can put one
    inland. A row REBUILD is such a path and was missing it: swap_row and
    _reconsider_row lift a whole row and re-lay it, cut stones included, and they
    only re-checked the notch. A cross that had been seated correctly against the
    curve could therefore be put back 17.30 mm inland on a Ø100 plate — the notch
    test passes there, because a notch facing open plate is not clipped by the
    boundary, so nothing caught it.

    THREE halves matter, and the third was missing until a real plate showed it:

      * the stone has to be within RIM_FLUSH of the boundary;
      * the corner it is missing must not open into the plate by more than
        CUT_NOTCH_MAX;
      * the corner must POINT OUTWARD, away from the plate centre.

    The direction test is not a refinement of the notch test — a small ground
    corner turned inward passes the notch test outright. The corner triangles in
    stock run a median 2.22 mm2 against a CUT_NOTCH_MAX of 3.0, so most cut
    stones could sit at the rim with the cross facing IN and nothing objected.
    The main scan in _pack_once has always tested the direction inline
    (`out > 0.0`), but the three paths that seat a stone WITHOUT going through
    that scan — _fill_row_ends, _fill_remaining_space and the row rebuilds — all
    came here instead, and this function did not. That is how two inward-facing
    crosses reached a finished Ø90 plate.

    A cross pointing inward is not a cosmetic defect: the missing corner opens a
    wedge against a flat neighbour that no stone can enter, and the shop floor
    rejects the seat. Outward it follows the curve the stone was ground for and
    costs nothing.
    """
    from shapely.geometry import Point

    gap = R - max(math.hypot(px, py) for px, py in g.exterior.coords)
    if gap > RIM_FLUSH:
        return False
    # `g` is already in plate coordinates, so its cut direction needs no further
    # rotation — score it as placed (deg = 0).
    cut = _cut_direction(g)
    if cut is not None and _outward_score(g, cut, 0) < CUT_OUTWARD_MIN:
        return False
    return _wasted_area(g, Point(0.0, 0.0).buffer(R, resolution=180)) <= CUT_NOTCH_MAX


def _wasted_area(g, disc):
    """Plate area lost to this seed's ground-off corner, in mm2.

    Only the part of the missing corner that lies INSIDE the plate counts. A cut
    seed set against the rim loses almost nothing, because the corner it lacks is
    beyond the boundary in any case; the same seed in the middle of the plate
    loses the whole triangle. Zero for a plain square or rectangle.
    """
    from shapely.geometry import box

    bx0, by0, bx1, by1 = g.bounds
    missing = box(bx0, by0, bx1, by1).difference(g)
    if missing.is_empty or missing.area < 0.01:
        return 0.0
    return missing.intersection(disc).area


def _outward_score(g, base_cut, deg):
    """How well a placed seed's cut faces AWAY from the plate centre: +1 straight
    out, -1 straight in, 0 for a seed with no cut.

    A seed is ground down because the plate is round — the cut exists to sit
    against the curve. Turned inward it does the opposite of its purpose and
    opens a wedge against its neighbour; turned outward it follows the rim and
    costs nothing. Scoring every orientation and keeping the outward-most is what
    puts a right-cut seed on the right of the plate.
    """
    if not base_cut:
        return 0.0
    a = math.radians(deg)
    ca, sa = math.cos(a), math.sin(a)
    dx = base_cut[0] * ca - base_cut[1] * sa
    dy = base_cut[0] * sa + base_cut[1] * ca
    c = g.centroid
    r = math.hypot(c.x, c.y)
    if r < 1e-6:
        return 0.0
    return (dx * c.x + dy * c.y) / r


def _seed_footprint(s):
    """The seed's REAL outline, moved so its bottom-left corner sits at (0, 0).

    A seed measured as a cut-corner outline uses that outline; a plain square or
    rectangle uses its L x W box. Either way this is the footprint the stone
    actually has — the packer never invents a cut.
    """
    from shapely import affinity
    from shapely.geometry import Polygon as ShPoly

    poly = s.get("poly")
    if poly:
        g = ShPoly(poly)
        if not g.is_valid:
            g = g.buffer(0)
        if not g.is_empty and g.area > 0 and g.geom_type == "Polygon":
            x0, y0 = g.bounds[0], g.bounds[1]
            return affinity.translate(g, -x0, -y0)
    w = float(s.get("L") or s.get("w") or 0.0)
    h = float(s.get("W") or s.get("h") or 0.0)
    if w <= 0 or h <= 0:
        return None
    return ShPoly([(0.0, 0.0), (w, 0.0), (w, h), (0.0, h)])


# Rotation is allowed for every seed — squares, rectangles and cut/trimmed alike.
# Turning a stone on the plate is a real operation; MIRRORING one is not, and
# affinity.rotate() cannot produce a mirror, so every orientation reached here is
# physically buildable.
#
# The rule from the shop floor is "clockwise only". That does NOT narrow this
# tuple: a 90 degree clockwise turn is a 270 degree anticlockwise one, so turning
# only clockwise still reaches all four orientations. Direction of travel does not
# constrain where a stone ends up — only which final orientations are permitted
# does, and all four are.
#
# 180 is tried straight after 0 on purpose: a cut corner and its neighbour's cut
# corner only nest when one of them is turned to face the other.
#
# HISTORY, so this is not flip-flopped again without cause. This was briefly
# (0,) after a real plate showed 7 of 11 cut seeds facing left when every real
# stone faced right. Locking to 0 hid that, and happened to pack better too:
#
#   rotation allowed   3 plates   84.0 / 67.7 / 11.4 %   run avg 54.4%
#   no rotation        2 plates   83.9 / 79.1 %          run avg 81.5%
#
# But 0-only treated the symptom. The cause is that the sheet cannot say which
# side the cut is on, so the outline is rebuilt from the L1/W2/L3/W4 ORDER alone
# and a mirrored reading is indistinguishable from a correct one. The datasheet
# is gaining an explicit cut-side column; once orientation is stated rather than
# inferred, rotation is safe and the packer keeps the freedom it needs.
#
# UNTIL that column is imported, cut seeds can still be turned to face the wrong
# way — the run-average figures above are what that costs.
ENHANCED_ANGLES = (0, 180, 90, 270)
NEST_STEP = 0.25          # mm — how finely a seed is slid left to close a gap
# How far to step up when a row height fits nothing. This is a SEARCH step, not
# a spacing: whatever it skips past becomes a horizontal band of bare plate
# between two rows, because the next row simply opens higher up.
#
# It was 1.0 mm, and that is what put a visible gap under the bottom row and
# above the top one — 0.94 mm and 1.04 mm on a Ø90 plate, while every row in the
# middle sat flush. Only the first and last rows show it, because those are the
# heights where the narrowing chord starts refusing rows. Measured on the
# 95-stone pool: 1.00 mm leaves gaps of 0.94/1.04 at 83.09%, 0.50 mm leaves one
# of 0.44 at 84.23%, 0.25 mm leaves NONE at 84.57%. The finer step both closes
# the band and finds a height where a row actually fits. It costs about 10% of
# the packing time and changes nothing on a pool whose rows never fail.
ROW_PROBE = 0.25          # mm — how far to step up when a row height fits nothing
RIM_EPS = 0.05            # mm — start this far inside the chord, so a corner
                          # lying exactly on the circle is not rejected (the disc
                          # is a 180-gon, marginally inside the true circle)
ROW_TOL = 0.10            # mm — how much taller than its row a seed may stand
SWEEP_ANCHORS = 48        # max pocket corners the sweep-up pass probes

# UNIFORM ROWS — a row is built from ONE height class and nothing else.
#
# ROW_TOL alone only stopped a seed standing TALLER than its row. A shorter one
# was always admitted, and it leaves a strip of bare plate above itself that no
# later row can reach. Two things are wrong with that: the strip is waste, and a
# row whose stones are not the same height cannot be built on the floor — which
# makes the plate unacceptable however good its coverage looks.
#
# ROW_LEVEL_TOL is the mirror of ROW_TOL: how much SHORTER than its row a seed
# may be. Together they hold a row inside a 0.20 mm band.
#
# The band is NOT slack for measurement error — the stones are measured to
# 0.01 mm and that is exactly the difficulty. 163 stones in stock carry 102
# DISTINCT short-side heights and 66 of them are unique, so a row can only be
# built at all by admitting stones whose true heights differ. Taken to zero the
# plate collapses: 11 seats and 20.68% on the 163-stone pool, 2 seats and 3.37%
# on the 95-stone one, because no two stones match exactly.
#
# 0.10/0.10 was measured against the 0.15/0.10 it replaces and is better on both
# pools at once — tighter rows AND more plate covered:
#
#   band          163-stone pool             95-stone pool
#   0.15 / 0.10   43 seats 86.44%, step 0.23  50 seats 82.40%, step 0.22
#   0.10 / 0.10   45 seats 86.58%, step 0.16  49 seats 83.09%, step 0.19
#
# Tightening further does cost: 0.08/0.06 gives a 0.13-0.14 mm step for 83.71%
# and 80.11%, and 0.02/0.02 drops to 76.66% / 51.43%.
#
# This only works because the height class is chosen from what is actually in
# stock — see pick_level() — so the row is opened at a height the inventory can
# finish. Choosing it from the first stone that happened to fit would strand
# every row the moment its class ran out.
ROW_LEVEL_TOL = 0.10      # mm — how much shorter than its row a seed may be
LEVEL_BAND = 0.10         # mm — width of a height class in the stock census
LEVEL_MARGIN = 1.0        # how much more stock than chord a class needs to qualify
# How many height classes are actually laid out per row before one is chosen.
#
# A row's height decides how WIDE it can be, because the plate curves in: a row
# 9.64 mm tall reaches a 14.60 mm chord, the same row at 7.37 mm reaches 20.07 mm.
# So a class of big stones can win a row on the stones it seats and still cost
# the plate, by making that row narrower than a shorter class would have.
#
# At 3 the search never reached the shorter classes for the top row: it built it
# from 12.7x9.58 and 12.88x9.64 stones and stopped 3 wide. Measured per plate:
#
#   tries   95-stone pool           163-stone pool
#     3     49 seats 84.98%  30 s   45 seats 86.58%   54 s
#     5     50 seats 85.44%  33 s   45 seats 87.25%   68 s
#     8     52 seats 86.52%  45 s   45 seats 87.25%   88 s
#    99     51 seats 85.93%  70 s   45 seats 85.44%  149 s
#
# 8 is the peak on both pools, not merely a budget: trying EVERY class is worse
# than trying eight, because each row is chosen greedily and a class spent on one
# row is gone from the next. More search is not monotonically better here.
LEVEL_TRIES = 8

# Who wins a seat when a cut stone and a plain one both fit. BOTH are tried on
# every plate and the better result is kept — see enhanced_plate_job.
#
#   "cut-first"   — a cut stone takes the rim seat that opens a row, so the cut
#                   pile drains a few per plate. Packs more stones in, but spends
#                   crosses in the middle of the plate where they leave a notch.
#   "plain-first" — a plain stone wins any tie. Fewer stones, far less waste.
#
# Measured on the reference plate, and the reason neither is hard-coded:
#
#   cut-first     42 seats  84.90%  11 cut used  10 unbuildable seats  24.6 mm2
#   plain-first   39 seats  84.37%   4 cut used   2 unbuildable seats   2.3 mm2
#
# Coverage alone would pick cut-first; buildable area picks plain-first. Which
# one wins depends on the mix on the day, so the packer runs both.
CUT_POLICIES = ("cut-first", "plain-first")

# How each row is laid. CENTRE-OUT only: rows start at the middle of the plate
# and grow both ways, so what is left over is shared between the two rims.
#
# The old "left" mode opened every row at the left chord and filled across,
# which pushed all the slack to one side — measured on the Ø80 reference plate,
# 27.99 mm of clear rim on the left against 30.06 mm on the right. Centre-out
# balances that to 29.13 against 29.09, and a symmetric plate is what the shop
# floor wants to work from.
#
# It is also better on the numbers, once the whole run is counted:
#
#   centre   plate 1: 40 seats 84.63%  7 cut  3 unbuildable | run: 35 cut, 27 bad
#   left     plate 1: 38 seats 83.84%  1 cut  0 unbuildable | run: 35 cut, 28 bad
#
# Left-to-right keeps plate 1 spotless only by pushing the cut stones onto plate
# 2, where they cost more. Over three plates the two place the same 35 cut stones
# and centre-out ends one unbuildable seat ahead.
#
# "left" is still implemented and can be put back here as a candidate; the packer
# tries every entry in this tuple and keeps the best-scoring plate.
FILL_DIRECTIONS = ("centre",)

# Where the first row's baseline sits, as an offset above the bottom of the
# plate. Rows stack from there, so this is the PHASE of the row grid against the
# circle: it decides which chords every row lands on, and a chord 1 mm higher can
# be several mm wider. Packing used to start hard at -R, which is an arbitrary
# phase, not a good one.
#
# On the Ø80 reference plate the phase alone moved coverage across a 7.7-point
# band — 76.3% at its worst, 84.0% at its best, with the shipped -R landing at
# 78.7%. It is far too big a lever to leave to chance, and the best phase depends
# on the seed mix, so it cannot simply be hard-coded either. enhanced_plate_job
# therefore packs the plate once per phase and keeps the best result.
#
# Coarse sweep first, then a refine step either side of the winner. Rows run
# about 9 mm tall, so phases beyond that repeat what a lower one already tried.
#
# The step is COARSE on purpose. Every phase is packed once per cut policy and
# once per fill direction, so the phase count multiplies by four. With real cut
# outlines a single pack costs about 30 s — polygon-against-polygon tests are far
# dearer than box-against-box — and a seven-phase sweep took 946 s a plate. Four
# phases plus the refine step below covers the same 0-6 mm range and brings that
# back to roughly a third.
ROW_PHASES = (0.0, 2.0, 4.0, 6.0)
ROW_PHASE_REFINE = 1.0

# How a seat picks between stones that fit it equally well.
#
#   "compact"  the original: closest fit, cut direction breaks the tie.
#   "area"     among equally good fits, take the stone that COVERS MORE.
#
# Neither wins everywhere, which is why both are swept and the better result is
# kept. On a pool of one size "compact" leads (84.59% vs 82.33% on 74 stones);
# on a mixed pool "area" leads (83.77% vs 82.83% on 163). The mixed case is the
# one that matters in practice, and it also fixes a wrong behaviour: under
# "compact" alone, ADDING stock lowered the plate — 84.59% fell to 82.83% when
# 89 narrow stones joined 74 wide ones, because the greedy spent seats on narrow
# stones that fit the seat but covered less. "area" reverses that (+1.44).
#
# Two other candidates were measured and rejected: penalising a placement that
# leaves a gap too narrow for any stone ("fit", 81.99%) and combining it with
# area (81.93%). Both LOWERED coverage — avoiding a dead remainder costs more
# mid-row than the rim scrap it saves.
SEAT_SCORES = ("compact", "area")

# How close to the plate boundary a cut/trimmed stone must sit to count as being
# ON THE RIM, in mm of clear space left outside it.
#
# CUT/TRIMMED STONES GO ON THE RIM AND NOWHERE ELSE. A ground corner only works
# where the plate curves away from it; anywhere else the missing corner is an
# open notch against a flat neighbour and the seat cannot be built.
#
# "Last stone in its row" is NOT the same thing, and that is what this used to
# test (_rem < RIM_SEAT, i.e. up to 8 mm of bare plate could sit outside a
# cross). It let a cut stone take the end of a row while still 6 mm short of the
# boundary — inside the plate by any reading, with the gap outside it. Refusing
# that seat leaves it to a plain square or rectangle, which is what belongs
# there, and the cut stone is only used where it genuinely follows the curve.
#
# 2.0 mm is set from the stones themselves, not picked round. On the Ø90 plate
# the five correctly-seated crosses sat 0.02, 0.10, 0.48, 0.48 and 1.23 mm from
# the boundary; the one sitting inland sat 5.96 mm out. Anything from about 1.5
# to 5.5 mm separates them, so 2.0 mm keeps every genuine rim seat and refuses
# the inland one — the smallest rule that fixes the seat without disturbing the
# rest of the plate. Tightening it to 0.5 mm also works but re-seats crosses
# that were already right, and moves the whole layout with them.
RIM_FLUSH = 2.0

# Seat cut stones at both rims BEFORE filling the row. Measured far worse than
# filling whole-first and capping afterwards; see seed_row_ends().
CUTS_FIRST = False

# The rows across the middle of the plate are its widest and most useful part,
# so they stay for whole stones — cut ones are pushed to the rows above and
# below. Half-height of that protected band, in mm.
CENTRE_BAND = 10.0

# Most plate a cut stone's missing corner may waste INSIDE the circle, mm2.
#
# This was 1.0, which is not a strict gate — it is a closed one. Measured over
# the 56 cut stones in stock, taking each stone's BEST orientation at its BEST
# rim seat on any row of a Ø80 plate:
#
#   ground-corner triangle  median 2.22 mm2, up to 12.84
#   best in-plate notch     min 0.08, median 2.22, p75 5.33
#   admitted at 1.0 mm2      7 of 56          at 3.0 mm2   34 of 56
#
# The median stone cannot get its notch below the size of the corner itself,
# because a chord is straight and the rim only curves away from a corner in the
# few places the two happen to match. Holding the gate at 1.0 therefore did not
# keep cut stones out of bad seats — it kept them out of every seat, and 0 of 56
# were used on a plate. 3.0 mm2 is about one corner triangle: a notch that size
# sits under the stone's own footprint, and the pile actually drains.
CUT_NOTCH_MAX = 3.0

# What using a cut stone is WORTH, as plate area in mm2, when it competes with a
# plain one for the same seat.
#
# Ranking cut stones as an absolute winner (-1 ahead of a plain stone's 0) is
# what made them unusable in the other direction: on a short row near the rim
# every seat counts as a rim seat, so a narrow cross took the FIRST seat of the
# row and closed it, leaving 17 mm of bare plate behind. Scoring it as a bonus
# instead keeps the choice honest — a cut stone wins the seat unless a plain one
# is better by more than this — and a cut stone is genuinely worth something:
# left in the drawer it is 100% waste, and the rim is the only place it can go.
# 30, measured across four plate sizes on the full 163-seed inventory. It is NOT
# the lever that decides how many cut stones get used — raising it from 12 to 60
# left the count identical at every size (5, 5, 6, 8), because what limits cut
# stones is how many seats exist where one can sit flush on the rim AND leave a
# channel under ROW_END_CUT_GAP, not how hard the ranking prefers them. What it
# does buy is a better choice among stones that were already competing: Ø90 rose
# 86.29% to 87.32%, and Ø80, Ø100 and Ø110 were unchanged.
CUT_BONUS = 30.0

# RECONSIDERATION: after the plate is packed, revisit finished rows and keep any
# exchange that seats one more stone. The packer is greedy — it commits a seat
# the moment it fills one and never looks back — so a row can end 7 mm short
# while carrying a 13 mm stone a narrower one would have served.
#
# ON. It was held off after its first attempt returned a plate with the rows
# shoved against the left rim; that was the re-lay opening every rebuilt row at
# the left chord, and it now centres the row instead. The exchange is also
# atomic — a rebuild that does not actually gain a seat puts the old row back —
# and it keeps the row's height line, so a reconsidered row is still level.
RECONSIDER = True
RECONSIDER_ROUNDS = 3

# How far a sweep-up seat may sit off an established row baseline, mm.
SWEEP_ROW_TOL = 0.6

# ROW-END FILL — the last pass, on the winning plate only. See _fill_row_ends.
# ROW_END_STEP is how finely a stone is slid along a row baseline looking for a
# corner seat; ROW_END_ROUNDS caps how many it may add.
ROW_END_STEP = 0.25
# How many row-end seats this pass may add. A BUDGET, not a target: the pass
# stops early when it runs out of seats it can legally fill.
#
# Held at the validated 6. It was briefly raised to 12 after measuring that the
# cap was binding on a wide band (63 -> 68 seats, 90.52% -> 91.20%), and that
# was the wrong trade: the extra seats are small stones dotted along the rim,
# which is precisely the appearance the shop floor rejected. Coverage is not the
# objective — a buildable, uniform plate is. See ROW_END_MIN_HEIGHT_FRAC.
#
# Back up to 20 now that the seats this pass finds are constrained on quality
# rather than on count. When it was raised to 12 before, the only limit on a
# seat was that it fitted, so a bigger budget bought more scattered stones and
# the plate got worse. Two rules now decide admission instead —
# ROW_END_MIN_HEIGHT_FRAC (how big a step the seat leaves) and ROW_END_MAX_GAP
# (whether it is attached to its row at all) — so the budget can go back to
# being a runaway guard rather than the thing shaping the plate.
ROW_END_ROUNDS = 20

# How tall a row-end stone must be relative to the ROW it joins, as a fraction.
#
# This pass exists to reach the trapezoid corners a full-height row cannot, so
# the stone it seats is legitimately SHORTER than its row — but there is a limit
# past which the result stops looking like a row with a corner filled and starts
# looking like debris swept against the rim. A 2.4 mm stone at the end of an
# 11.3 mm row leaves a 9 mm step; on the plate that reads as a gap with a chip in
# it, and the floor will not build it.
#
# Measured on the rejected Ø90 wide-band plate, the row-end stones were 3.91,
# 2.41, 4.94, 4.83, 3.40 and 2.41 mm against rows of 11.29, 10.35, 10.23, 10.23,
# 6.44 and 6.44 mm — ratios of 0.35, 0.23, 0.48, 0.47, 0.53 and 0.37.
#
# 0.45 rather than the 0.60 first tried. A row END is a TRAPEZOID: the row is
# bounded by its narrow outer edge, so the space beyond the last stone is a
# triangle that is only tall enough for a shorter stone. Measured on the current
# Ø90 plate, the eight rows leave 1.6, 7.6, 8.8, 12.1, 12.1, 12.7, 13.4 and
# 29.7 mm of chord unused at their ends — that is where the plate's remaining
# 13% lives, and a 0.60 floor refuses almost every stone that could reach it.
#
# The height ratio alone was never the thing that made the rejected plate look
# wrong, either. What made it look wrong was stones sitting apart from anything,
# which is what ROW_END_MAX_GAP now governs. Ratio controls how big a STEP the
# seat leaves; the gap rule controls whether it reads as attached at all.
ROW_END_MIN_HEIGHT_FRAC = 0.45

# How far a row-end stone may finish from the row it is filling beside, in mm.
#
# This is the rule that separates "the corner of a row is filled" from "there is
# debris against the rim". A stone tucked hard against the last seed of its row
# reads as part of that row however much shorter it is; the same stone sitting
# 3 mm out with a channel behind it reads as a loose chip, and the channel is
# unfillable besides.
#
# The pass already applied exactly this rule to CUT stones (ROW_END_CUT_GAP,
# 1.0 mm) with the reasoning that a stone which cannot nest in "holds a seat open
# and leaves a channel nothing can fill". That is just as true of a whole stone;
# it simply was not enforced, so whole stones were free to land anywhere the
# sweep first found room. Applying it to every stone is what lets the height
# ratio be relaxed without the plate degenerating again.
ROW_END_MAX_GAP = 1.0

# The RESIDUAL fill — a final sweep for any pocket a real seed still fits,
# without requiring it to line up with a row at all.
#
# OFF. It is the single largest source of both the appearance the floor rejected
# and the generation time users complain about, and it loses on both counts at
# once. Measured on the live Ø90 pool, band 2-12 mm:
#
#                              seats   fill%    plate structure        time
#   row sweep only             49      86.25    9 rows, 0.18 mm spread  715 s
#   + row-end + residual       63      90.52    16 baselines, 6 of      +620 s
#                                               them a single stone
#
# The 4.27 points it adds are 14 small stones scattered around the rim as
# isolated blobs — its own docstring says it "does NOT require the additional
# seed to match an existing row height", which is exactly the property that
# makes a plate unbuildable. And it costs as much wall-clock as the entire
# 20-pack search that produced the good plate.
#
# Left in place rather than deleted: the pocket search is sound work and could
# be reused if it is taught row discipline. Turning it on again needs a plate
# rendered and looked at, not a coverage number.
RESIDUAL_FILL = False
ROW_END_NEST = 0.02       # mm — how finely the seat is slid back to touch
# A row-end seat is laid FLUSH WITH ONE OF ITS ROW'S TWO EDGES, and nowhere
# between them. Offsetting it by hundredths to squeeze past a protruding
# neighbour was tried and is worse: it splits one row into two baselines a tenth
# of a millimetre apart, which reads as a broken row. The protrusion is prevented
# in the sweep-up pass instead.
#
# Which edge matters enormously, and trying only the baseline is why this pass
# found nothing on the lower half of the plate. A row-end gap is a TRIANGLE, and
# it is widest at the row's INNER edge — the side nearer the plate centre, where
# the chord is longest. For a row above centre that is its baseline; for a row
# BELOW centre it is its top, and probing the baseline there probes the one part
# of the wedge with no room in it.
#
# Measured on the Ø80 usable plate, room for a 4 mm stone at each row end:
#
#   row y     bottom-flush   top-flush
#   -37.50      0.22 mm      10.22 mm
#   -28.06     -0.69 mm       4.45 mm    (baseline has NO room at all)
#   -17.62      0.21 mm       2.98 mm
#    -6.18      0.57 mm       0.82 mm
#    +4.25      3.74 mm       1.47 mm    <- above centre, baseline wins
#   +15.51      4.37 mm       0.05 mm
#   +25.76      2.18 mm      -5.20 mm
#
# So both are tried and the seat scoring picks between them. A top-flush stone is
# not a floating one: its top edge lies on the line the NEXT ROW starts from, so
# the row above rests on it and the eye reads one continuous seam.
ROW_END_LIFTS = (0.0,)          # kept: the baseline is always one of the candidates


def _row_end_lifts(rh, h):
    """Flush positions for a stone of height `h` in a row-end gap `rh` tall.

    Always the row's baseline, plus the top-flush position when the stone is
    genuinely shorter than the gap. Never anything between the two.
    """
    lifts = [0.0]
    top = rh - h
    if top > 1e-6:
        lifts.append(top)
    return lifts
# Widest channel a rim-seated CUT stone may leave between itself and its row.
# It cannot be nested in without coming off the rim, so beyond this the seat is
# better given to a whole stone that can close up.
ROW_END_CUT_GAP = 1.0

# CENTRE-TO-OUTER SEED SIZING for Max Coverage: bias tall height classes toward
# the middle rows of the plate and short ones toward the rim.
#
# 0.0 = OFF, and off is the shipped default. Read the note in pick_level() before
# raising it: ranking classes by height was already tried there and LOST coverage
# on every pool measured (86.44% -> 83.94% on the 163-stone pool), because
# draining the tall classes a row at a time leaves no class able to cover a
# chord. This knob is not that change — it only shifts the preference by WHERE
# the row sits, and leaves the "can this class finish the chord" gate untouched —
# but it is the same family, so it does not go on without numbers behind it.
#
# The Arrange packer gets centre-out sizing unconditionally (see
# pack_v2._centre_out_row); there it is a pure reorder within a row and cannot
# cost anything. Here it changes which stones are picked, so it must be paid for.
#
# MEASURED INERT — do not expect this knob to do anything, and do not raise it
# hoping for a centre-to-outer gradient. It was added when the outward-cross rule
# was still `out > 0.0`, and against THAT broken baseline it appeared to gain
# 0.69 points. Once CUT_OUTWARD_MIN was fixed the apparent gain vanished:
#
#   band     bias   seats   fill%    inner mm2   outer mm2   ratio
#   8-12     0.0    35      85.01    123.3       108.5       1.14
#   8-12     3.0    35      85.01    123.3       108.5       1.14   identical
#   2-12     0.0    63      90.52    112.9        30.4       3.72
#   2-12     3.0    63      90.52    112.9        30.4       3.72   identical
#
# The lever that actually produces the gradient is the SEED-WIDTH BAND on the
# criteria form. Widening it from 8-12 to 2-12 moves the inner/outer area ratio
# from 1.14 to 3.72 and coverage from 85.01% to 90.52%, because the row-height
# census in pick_level then HAS short classes to open the narrow rim rows with.
# Starve it of small stones and no ranking bias can invent them; give it them and
# it sorts them outward on its own. That is why the band is the user-facing
# control and this is not.
#
# Kept, not deleted, because the pick_level ranking it hooks into is pool
# dependent and a different inventory could make it bite. If you do turn it on,
# measure against the current default first — and remember the 2026-08-15
# physical validation was run at 0.0.
CENTRE_SIZE_BIAS = 0.0

# CENTRE-TO-OUTER SIZE GRADIENT, second attempt — the one that can actually move
# the ranking. How wide a band counts as "equally suited to this row", as a
# FRACTION of the height span the pool actually offers; 0.0 = OFF.
#
# WHY CENTRE_SIZE_BIAS CANNOT DO THIS. It adds its term to `stock`, and `stock`
# is a total stocked WIDTH in mm, so the two are not the same size of number.
# Measured on the live Ø70 feed, 236 stones, band 7-12, at a rim row (chord
# 11.5 mm) every one of the 15 classes clears `fits` — the 7.00 class holds
# 37.8 mm against an 11.5 mm chord — and the ranking still goes to class 9.40 on
# 193.6 mm of stock. Second place is 86.9 mm, a gap of 106.7 mm, while
# CENTRE_SIZE_BIAS at 3.0 can move a class by at most 18.4 mm. It is an order of
# magnitude short, which is exactly why it measured byte-identical.
#
# The small classes are not being REFUSED, they are being out-stocked, and they
# are out-stocked at the rim by the same margin as at the centre. That is the
# whole defect: one best-stocked class wins every row whatever its position, so
# the plate comes out of a single size with no gradient and no mixture. On the
# Ø70 7-12 plate nothing under 9.42 mm was placed although 132 of the 236 stones
# are smaller than that.
#
# WHAT THIS DOES. Among the classes that already pass `fits`, prefer the one
# whose height suits WHERE the row sits: the tallest fitting class at the middle
# of the plate, the shortest at the rim, interpolated in between. Classes within
# the tolerance of that target tie, and `stock` breaks the tie exactly as it does
# today — so the anti-fragmentation property that ranking buys is kept, and only
# the choice BETWEEN equally-well-stocked-enough classes changes.
#
# WHY A FRACTION AND NOT MILLIMETRES. This shipped as an absolute 1.5 mm and that
# was a latent bug, not a tuning choice: the TARGET is interpolated across the
# span the pool offers, so comparing it against a fixed millimetre count means
# the rule changes character whenever the stock does. Measured, on Ø90:
#
#   pool                       span     best tol   as a fraction
#   old, width 7.01-11.44      4.43 mm  1.5 mm     34%
#   new, width 7.00-9.91       2.91 mm  1.0 mm     34%
#
# The same fraction, a different number of millimetres. Left at 1.5 mm the new
# inventory ran at 52% of its span, which ties over half the classes at every
# row, hands the decision back to `stock`, and lands WORSE than switching the
# gradient off — 43 seats/85.66% against 44/86.35% off and 47/87.21% at 34%.
# A fixed millimetre value has to be re-measured against every new inventory;
# a fraction does not, which is the whole point.
#
# Bounded below by LEVEL_BAND so the tolerance can never be finer than the census
# can resolve, and disabled outright when the span is one band or less — with a
# single height in play there is nothing to grade.
#
# `fits` stays the primary key and is untouched. A class that cannot finish the
# chord still never opens the row, which is the gate that stops the tall classes
# being drained a row at a time — the failure mode documented in pick_level().
#
# VALIDATED ON EVERY SUPPORTED PLATE SIZE before being turned on, band 7-12,
# 236-stone pool, against the ranking that preceded it:
#
#   plate   seats            fill%              radius/size corr   classes used
#   Ø70     18 -> 22         79.47 -> 81.80     +0.025 -> -0.640    3 -> 8
#   Ø90     36 -> 38         83.96 -> 86.16     -0.315 -> -0.338    ? -> 8
#   Ø158    150 -> 155       90.01 -> 90.61     -0.251 -> -0.344    ? -> 24
#
# Seats up and coverage up on all three, the correlation negative on all three
# (bigger toward the middle), and the plate drawn from many classes instead of
# one. The Ø70 plate had been using NOTHING under 9.42 mm; it now runs 7.03 to
# 10.44 across 8 classes.
#
# The response is NOT monotonic in the fraction — on the old pool at Ø90, 23%
# gave 35 seats/83.06% and 45% gave 34/84.34%, both regressions, while 34% gave
# 38/86.16%. So this is still a measured value and not a free parameter: re-run
# the three-size validation before moving it. What it no longer needs is
# re-measuring for every new INVENTORY, which is what the millimetre form did.
#
# WHAT IT DOES NOT FIX: the horizontal bands between rows on Ø158. Those come
# from stones being shorter than the row they sit in, and the gradient slightly
# WORSENS that measure (229 -> 249 mm2 lost to height mismatch, worst spread
# 2.11 -> 3.82 mm) even as it cuts total empty plate from 1719 to 1615 mm2. It
# buys more seats and more coverage, not flusher rows. Closing those bands is a
# row-height-discipline change and is still open.
SIZE_GRADIENT_FRAC = 0.34

# HOW FAR OUTWARD a ground corner must actually point, as the cosine of the
# angle between the cut direction and the radius through the stone.
#
#   1.00  straight out along the radius
#   0.71  45 degrees off
#   0.50  60 degrees off
#   0.00  exactly tangential — the cut points ALONG the rim, not through it
#
# The rule used to be `out > 0.0`, which is not "facing outward" but merely "not
# facing inward". A stone scoring +0.06 has its cross within four degrees of
# tangential: on the plate it reads as a notch in the side of the row, and the
# floor rejects it exactly as it rejects an inward one. Two of the four cut
# stones on the reported Ø90 plate passed at +0.060 and +0.543.
#
# Set from the plate, not picked round. Measured on the live Ø90 pool, seed
# width 8-12 mm, 202 stones of which 82 are cut, CENTRE_SIZE_BIAS 3.0:
#
#   min    seats   fill%    cut used   worst outward on the plate
#   0.00   34      84.27    4          0.060   <- tangential, the reported defect
#   0.35   35      85.01    4          0.603
#   0.50   35      85.01    4          0.603
#   0.71   34      84.01    0          none placed — too strict to satisfy
#
# 0.50 was the first fix and was NOT enough. It admits a cross 60 degrees off
# radial, and on the Ø90 2-12 mm plate one stone (DOMI002072) sat at +0.531 while
# every other cut stone scored 0.918 or better. At a glance it reads as facing
# sideways, and the floor rejected it exactly as it rejects an inward one — the
# distribution is bimodal, so a threshold in the middle of the gap admits the
# outlier and nothing else.
#
# Re-measured on the Ø90 pool, band 2-12 mm, with the residual pass off:
#
#   min    seats   fill%    cut used   worst cross on the plate
#   0.50   51      86.63    4          +0.531   <- the reported defect
#   0.70   46      87.58    3          +0.934
#
# 0.70 costs one cut stone and BUYS coverage (86.63 -> 87.58): the rim seat that
# stone was holding goes to a whole stone that fits it better. An earlier sweep
# found 0.71 unusable, but that was on the 8-12 mm band where only four cut
# stones existed at all; on a band that carries the small cut stock it is
# comfortably satisfiable.
CUT_OUTWARD_MIN = 0.70


# RIM POCKET FILL — the last pass, after the rows and the row ends are settled.
#
# What is left at that point is the crescent between the outermost stones and the
# rim, broken into small pockets. Measured on the accepted Ø90 2-12 mm plate:
# 546 mm2 across the ring, of which eight pockets can each take a real stone from
# the leftover inventory — worth 1.84 points (89.14% -> 90.98%).
#
# This is NOT the old residual fill that was turned off. That one placed a stone
# anywhere it fitted, which is how the plate ended up with stones floating in
# open space. The rule that makes this acceptable is RIM_POCKET_TOUCH: a stone
# may only take a pocket if it lands against something already on the plate. Plus
# the rules the rest of the plate already obeys — inside the usable circle, the
# form's seed spacing, and a cut stone only where _cut_on_rim allows it, so the
# outward-cross requirement is carried into this pass unchanged.
#
# Nothing already placed is moved, resized, rotated or removed.
RIM_POCKET_FILL = True
# Ignore free scraps below this. Sized from the SMALLEST stone the run could
# ever place — P.MINSEED, the criteria form's "min filler size" — rather than a
# fixed number, so a run whose stones are all large does not waste time probing
# slivers, and a run carrying 2 mm stock still sees the pockets that suit it.
RIM_POCKET_MIN_FACTOR = 0.8
# Candidates are collapsed by size so the same seat is not probed once per stone
# that would fill it identically. One representative is tried and the stock
# number is drawn from the matching stones only once a seat is won.
#
# The match must be EXACT. Bucketing to a tolerance (0.1 mm was tried) places the
# representative's outline and then labels it with a different stone's number, so
# the plate shows a seat the chosen stone does not actually have. It surfaced as
# three seeds reported "trimmed" on a plate where nothing is ever cut — the
# stored rawL x rawW no longer matched the polygon drawn. Measurements are held
# to 2 dp, so exact matching still collapses genuine duplicates; it simply never
# claims two different stones are interchangeable when they are not.
RIM_POCKET_SIZE_RES = None  # exact match only — see above
RIM_POCKET_CELL = 8.0       # mm — neighbourhood each search is bounded to
RIM_POCKET_STEP = 0.5       # mm — finest probe spacing
# How many probe positions across a pocket, per axis. The old search tried each
# corner of the pocket plus a couple of nudges — 25 positions — which is dense
# enough for a pocket the size of the stone and far too sparse for a big one. A
# 22.4 x 13.0 mm sliver on the reference plate could take a 2.58 x 2.04 stone and
# the corner probes all missed it, because the only seat was in the middle.
#
# Sampling a fixed COUNT per axis rather than a fixed step keeps the cost
# constant whatever the pocket's size: a small pocket is probed finely, a large
# one coarsely but everywhere, and neither explodes.
#
# It is a FALLBACK, not a replacement for the corner probes. Sweeping the pocket
# uniformly INSTEAD of trying its corners was measured and is worse — 61 seats
# and 90.88% against 63 and 92.20% — because a stone belongs tucked into a corner
# of a gap, and a uniform grid spends most of its probes in the middle where
# nothing can sit. Corners first, then the grid for what the corners miss.
RIM_POCKET_PROBES = 14
_NUDGES = (0.0, RIM_POCKET_STEP, -RIM_POCKET_STEP,
           2 * RIM_POCKET_STEP, -2 * RIM_POCKET_STEP)
RIM_POCKET_TOUCH = 0.05     # mm — how close counts as "against a neighbour"
RIM_POCKET_MAX = 40         # runaway guard, not a target
# Revisiting this pass's own seats to see whether a bigger stone now fits.
#
# OFF (0 rounds). It WORKS — on the reference plate it found seed 60, a
# 2.08 x 2.56 stone sitting in a hole that later placements had grown to take a
# 3.12 x 3.91 one. It is simply not worth what it costs. Measured on the live
# Ø90 2-12 mm plate, rim pass only:
#
#   configuration                       seats added   fill%    rim pass
#   corners + grid                          19        92.72     1010 s
#   + vertex probe + upgrade (3 rounds)     21        92.88     2635 s
#
# 0.16 points of coverage for 27 extra minutes a plate, against a target of 90%
# that the cheap configuration already clears by 2.7 points. Raise it only if a
# particular inventory leaves obvious upgrades on the plate and the time is
# acceptable that day.
RIM_UPGRADE_ROUNDS = 0
RIM_UPGRADE_MIN_GAIN = 0.5  # mm2 — ignore swaps that gain next to nothing
# How far beyond its own seat an upgrade may reach, in mm. A stone can only grow
# into space that touches it, and bounding the region here is what makes the
# candidate filter bite — see the note in the upgrade loop.
RIM_UPGRADE_REACH = 4.0
# Probe the free region's OWN vertices as well as its bounding-box corners.
# Shape-aware, and it does find seats the corners and the grid both miss — but it
# runs for every candidate the corners rejected, which is most of them. Together
# with the upgrade pass above it turned a 1010 s pass into 2635 s for two extra
# stones. OFF for the same reason: see the table beside RIM_UPGRADE_ROUNDS.
RIM_VERTEX_PROBE = True
RIM_POCKET_ROUNDS = 8       # how many times the ring is re-examined


def _fill_rim_pockets(real, placed, fill, R):
    """Seat leftover stones in the rim pockets. Additive only.

    Works pocket-first rather than stone-first: the leftover ring is diced on a
    coarse grid, and each pocket asks the unused inventory for the LARGEST stone
    that legally fits it. Sweeping every stone over the whole plate instead is
    what made the old residual pass cost as much as the entire row search.
    """
    if not RIM_POCKET_FILL or not placed or not real:
        return placed, fill

    from shapely import affinity
    from shapely.geometry import Point, Polygon, box as shbox
    from shapely.ops import unary_union
    from shapely.strtree import STRtree

    disc = Point(0.0, 0.0).buffer(R, resolution=180)
    RSQ = (R - RIM_EPS) ** 2
    geoms = [Polygon(p["poly"]).buffer(0) for p in placed]
    clear = max(0.0, float(getattr(P, "CLEARANCE", 0.0) or 0.0))
    # Smallest scrap worth probing, derived from the run's own minimum seed size
    # rather than hard-coded — see RIM_POCKET_MIN_FACTOR.
    min_seed = max(0.5, float(getattr(P, "MINSEED", 2.0) or 2.0))
    min_area = (min_seed ** 2) * RIM_POCKET_MIN_FACTOR

    used = {p["stock"] for p in placed}
    # Build ONE candidate per distinct (size, orientation, cut?) and remember
    # which stones can supply it. Cut stones are never collapsed together: their
    # outlines differ in ways a bounding box does not capture, and which way the
    # ground corner faces is the whole question for them.
    by_shape = {}
    for s in real:
        if s["stock"] in used:
            continue
        g0 = _seed_footprint(s)
        if g0 is None or g0.is_empty:
            continue
        cut = bool(s.get("poly"))
        for deg in ENHANCED_ANGLES:
            rg = affinity.rotate(g0, deg, origin="centroid") if deg else g0
            b = rg.bounds
            w, h = b[2] - b[0], b[3] - b[1]
            # Cut stones are never collapsed at all: their outlines differ in
            # ways a bounding box does not capture, and which way the ground
            # corner faces is the whole question for them.
            key = (s["stock"] if cut else (round(w, 6), round(h, 6)), deg, cut)
            entry = by_shape.get(key)
            if entry is None:
                by_shape[key] = [affinity.translate(rg, -b[0], -b[1]),
                                 w, h, cut, deg, [s]]
            else:
                entry[5].append(s)
    if not by_shape:
        return placed, fill
    # LARGEST first. A pocket should be filled by the biggest stone that fits it,
    # not the first one tried — two 3 mm stones where one 6 mm stone belongs is a
    # worse plate for the same coverage.
    poses = sorted(by_shape.values(), key=lambda e: -e[0].area)

    added = []

    tree = [STRtree(geoms)]

    # A box-level prefilter was tried here — compute the candidate's bounds
    # arithmetically and skip the translate when the index shows no neighbour
    # within touch range. It is logically sound (legal() cannot pass without a
    # neighbour) and produced byte-identical plates, but it is SLOWER: 1209.7 s
    # -> 1552.1 s. Almost every probe in a rim crescent DOES have a neighbour,
    # so the prefilter never fires and only adds a box construction and an index
    # query per candidate. Do not reintroduce it without measuring.
    def legal(g):
        # Analytic rim test first — disc.contains() walks a 180-gon and this is
        # the hottest call in the pass. A bounding box inside the circle
        # guarantees the stone inside it is too.
        bx0, by0, bx1, by1 = g.bounds
        if max(bx0 * bx0, bx1 * bx1) + max(by0 * by0, by1 * by1) > RSQ:
            if not disc.contains(g):
                return False
        # Then only the handful of stones that could possibly be in the way,
        # via the spatial index rather than all of them. The query region is the
        # stone's BOX grown by the tolerance, not g.buffer() — buffering a
        # polygon rounds every corner with a fresh ring of vertices and, called
        # once per candidate seat, it dominated this pass outright.
        t = clear + RIM_POCKET_TOUCH
        near = tree[0].query(shbox(bx0 - t, by0 - t, bx1 + t, by1 + t))
        touching = False
        for i in near:
            o = geoms[i]
            if g.intersection(o).area > 1e-9:
                return False
            d = g.distance(o)
            if clear > 0.0 and d < clear - 1e-9:
                return False
            if d <= clear + RIM_POCKET_TOUCH:
                touching = True
        # ATTACHED, not floating — the rule the old residual pass lacked. The
        # query radius is exactly the touch distance, so anything it returned at
        # that range IS a neighbour; an empty result means the stone is adrift.
        return touching

    # Re-examine the ring after every round: placing one stone changes the free
    # space, and a pocket that held one 6 mm stone often still has room for a
    # 3 mm one beside it. The fixed-grid version served each cell once and moved
    # on, which is what left large regions holding a single small stone.
    for _round in range(RIM_POCKET_ROUNDS):
        if len(added) >= RIM_POCKET_MAX:
            break
        free = disc.difference(unary_union(geoms))
        # Dice the leftover ring into CELLS. The ring is one connected crescent
        # whose bounding box is the whole plate, so searching it as a single
        # region means probing the full diameter to rediscover that the middle is
        # occupied. A cell bounds each search to a neighbourhood — and because
        # the free space is recomputed every round, a region wider than one cell
        # is served repeatedly rather than once, which is what lets a large
        # pocket take a big stone AND a small one beside it.
        parts = []
        n = int(math.ceil(2 * R / RIM_POCKET_CELL))
        for i in range(n):
            for j in range(n):
                cell = shbox(-R + i * RIM_POCKET_CELL, -R + j * RIM_POCKET_CELL,
                             -R + (i + 1) * RIM_POCKET_CELL,
                             -R + (j + 1) * RIM_POCKET_CELL)
                bit = free.intersection(cell)
                if bit.area >= min_area:
                    parts.append(bit)
        if not parts:
            break
        seated_this_round = False
        for part in sorted(parts, key=lambda c: -c.area):
            if len(added) >= RIM_POCKET_MAX:
                break
            px0, py0, px1, py1 = part.bounds
            pw, ph = px1 - px0, py1 - py0
            best = None
            for rg, w, h, is_cut, deg, supply in poses:
                # A stone may overhang the cell into the neighbouring free space;
                # the cell bounds the SEARCH, not the stone.
                if w > pw + RIM_POCKET_CELL or h > ph + RIM_POCKET_CELL:
                    continue
                if best is not None and rg.area <= best[0]:
                    break                  # sorted by area: nothing left can win
                s = next((c for c in supply if c["stock"] not in used), None)
                if s is None:
                    continue               # every stone of this size is spent
                # CORNERS FIRST — a stone belongs tucked into a corner of the gap.
                for ax in (px0, px1 - w):
                    for ay in (py0, py1 - h):
                        for ddx in _NUDGES:
                            for ddy in _NUDGES:
                                g = affinity.translate(rg, ax + ddx, ay + ddy)
                                if legal(g) and ((not is_cut) or _cut_on_rim(g, R)):
                                    best = (g.area, s, deg, g)
                                    break
                            if best is not None:
                                break
                        if best is not None:
                            break
                    if best is not None:
                        break
                if best is not None:
                    continue
                # THEN THE POCKET'S OWN VERTICES. A rim pocket is a crescent, and
                # its seat is usually against a corner of the crescent itself
                # rather than of its bounding box — the box corners sit out in
                # the curve where nothing fits. These are the points where the
                # rim meets a stone, which is exactly where a small stone tucks
                # in. Cheap: a few dozen points after rounding, against a grid's
                # couple of hundred.
                # A cell can slice the ring into several disjoint pieces, so the
                # free region here may be a MultiPolygon — walk whatever it is.
                verts = set() if RIM_VERTEX_PROBE else ()
                for piece in (getattr(part, "geoms", [part]) if RIM_VERTEX_PROBE else ()):
                    ext = getattr(piece, "exterior", None)
                    if ext is None:
                        continue           # a stray line or point: no seat in it
                    for ring in [ext] + list(piece.interiors):
                        for vx, vy in ring.coords:
                            verts.add((round(vx * 4.0) / 4.0, round(vy * 4.0) / 4.0))
                for vx, vy in sorted(verts):
                    for ox, oy in ((0.0, 0.0), (-w, 0.0), (0.0, -h), (-w, -h)):
                        g = affinity.translate(rg, vx + ox, vy + oy)
                        if legal(g) and ((not is_cut) or _cut_on_rim(g, R)):
                            best = (g.area, s, deg, g)
                            break
                    if best is not None:
                        break
                if best is not None:
                    continue
                # FALLBACK: sweep the pocket. Only reached when no corner works,
                # which is the case a big sliver presents — its one seat is in
                # the middle. Bounded to RIM_POCKET_PROBES positions per axis so
                # the cost does not grow with the pocket.
                sx, sy = max(0.0, pw - w), max(0.0, ph - h)
                nx = min(RIM_POCKET_PROBES, int(sx / RIM_POCKET_STEP) + 1)
                ny = min(RIM_POCKET_PROBES, int(sy / RIM_POCKET_STEP) + 1)
                stx = sx / (nx - 1) if nx > 1 else 0.0
                sty = sy / (ny - 1) if ny > 1 else 0.0
                for ix in range(nx):
                    for iy in range(ny):
                        g = affinity.translate(rg, px0 + ix * stx, py0 + iy * sty)
                        if legal(g) and ((not is_cut) or _cut_on_rim(g, R)):
                            best = (g.area, s, deg, g)
                            break
                    if best is not None:
                        break
            if best is None:
                continue
            _a, s, deg, g = best
            added.append(g)
            geoms.append(g)                # later stones may lean on this one
            tree[0] = STRtree(geoms)       # keep the index in step with them
            used.add(s["stock"])
            seated_this_round = True
            bx0, by0, bx1, by1 = g.bounds
            rp = g.representative_point()
            placed.append({
                "stock": s["stock"], "cts": s.get("cts", 0.0),
                "L": round(bx1 - bx0, 1), "W": round(by1 - by0, 1),
                "H": s.get("H", round((P.T_LO + P.T_HI) / 2.0, 3)),
                "rawL": s.get("L"), "rawW": s.get("W"),
                "x": bx0, "y": by0, "w": bx1 - bx0, "h": by1 - by0, "angle": deg,
                "kind": "real",
                "poly": [(round(px, 3), round(py, 3)) for px, py in g.exterior.coords],
                "area": g.area,
                "irregular": bool(s.get("poly")),
                "lx": rp.x, "ly": rp.y,
            })
        if not seated_this_round:
            break                          # nothing more will fit

    if not added:
        return placed, fill

    # ---- UPGRADE ------------------------------------------------------------
    # Revisit the stones THIS PASS placed and see whether a bigger one now fits
    # where they sit. The fill is greedy: each pocket takes the largest stone
    # that fitted AT THE TIME, and every later placement changes the free space
    # around it — so an early choice can end up smaller than the space it now
    # has. Measured on the reference plate, seed 60 was a 2.08 x 2.56 stone in a
    # hole that had grown to take a 3.12 x 3.91 one.
    #
    # Strictly limited to this pass's OWN seats: `first` is where they begin, and
    # nothing before it — no row, no row-end seat — is ever touched.
    first = len(placed) - len(added)
    for _round in range(RIM_UPGRADE_ROUNDS):
        improved = False
        for idx in range(first, len(placed)):
            cur = placed[idx]
            cur_area = cur["area"]
            # The room this seat would have if its stone were lifted out.
            others = [g for k, g in enumerate(geoms) if k != idx]
            # LOCAL room only. `disc.difference(others)` is the whole leftover
            # ring, whose bounding box is the entire plate — filtering candidate
            # stones against that rejects nothing, so the pass probed every stone
            # bigger than the current one, everywhere, and cost more than the
            # rest of the plate put together. A seat can only grow into the space
            # touching it, so bound the region to its own neighbourhood first.
            box = shbox(cur["x"] - RIM_UPGRADE_REACH, cur["y"] - RIM_UPGRADE_REACH,
                        cur["x"] + cur["w"] + RIM_UPGRADE_REACH,
                        cur["y"] + cur["h"] + RIM_UPGRADE_REACH)
            room = (disc.difference(unary_union(others)) if others else disc)
            room = room.union(geoms[idx]).intersection(box)
            if room.is_empty:
                continue
            # Index them ONCE per seat. Without this every probe compared the
            # candidate against all ~70 stones on the plate, and the pass ran
            # 14x14 probes per candidate over hundreds of candidates per seat —
            # the arithmetic works out to millions of polygon intersections and
            # took longer than the entire rest of the plate.
            otree = STRtree(others) if others else None
            spot = None
            for rg, w, h, is_cut, deg, supply in poses:
                if rg.area <= cur_area + RIM_UPGRADE_MIN_GAIN:
                    break                  # sorted by area: no improvement left
                s = next((c for c in supply
                          if c["stock"] not in used or c["stock"] == cur["stock"]),
                         None)
                if s is None:
                    continue
                b = room.bounds
                if w > b[2] - b[0] + 1e-9 or h > b[3] - b[1] + 1e-9:
                    continue
                # Probe around where the current stone already is — the space it
                # can grow into is adjacent to it, not across the plate.
                near_box = cur["x"] - w, cur["y"] - h, cur["x"] + cur["w"], cur["y"] + cur["h"]
                nx = ny = RIM_POCKET_PROBES
                stx = (near_box[2] - near_box[0]) / max(1, nx - 1)
                sty = (near_box[3] - near_box[1]) / max(1, ny - 1)
                for ix in range(nx):
                    for iy in range(ny):
                        g = affinity.translate(rg, near_box[0] + ix * stx,
                                               near_box[1] + iy * sty)
                        if not _legal_against(g, others, otree, disc, RSQ, clear):
                            continue
                        if is_cut and not _cut_on_rim(g, R):
                            continue
                        spot = (s, deg, g)
                        break
                    if spot is not None:
                        break
                if spot is not None:
                    break
            if spot is None:
                continue
            s, deg, g = spot
            used.discard(cur["stock"])
            used.add(s["stock"])
            bx0, by0, bx1, by1 = g.bounds
            rp = g.representative_point()
            placed[idx] = {
                "stock": s["stock"], "cts": s.get("cts", 0.0),
                "L": round(bx1 - bx0, 1), "W": round(by1 - by0, 1),
                "H": s.get("H", round((P.T_LO + P.T_HI) / 2.0, 3)),
                "rawL": s.get("L"), "rawW": s.get("W"),
                "x": bx0, "y": by0, "w": bx1 - bx0, "h": by1 - by0, "angle": deg,
                "kind": "real",
                "poly": [(round(px, 3), round(py, 3)) for px, py in g.exterior.coords],
                "area": g.area,
                "irregular": bool(s.get("poly")),
                "lx": rp.x, "ly": rp.y,
            }
            geoms[idx] = g
            improved = True
        if not improved:
            break

    covered = unary_union([Polygon(p["poly"]).buffer(0) for p in placed]).area
    return placed, 100.0 * covered / (math.pi * R * R)


def _legal_against(g, others, otree, disc, rsq, clear):
    """Is `g` inside the plate, clear of `others`, and touching at least one?

    Same contract as the fill's own `legal`, but against an explicit list rather
    than the running index — the upgrade pass asks "would this fit if that stone
    were not there", which the running index cannot answer. `otree` indexes that
    list; only the neighbours it returns can possibly be in the way.
    """
    from shapely.geometry import box as shbox

    bx0, by0, bx1, by1 = g.bounds
    if max(bx0 * bx0, bx1 * bx1) + max(by0 * by0, by1 * by1) > rsq:
        if not disc.contains(g):
            return False
    if otree is None:
        return False                       # nothing to lean on
    t = clear + RIM_POCKET_TOUCH
    touching = False
    for i in otree.query(shbox(bx0 - t, by0 - t, bx1 + t, by1 + t)):
        o = others[i]
        if g.intersection(o).area > 1e-9:
            return False
        d = g.distance(o)
        if clear > 0.0 and d < clear - 1e-9:
            return False
        if d <= t:
            touching = True
    return touching


def _interior_waste(placed, disc):
    """What a layout wastes because a cross has nothing to face.

    Returns ``(notch_area, stone_count, stone_area)`` — the open notch beside
    such stones, how many there are, and how much stone is sitting in those
    unbuildable seats.

    A cut stone's missing corner has to go somewhere. Against the rim it costs
    nothing — that crescent was unusable anyway. Against another cross it costs
    nothing either, because the two sit flush. Against a FLAT neighbour it leaves
    a notch no stone can enter, and the shop floor cannot build that seat.
    """
    from shapely.geometry import Polygon as ShPoly, box as shbox
    from shapely.ops import unary_union

    if not placed:
        return 0.0, 0, 0.0
    occ = unary_union([ShPoly(p["poly"]).buffer(0) for p in placed])
    waste, n, stone = 0.0, 0, 0.0
    for p in placed:
        g = ShPoly(p["poly"]).buffer(0)
        void = shbox(*g.bounds).difference(g)
        if void.is_empty or void.area < 0.05:
            continue                                  # a plain rectangle
        inside = void.intersection(disc)
        if void.area - inside.area > void.area * 0.5:
            continue                                  # mostly past the rim
        open_part = inside.difference(occ)             # not taken by a neighbour
        if open_part.area >= void.area * 0.25:
            waste += open_part.area
            stone += g.area
            n += 1
    return waste, n, stone


def _pack_once(args, y0=None, policy="cut-first", fill="left", seat="compact"):
    """ONE complete packing run for MAX COVERAGE. Returns (placed, fill_pct).

    Draws nothing and mutates no caller state, so it is safe to call repeatedly
    on the same pool — enhanced_plate_job runs it once per candidate row phase
    and renders only the best result.

    Seeds arrive already cut or trimmed on the shop floor, so this packer never
    cuts anything. Each seed is placed at its true measured footprint or not at
    all; one that will not fit in any orientation is left for the next plate.
    That is why the rim stays ragged rather than being shaved flush.

    Seats are not a fixed grid. A row opens at the height of the seed that starts
    it, each seed is laid at its OWN width, and every seed is slid left until it
    touches its neighbour — the slide is what lets two opposing cut corners nest
    instead of each leaving a triangle of scrap.

    `args` = (real, pi, plate_d, R, min_size, path); `y0` is the absolute y of
    the first row's baseline, None meaning the bottom of the plate; `policy`
    decides who wins a seat when a cut stone and a plain one both fit:

      "cut-first"   — a cut stone takes the rim seat that opens a row, so the
                      cut pile drains a few per plate
      "plain-first" — a plain stone wins any tie, so crosses are not spent in
                      the middle of the plate where they leave a notch

    `fill` is the direction each row is laid:

      "left"   — open at the left chord and fill across, as it always has
      "centre" — start at the middle of the plate and grow both ways, so a row
                 ends ragged at both rims instead of flush left and ragged right
    """
    real, pi, plate_d, R, min_size, path = args
    P.PLATE_D = plate_d
    from shapely import STRtree, affinity
    from shapely.geometry import Point, Polygon as ShPoly, box as shbox
    from shapely.ops import unary_union

    disc = Point(0.0, 0.0).buffer(R, resolution=180)
    # The disc is a 180-gon inscribed in the true circle, so a box that clears
    # this radius is inside the polygon too. RIM_EPS keeps the analytic test on
    # the safe side of that difference.
    RSQ = (R - RIM_EPS) ** 2
    if not real:
        return [], 0.0

    defH = round((P.T_LO + P.T_HI) / 2.0, 3)
    EPS = 1e-6
    # "Distance between seeds" from the criteria form, in mm. Kept between seeds
    # only — the clear ring around the plate is the MARGIN, already taken out of
    # R before this runs. 0 packs edge to edge, as before.
    clear = max(0.0, float(getattr(P, "CLEARANCE", 0.0) or 0.0))
    # Narrowest stone in the pool. A leftover slimmer than this can never be
    # seated, which is what makes the SQUEEZE below worth attempting.
    narrowest = min((min(float(s.get("L") or 99.0), float(s.get("W") or 99.0))
                     for s in real), default=0.0)
    # Where each half of a centre-out row stops. The two halves grow towards each
    # other and must not touch, so the middle needs the gap as well: the
    # rightward half owns x >= +half, the leftward half x <= -half, leaving
    # exactly `clear` across the centre line and the same margin at both rims.
    #
    # Getting this wrong is not a cosmetic error. With both halves stopped at 0
    # the first leftward seed of every row landed flush against the first
    # rightward one, free() refused it for being 0 mm away when `clear` mm were
    # required, and — since no orientation of any seed could ever satisfy that
    # seat — the whole left side of the row was marked dead. Rows collapsed into
    # short fragments: at 0.5 mm a Ø90 plate fell to 19 stunted rows, 31 seats,
    # 68.10%, against 84.63% at 0 mm where the bug cannot bite.
    half = clear / 2.0

    shapes, cutdirs, poses = {}, {}, {}
    for s in real:
        g = _seed_footprint(s)
        if g is not None:
            shapes[id(s)] = g
            cutdirs[id(s)] = _cut_direction(g)
            # A seed only ever has these four orientations, and scan() reaches
            # for them once per seat. Rotating on demand redid the same few
            # hundred rotations ten thousand times a plate; turning each seed
            # once here and keeping its size alongside cuts a pack by a fifth.
            pl = []
            for deg in ENHANCED_ANGLES:
                # ROTATION ONLY, never a mirror: affinity.rotate preserves
                # handedness, matching how a seed can actually be laid down.
                rg = affinity.rotate(g, deg, origin="centroid") if deg else g
                b = rg.bounds
                # NORMALISED to the origin, so every pose's bounds are exactly
                # (0, 0, ww, hh). Seating one is then a translate to the seat's
                # own corner, and neither the pose nor the placement has to be
                # asked for its bounds — that question was being put to shapely
                # 85,000 times per pack, once for every candidate orientation
                # the scan looked at and discarded.
                rg = affinity.translate(rg, -b[0], -b[1])
                pl.append((deg, rg, b[2] - b[0], b[3] - b[1]))
            poses[id(s)] = pl
    queue = [s for s in real if id(s) in shapes]
    # Tall-first: the seed that opens a row sets that row's height, so starting
    # with the tallest stops a row being opened too short to be reused.
    #
    # Then WIDEST first among equal heights. Ties were previously broken by
    # whatever order the pool arrived in, and with far more seeds available than
    # one plate holds that choice is a real lever — spending a row's width on
    # fewer, wider seeds leaves less end-of-row remainder. Worth 1.6 points on a
    # Ø80 plate (77.08% to 78.71%) for a change of sort key alone.
    #
    # ASSUMED corners go last. A stone whose datasheet left the cross corner
    # blank is carrying a guess — LEFT-TOP — and a guess is right only about two
    # thirds of the time, because rotation turns LEFT-TOP into RIGHT-BOTTOM but
    # can never reach LEFT-BOTTOM or RIGHT-TOP, which are mirror images. Placing
    # the declared stones first means a plate that fills on known-good data never
    # spends a guessed one at all.
    queue.sort(key=lambda s: (1 if s.get("corner_assumed") else 0,
                              -(shapes[id(s)].bounds[3] - shapes[id(s)].bounds[1]),
                              -(shapes[id(s)].bounds[2] - shapes[id(s)].bounds[0])))

    placed, occ, used = [], [], set()
    srcs = []          # seed behind each placement, so one can be undone

    def pick_level(y):
        """Height classes for the row starting at `y`, best first.

        Every stone still in the pool is filed under the height it lies flat at,
        in LEVEL_BAND-wide classes. A class is judged on whether it can FINISH
        the row about to be laid: the chord at this height is measured, and the
        class's stones are counted against it — WHOLE stones only, since a cut
        stone can only ever take the one seat at each end.

        Choosing the class before the row opens is the whole point. The row is
        then locked to a height the inventory can actually finish, instead of to
        whichever stone happened to win the opening seat — which is how rows
        ended up 2 mm taller than anything left to fill them.

        Ranking on stock alone was not enough. A class holding three wide stones
        outranked one holding eight narrower ones, opened a row it could fill
        barely a third of, and left the rest of that chord bare — one plate came
        back with 5 seats on it while the NEXT plate, from almost the same pool,
        took 19. Whether the class covers the chord is what matters; how much it
        has spare only breaks the tie.
        """
        census = {}
        for s in queue:
            if id(s) in used:
                continue
            g = shapes.get(id(s))
            if g is None:
                continue
            bx0, by0, bx1, by1 = g.bounds
            h = min(bx1 - bx0, by1 - by0)
            w = max(bx1 - bx0, by1 - by0)
            k = round(round(h / LEVEL_BAND) * LEVEL_BAND, 2)
            c = census.setdefault(k, [0.0, 0.0])
            c[1] = max(c[1], h)
            if not s.get("poly"):
                c[0] += w
        if not census:
            return None

        # CENTRE-TO-OUTER GRADIENT (opt-in, see SIZE_GRADIENT_FRAC). The target
        # height for this row is interpolated across the heights of the classes
        # that can actually finish it, so it follows the user's band and the
        # plate diameter with nothing to tune. Computed once here rather than
        # inside rank(), which runs per class.
        _grad = SIZE_GRADIENT_FRAC > 0.0
        _target = 0.0
        if _grad:
            _fit_h = []
            for _k in census:
                _h = census[_k][1]
                _yo = max(abs(y), abs(y + _h))
                if _yo >= R:
                    continue
                if census[_k][0] >= LEVEL_MARGIN * 2.0 * math.sqrt(
                        R * R - _yo * _yo):
                    _fit_h.append(_h)
            if _fit_h:
                # 1 at the middle of the plate, 0 at the rim. Taken from the row
                # baseline, which is fixed before a class is chosen — using each
                # class's own height here would give every class a different
                # target and there would be nothing to rank against.
                _near = 1.0 - min(1.0, abs(y) / R)
                _span = max(_fit_h) - min(_fit_h)
                _target = min(_fit_h) + _span * _near
                # The tolerance is a FRACTION of the span the pool actually
                # offers, never a fixed millimetre count — see SIZE_GRADIENT_FRAC.
                # Never below one census band, or classes that the census cannot
                # tell apart would be ranked against each other.
                _tol = max(LEVEL_BAND, SIZE_GRADIENT_FRAC * _span)
                if _span <= LEVEL_BAND:
                    _grad = False        # one height in play; nothing to grade
            else:
                _grad = False

        def rank(k):
            h = census[k][1]
            yout = max(abs(y), abs(y + h))
            if yout >= R:
                return (0, 0.0, 0.0, k) if _grad else (0, 0.0, k)
            chord2 = 2.0 * math.sqrt(R * R - yout * yout)
            stock = census[k][0]
            # Among classes that can finish the row, the best STOCKED wins, and
            # height only breaks the tie.
            #
            # Both alternatives were tried and both are worse. Ranking by height
            # drains the tall classes a row at a time and leaves every class
            # holding a little, none able to cover a chord; the job dribbles out
            # over plates of 14, 8, 6 and 5 seats. Weighting stock BY height
            # spreads the size classes more evenly and narrows the seat-count gap
            # against Arrange, which looks reassuring — and it costs coverage on
            # every pool measured and puts a 22.89 mm hole back at the end of a
            # row on thin stock:
            #
            #   ranking        163-stone pool        95-stone pool
            #   stock          43 seats / 86.44%     50 / 82.40%, no holes
            #   stock x height 42 / 83.94%           47 / 79.91%, 22.89 mm hole
            #
            # The gap against Arrange is not a quality measure — it tracks which
            # stones each method happens to hold. This ranking is the one whose
            # output was validated on a real plate on 2026-08-15; do not trade it
            # for a more comfortable-looking comparison.
            fits = 1 if stock >= LEVEL_MARGIN * chord2 else 0
            if _grad:
                # Closeness to the row's target height, quantised so that classes
                # within _tol of it are EQUAL and fall through to `stock` — the
                # gradient decides which sizes belong at this radius, stock still
                # decides which of those to spend.
                near_t = -int(abs(h - _target) / _tol)
                return (fits, near_t, stock, k)
            if CENTRE_SIZE_BIAS <= 0.0:
                return (fits, stock, k)
            # CENTRE-TO-OUTER (opt-in, see CENTRE_SIZE_BIAS). `near` is 1 at the
            # middle of the plate and 0 at the rim, so the term is a BONUS for a
            # tall class on a middle row and the same size PENALTY for one on a
            # rim row. `fits` stays the primary key, so a class that cannot cover
            # this chord still never wins the row however tall it is — that gate
            # is what stops the tall classes being drained a row at a time.
            near = 1.0 - min(1.0, abs(y + h / 2.0) / R)
            return (fits, stock + CENTRE_SIZE_BIAS * h * (2.0 * near - 1.0), k)

        return [census[k][1] for k in sorted(census, key=rank, reverse=True)]

    def rollback(n, used_mark):
        """Undo every placement made since a mark, so a row can be tried again."""
        while len(placed) > n:
            placed.pop()
            occ.pop()
            srcs.pop()
        used.clear()
        used.update(used_mark)
        reindex()

    # Spatial index over what is already placed. free() is called thousands of
    # times — every probe, and every 0.25 mm step of every nesting slide — and
    # testing each call against all 40+ placed seeds made a plate take 25 s.
    # The index narrows it to the handful that could possibly be in the way.
    tree = [None]

    def reindex():
        tree[0] = STRtree(occ) if occ else None

    def free(g):
        """Wholly inside the plate, and at least `clear` mm from every neighbour."""
        # Cheap analytic rim test first. disc.contains() walks a 180-gon and is
        # called tens of thousands of times a pack; the four corners of the
        # bounding box against R settles almost every case without it, and a box
        # inside the circle guarantees the stone inside it is too.
        bx0, by0, bx1, by1 = g.bounds
        far = max(bx0 * bx0, bx1 * bx1) + max(by0 * by0, by1 * by1)
        if far > RSQ:
            # A corner of the BOX is outside — the stone itself may still be in
            # (a ground-off corner), so fall back to the exact test.
            if not disc.contains(g):
                return False
        t = tree[0]
        if t is None:
            return True
        probe = g.buffer(clear) if clear > 0.0 else g
        for i in t.query(probe):
            o = occ[i]
            if clear > 0.0:
                if g.distance(o) < clear - EPS:
                    return False
                continue
            # Seeds are laid touching, so `intersects` is true for every
            # neighbour and the old test paid for a full polygon intersection on
            # each one. Overlapping BOXES are a prerequisite for overlapping
            # areas, so reject on the boxes first — that settles a touching
            # neighbour with four comparisons instead.
            ob0, ob1, ob2, ob3 = o.bounds
            if (min(bx1, ob2) - max(bx0, ob0) <= EPS
                    or min(by1, ob3) - max(by0, ob1) <= EPS):
                continue
            if g.intersects(o) and g.intersection(o).area > EPS:
                return False
        return True

    def slide_left(g, limit=None):
        """Nudge a seed left while it stays legal — this is the nesting step.

        `limit` stops the seed's left edge going past a line. Centre-out needs
        it: the first stone of a row has nothing to its left, so without a stop
        it slides the full width of the plate to the rim — which defeats the
        point of starting at the centre AND costs hundreds of collision tests,
        since every 0.25 mm step is a full free() check.
        """
        cur = g
        while True:
            if limit is not None and cur.bounds[0] - NEST_STEP < limit:
                return cur
            nxt = affinity.translate(cur, -NEST_STEP, 0.0)
            if not free(nxt):
                return cur
            cur = nxt

    def slide_right(g, limit=None):
        """The mirror of slide_left, for a row growing leftwards from the middle.
        `limit` caps the seed's RIGHT edge."""
        cur = g
        while True:
            if limit is not None and cur.bounds[2] + NEST_STEP > limit:
                return cur
            nxt = affinity.translate(cur, NEST_STEP, 0.0)
            if not free(nxt):
                return cur
            cur = nxt

    def record(s, deg, g):
        """Commit a placement. Used by both the row sweep and the sweep-up pass,
        so the two can never drift in what they store."""
        used.add(id(s))
        srcs.append(s)
        occ.append(g)
        reindex()
        bx0, by0, bx1, by1 = g.bounds
        rp = g.representative_point()
        placed.append({
            "stock": s["stock"], "cts": s.get("cts", 0.0),
            "L": round(bx1 - bx0, 1), "W": round(by1 - by0, 1),
            "H": s.get("H", defH),
            # The stone's measurement from the import, for the seed list.
            "rawL": s.get("L"), "rawW": s.get("W"),
            "x": bx0, "y": by0, "w": bx1 - bx0, "h": by1 - by0, "angle": deg,
            "kind": "real",
            "poly": [(round(px, 3), round(py, 3)) for px, py in g.exterior.coords],
            "area": g.area,
            # No "clipped"/"nomarea": nothing is ever cut, so there is no trim to
            # report and dim_rows() renders these as plain real seeds.
            "irregular": bool(s.get("poly")),
            "lx": rp.x, "ly": rp.y,
        })

    def scan(y, x, row_h, allow_taller, going_left=False, want_cut=None,
             rim_seed=False, level=None):
        """Best seed for the seat at (x, y), or None. Returns (key, seed, deg, poly).

        `x` is the seat's NEAR edge: its left edge when the row grows rightwards,
        its right edge when the row grows leftwards from the middle of the plate.
        """
        best = None
        for s in queue:
            if id(s) in used:
                continue
            g0, cut0 = shapes[id(s)], cutdirs[id(s)]
            # WHOLE-first, then CAP the ends with cut stones. A row filled with
            # both at once lets a cut stone win an ordinary seat and close that
            # side early, which cost 5.8 points of coverage. Filling with whole
            # stones and capping afterwards puts the ground corners in the
            # crescent no rectangle can reach — the space they are FOR.
            if want_cut is True and cut0 is None:
                continue
            if want_cut is False and cut0 is not None:
                continue
            for deg, rg, ww, hh in poses[id(s)]:
                if row_h > 0.0 and hh > row_h + ROW_TOL and not allow_taller:
                    continue
                # LEVEL ROWS, both ways, against a FIXED line. Too tall was
                # always refused; too short is refused here, so a row cannot end
                # up carrying stones of different heights.
                #
                # `level` is the class chosen from stock before the row opens,
                # replaced by the opener's true height once one is seated — and
                # never by the running row_h. Measuring against row_h ratchets:
                # each stone may stand ROW_TOL taller than the row so far, that
                # raises row_h, and the next stone may stand ROW_TOL taller
                # again, so a row drifts well past its band and the earliest
                # stones end up under a 0.45 mm strip.
                if not allow_taller and level:
                    if not (level - ROW_LEVEL_TOL <= hh <= level + ROW_TOL):
                        continue
                # The binding corner is on whichever of THIS seed's edges is
                # nearer the rim, which depends on its own height, not the row's.
                y_out = max(abs(y), abs(y + hh))
                if y_out >= R:
                    continue
                chord = math.sqrt(R * R - y_out * y_out) - RIM_EPS
                # No cheap width pre-test here on purpose. Rejecting a seat
                # because the seed's BOUNDING BOX overruns the far chord throws
                # away legal placements of cut seeds: the ground-off corner means
                # the box crosses the chord while the stone itself is still
                # inside. free() tests the true outline, so let it decide.
                # Centre-out keeps each half on its own side of the middle: a
                # stone growing rightwards may nest back toward x=0 but not past
                # it, and vice versa. Left-to-right has no such stop — its stones
                # nest all the way back to whatever is already placed.
                # rim_seed places the stone FLUSH to the chord and leaves it
                # there. The cursor `x` is an edge — the right edge going left,
                # the left edge going right — so seating at the rim has to be
                # computed from the stone's own width, and it must NOT then be
                # slid back toward the middle, which is the whole point.
                #
                # The pose is normalised to the origin, so its bounds are
                # (0, 0, ww, hh) and a translate by (left, bottom) seats it.
                bw = ww
                if going_left:
                    # Right edge on the cursor, then nudged back toward the middle.
                    x_at = (-chord + bw) if rim_seed else min(x, chord)
                    g = affinity.translate(rg, x_at - bw, y)
                    if not free(g):
                        continue
                    if not rim_seed:
                        g = slide_right(g, -half if fill == "centre" else None)
                else:
                    x_at = (chord - bw) if rim_seed else max(x, -chord)
                    g = affinity.translate(rg, x_at, y)
                    if not free(g):
                        continue
                    if not rim_seed:
                        g = slide_left(g, half if fill == "centre" else None)

                # Plate actually LOST by taking this seat: the strip consumed
                # along the row, clipped to the plate, minus the seed itself.
                # Zero for a rectangle laid flush. This is what makes seeds nest:
                # two cut seeds sit flush when one is turned 180 degrees, and the
                # notch between them disappears from this measure.
                # The strip runs from the PREVIOUS seed to this one. For the seed
                # that OPENS a row there is no previous seed, and measuring from
                # the cursor (still at -R) charged it for the whole crescent
                # beside it — a cost that grows with the seed's height, which let
                # a 90-degree rotation win the opening slot and set the row 2 mm
                # taller than it needed to be. Every shorter seed in that row then
                # left a band above it.
                if going_left:
                    lo, hi = g.bounds[0], (g.bounds[2] if row_h <= 0.0 else x)
                else:
                    lo, hi = (g.bounds[0] if row_h <= 0.0 else x), g.bounds[2]
                sx0, sx1 = min(lo, hi), max(lo, hi)
                # Clipping the strip to the plate only matters when it actually
                # reaches the rim. Inside the circle its area is just width x
                # height, and skipping the clip removes one polygon intersection
                # per candidate seat — the single hottest call in the packer.
                if max(sx0 * sx0, sx1 * sx1) + max(y * y, (y + hh) ** 2) <= RSQ:
                    strip_area = (sx1 - sx0) * hh
                else:
                    strip_area = shbox(sx0, y, sx1, y + hh).intersection(disc).area
                lost = max(0.0, strip_area - g.area)
                out = _outward_score(g, cut0, deg)
                gap = max(0.0, row_h - hh) if row_h > 0.0 else 0.0
                # A seed lying flat — its shorter side as the height. Only that
                # orientation may open a row: standing one on its long side sets
                # the row taller than anything else will fill, and every later
                # seed then leaves a band above it.
                lying_flat = hh <= min(g0.bounds[2] - g0.bounds[0],
                                       g0.bounds[3] - g0.bounds[1]) + 0.05
                # A stone whose cross corner was ASSUMED is a last resort at
                # every seat, not merely later in the queue. Sorting the queue
                # only breaks ties — scan() picks the best-scoring stone for each
                # seat, so an assumed one still won seats while declared stones
                # were going spare. Leading the key with this makes the rule
                # absolute: an assumed stone is chosen only when no declared one
                # fits the seat at all.
                # A row's height is fixed by whatever opens it, and every row
                # above sits on that line — so a stone standing on its LONG side
                # opens a 13.9 mm row and knocks the whole plate out of square.
                # Measured: row heights ranged 7.8-13.9 mm before this, 8.9-9.9
                # after. Uniform rows matter more to the floor than the ~1 point
                # of coverage a tall opener occasionally buys.
                if row_h <= 0.0 and not lying_flat:
                    continue
                # Is this seat against the rim? Centre-out grows each row from
                # the middle, so the rim seats are the LAST of each half, not the
                # first — how much room is left to the chord is what says so.
                # Both halves qualify, which is how cut stones reach BOTH sides;
                # the old test (row_h <= 0) fired on the first seat of the row,
                # which under centre-out is the CENTRE, so cut stones were being
                # pushed into open plate and rejected. 2 of 56 were being used.
                _rem = (g.bounds[0] + chord) if going_left else (chord - g.bounds[2])
                _mid = abs(y + hh / 2.0) < CENTRE_BAND
                # -1 outranks a plain stone (0) outright. Merely tying is not
                # enough: a cut stone's notch counts as `lost`, so on any tie the
                # plain rectangle wins and the cross never reaches the rim.
                notch = 0.0
                if cut0 is None:
                    cutrim = 0
                elif _rem <= RIM_FLUSH and out >= CUT_OUTWARD_MIN and not _mid:
                    # WHERE a cut stone may sit is unchanged: the outermost seat
                    # of a row, outside the protected centre band, ground corner
                    # pointing outward. A cross facing a flat neighbour leaves a
                    # notch nothing can close and the floor rejects the seat.
                    #
                    # HOW it competes for that seat has changed. The notch is now
                    # priced into the seat's waste below rather than being tested
                    # against a threshold no real stone could meet, so the best
                    # orientation of the best stone wins on merit.
                    notch = _wasted_area(g, disc)
                    if notch > CUT_NOTCH_MAX:
                        continue
                    cutrim = 0
                else:
                    continue
                # EDGE UTILISATION: with cut stones allowed into the fill they
                # win rim seats outright (-1 beats a plain stone's 0). Every rim
                # seat outside the centre band then takes one — 12 of the 12
                # such seats a Ø90 plate has. The cost is ~4 points of coverage,
                # paid deliberately: a ground stone that never leaves the drawer
                # is 100% waste, and the rim is the only place it can go.
                risky = 1 if s.get("corner_assumed") else 0
                if (policy == "cut-first" and row_h <= 0.0
                        and cut0 is not None and out > 0.0 and lying_flat):
                    # ROW-END SEAT. The first seat of a row sits against the rim,
                    # and that is where a ground corner earns its keep. Cut seeds
                    # take it ahead of a rectangle so they drain a few per plate,
                    # rather than piling up until a plate holds nothing else.
                    #
                    # Height comes BEFORE the cut direction here. This seat sets
                    # the row's height, and every later seed in the row must fit
                    # under it — so an orientation standing the seed on its long
                    # side makes the whole row too tall and leaves a band above
                    # every neighbour. Ranking only on the outward cut chose
                    # exactly that: an 8.88 mm seed turned 90 degrees opened a row
                    # at 11.12 mm and cost 55 mm2 in one stripe.
                    key = (risky, -1.0, -out, round(lost, 2))
                else:
                    # Mid-row, FILLING THE ROW'S HEIGHT matters more than which
                    # way a cut points. A seed shorter than its row leaves a
                    # strip above it that no later row can reach — the horizontal
                    # bands between rows — and every orientation is available to
                    # close it, since turning a seed 90 degrees swaps which of
                    # its sides is the height. Ranking the cut direction first
                    # left those bands open for the sake of a cut that, away from
                    # the rim, has no curve to follow anyway.
                    #
                    # Under "plain-first" a plain stone also wins any tie against
                    # a cut one. A cut stone put here has its cross facing a flat
                    # neighbour, which leaves a notch nothing can fill and a seat
                    # the shop floor cannot build.
                    # Under "area", the stone that covers more wins any tie the
                    # fit and the row height leave open. Placed AFTER lost/gap so
                    # it never overrides a tighter fit or a better-filled row —
                    # it only decides between stones that were already equal, and
                    # there the wider one is simply more plate covered.
                    bulk = -g.area if seat == "area" else 0.0
                    # Charge the seat for BOTH kinds of waste it creates, as
                    # areas so they are comparable: what the stone fails to nest
                    # into along the row (`lost`), and the band it leaves above
                    # itself for being shorter than the row (`gap` x its width).
                    #
                    # Ranked separately, `gap` was a bare length placed AFTER
                    # `lost`, so any nesting difference above 0.01 mm2 outranked
                    # a band of any size — a stone 2.2 mm short of its row won
                    # the seat over one that filled it, and left a strip no later
                    # row can reach. Measured on the Ø90 reference pool: 46
                    # seats / 83.82% ranked separately, 49 / 84.95% combined,
                    # with the surviving bands under 0.55 mm instead of 2.2 mm.
                    #
                    # DEAD REMAINDER. What is left between this stone and the
                    # chord is charged to the seat when nothing in stock can fit
                    # there. Without it the packer cannot tell a stone that
                    # finishes a row from one that stops 6 mm short, because
                    # neither wastes anything BEHIND it — and a row-end gap is
                    # the most visible waste on the plate.
                    #
                    # NOTCH is the cut stone's missing corner, counted only where
                    # it falls inside the plate, and CUT_BONUS is what draining
                    # the cut pile is worth against it.
                    # Charging a cut stone for the WHOLE remainder beyond it —
                    # true in principle, since a cross closes its side — was
                    # tried and is worse: it prices crosses out of the rim seats
                    # they exist for (cut usage 6 -> 3 on the Ø90 pool) and the
                    # layout that replaces them left 9.5 mm and 24.7 mm holes,
                    # against 83.32% and one 9.08 mm gap this way.
                    _dead = _rem * hh if 0.0 < _rem < narrowest else 0.0
                    _waste = (lost + gap * bw + _dead + notch
                              - (CUT_BONUS if cut0 is not None else 0.0))
                    key = ((risky, cutrim, round(_waste, 2),
                            1 if cut0 is not None else 0, bulk, -out)
                           if policy == "plain-first"
                           else (risky, cutrim, round(_waste, 2), bulk, -out))
                if best is None or key < best[0]:
                    best = (key, s, deg, g)
        return best

    y = -R if y0 is None else y0
    guard = 0
    # The guard has to SCALE WITH THE PLATE. Each turn of this loop either seats
    # a row, advancing y by its height, or fails and advances by ROW_PROBE — so
    # the worst case is one turn per probe step across the whole diameter. A flat
    # 400 was ample at ROW_PROBE 1.0 (80 turns on a Ø80 plate); at 0.25 it is 320
    # turns for Ø80, 360 for Ø90 and 400 for Ø100, which means a large plate
    # could stop being packed part way up and come out missing its top rows.
    guard_max = int(2.0 * R / max(ROW_PROBE, 0.05)) + 50
    while y < R - 0.5 and guard < guard_max:
        guard += 1
        if all(id(s) in used for s in queue):
            break
        # CENTRE-OUT lays each row from the middle of the plate outwards, filling
        # rightwards and leftwards in turn, so a row ends ragged at BOTH rims
        # instead of flush left and ragged right. LEFT-TO-RIGHT opens each row at
        # the left chord and fills across.
        #
        # Both are packed on every plate and the better one is kept — the row
        # direction changes which chord each row is anchored to, and which wins
        # depends on the seed mix.
        xr = half if fill == "centre" else -R
        xl = -half              # only used by centre-out, growing leftwards
        row_h = 0.0
        dead = {False: False, True: False}   # has each side run dry?
        n0 = len(placed)        # first seat of THIS row, for the squeeze below
        level = None            # height class this row is built from
        anchored = False        # were the rims claimed by cut stones first?

        def seed_row_ends():
            """Seat a cut/trimmed stone at BOTH rims before the row is filled.

            Cut stones are full-size stones with one corner ground off, so they
            need a real seat, not the sliver a finished row leaves — capping
            afterwards placed 2 per plate out of 56 available. Seated first they
            take the seat they are actually good at: against the curve, where the
            ground corner follows the rim instead of wasting plate.

            The rows across the middle are skipped. That is the widest, most
            useful band and belongs to whole stones.

            Returns True if either end was seated, which suppresses the squeeze —
            a row already anchored at both rims has nothing to slide.
            """
            nonlocal row_h, xl, xr
            # OFF by default. Seating the rims before the row is filled fixes the
            # row height from a rim stone and leaves a gap between each anchor
            # and the centre fill: measured 31 seats / 52.28% against 47 / 84.17%
            # for filling whole-first. Kept because the idea is sound in
            # principle and may pay once smaller stones are stocked.
            if not CUTS_FIRST:
                return False
            if abs(y + (row_h or 9.0) / 2.0) < CENTRE_BAND:
                return False
            seated = False
            for going_left in (False, True):
                yout = max(abs(y), abs(y + (row_h or 9.0)))
                if yout >= R:
                    continue
                best = scan(y, 0.0, row_h, False, going_left,
                            want_cut=True, rim_seed=True, level=level)
                if not best:
                    continue
                _key, s, deg, g = best
                record(s, deg, g)
                row_h = max(row_h, g.bounds[3] - y)
                if going_left:
                    xl = min(xl, g.bounds[0] - clear)
                else:
                    xr = max(xr, g.bounds[2] + clear)
                seated = True
            return seated

        def swap_row(n0):
            """Trade one stone in the finished row for a different-width one so
            that a further stone fits — the move a greedy packer can never make.

            The packer commits a seat the moment it fills it and never looks
            back, so a row can end with, say, 7 mm spare while carrying a 13 mm
            stone that a narrower one would have served just as well. Swapping
            the two frees enough width for another stone, and the stone gained is
            worth more than the width given up.

            Only applied when it PAYS: width(new) > width(out) - width(in).
            Measured on a 163-stone pool: three rows qualified, +129 mm2, about
            +2.6 coverage points.
            """
            row = placed[n0:]
            if len(row) < 2 or row_h <= 0.0:
                return False
            yout = max(abs(y), abs(y + row_h))
            if yout >= R:
                return False
            chord = math.sqrt(R * R - yout * yout) - RIM_EPS
            lo = min(p["x"] for p in row)
            hi = max(p["x"] + p["w"] for p in row)
            gap = max(0.0, chord - hi) + max(0.0, lo + chord)
            if gap < 0.05:
                return False
            # CUT STONES MAY BE TRADED IN. Freeing width by swapping one stone
            # for a narrower one is exactly how a cross gets a row-end seat it
            # could not otherwise reach, and draining the cut pile is a standing
            # requirement, not a nicety.
            #
            # This was whole-only until the rebuild became safe. Two things make
            # it safe now: the rebuild is ATOMIC, so a cross that cannot be
            # seated costs nothing and the old row is restored; and every cut
            # stone it lays is checked with _cut_on_rim, so none can end up
            # inland. Before both, a rejected cross was simply dropped and the
            # row came out shorter than it started — the 10 and 15 mm holes.
            #
            # Measured, full inventory: Ø90 5 cut -> 10, Ø100 6 -> 8; on a
            # thinner band Ø100 3 -> 10 and coverage ROSE 82.25% to 84.03%.
            # Ø90 pays about 0.7 points for those five extra crosses. No gaps and
            # no inland seats appeared at any size.
            spare = [t for t in queue if id(t) not in used]

            def sides(t):
                """(width, height) for each orientation that fits this row.

                Level both ways, as in scan(): a trade that brings in a shorter
                stone buys its extra seat with a strip of bare plate along the
                whole row, and leaves a row the floor cannot build.
                """
                L, W = float(t["L"]), float(t["W"])
                return [(a, b) for a, b in ((L, W), (W, L))
                        if row_h - ROW_LEVEL_TOL <= b <= row_h]

            best = None
            for i, b in enumerate(row):
                if b.get("irregular"):
                    continue
                for a in spare:
                    for wa, _ha in sides(a):
                        freed = gap + (b["w"] - wa)
                        if freed <= 0.0:
                            continue
                        for c in spare:
                            if c is a:
                                continue
                            for wc, _hc in sides(c):
                                if wc > freed:
                                    continue
                                gain = wc - (b["w"] - wa)
                                if gain > 0.05 and (best is None or gain > best[0]):
                                    best = (gain, i, a, wa, c, wc)
            if best is None:
                return False
            _gain, idx, a, wa, c, wc = best

            # Rebuild the row: same stones, one swapped, one added, laid left to
            # right from the chord. Everything is lifted first so the strip is
            # empty and each piece can be seated without fighting its neighbours.
            order = [(srcs[n0 + k], placed[n0 + k]["w"]) for k in range(len(row))]
            order[idx] = (a, wa)
            order.append((c, wc))
            # Cut stones must finish at the ENDS of the rebuilt row. The added
            # stone goes on last, so a cross that used to be outermost would end
            # up with a neighbour beyond it — stranded against a flat face, and
            # rejected on the floor.
            _cuts = [t for t in order if t[0].get("poly")]
            _whole = [t for t in order if not t[0].get("poly")]
            if _cuts:
                order = _cuts[:1] + _whole + _cuts[1:]
            # Everything needed to put the row back exactly as it was. The
            # rebuild below can refuse any stone it is handed, so the old row has
            # to survive until the new one has proved itself.
            keep = [(placed[n0 + k], occ[n0 + k], srcs[n0 + k])
                    for k in range(len(row))]
            for k in range(len(placed) - 1, n0 - 1, -1):
                used.discard(id(srcs[k]))
                placed.pop(k)
                occ.pop(k)
                srcs.pop(k)
            reindex()

            x = -chord
            for t, w in order:
                pick = None
                for deg, rg, ww, hh in poses[id(t)]:
                    # Capped at row_h, NOT row_h + ROW_TOL. ROW_TOL is headroom
                    # for the FILL, which raises row_h as it goes and advances y
                    # from the final value. This rebuild runs after that: it never
                    # updates row_h, so a stone admitted on the tolerance finishes
                    # above the line the next row starts from. That 0.04 mm
                    # protrusion then blocks 124 mm2 of otherwise fillable plate
                    # at the next row's end, because any seat laid on that
                    # baseline catches on it.
                    if abs(ww - w) < 0.01 and hh <= row_h:
                        pick = (deg, rg)
                        break
                if pick is None:
                    continue
                deg, rg = pick
                g = affinity.translate(rg, x, y)   # pose is origin-normalised
                if not free(g):
                    continue
                # The rebuild seats stones directly, so scan()'s cut rules do not
                # apply here — re-check them. Without this a rebuilt row could
                # carry a cross facing a flat neighbour, which is exactly what
                # appeared beside seats 43 and 5.
                if t.get("poly") and not _cut_on_rim(g, R):
                    continue
                record(t, deg, g)
                x = g.bounds[2] + clear
            # ATOMIC. The whole point of the trade is to gain a seat: a stone was
            # given up for a narrower one to make room for one more. Every step
            # of the rebuild is allowed to refuse a stone, and when the one being
            # ADDED is refused the row keeps the narrower stone and gains
            # nothing — it ends shorter than it started, by exactly the width
            # that was traded away. That is how a whole-seed hole appeared at the
            # right rim of two rows on a Ø90 plate (10.66 mm and 7.86 mm wide)
            # while the seat count stayed the same, so nothing flagged it.
            #
            # Unless the row actually grew, put back what was lifted.
            if len(placed) - n0 <= len(row):
                for k in range(len(placed) - 1, n0 - 1, -1):
                    used.discard(id(srcs[k]))
                    placed.pop(k)
                    occ.pop(k)
                    srcs.pop(k)
                for p_, g_, s_ in keep:
                    used.add(id(s_))
                    placed.append(p_)
                    occ.append(g_)
                    srcs.append(s_)
                reindex()
                return False
            return True

        def cap_row():
            """Cap both ends of the FINISHED row with cut/trimmed stones.

            Run last, once the row can no longer move or grow, so a cross is
            always the outermost stone on its side and can never be stranded
            mid-row. The rows across the middle are skipped — that is the widest,
            most useful band and belongs to whole stones.
            """
            nonlocal xl, xr
            # The fill now seats cut stones at the rims itself and closes that
            # side behind them, so capping normally finds nothing left to do.
            # It still runs, because a side the fill closed for want of a WHOLE
            # stone can sometimes take a cut one, and a cross there costs
            # nothing.
            if row_h <= 0.0 or abs(y + row_h / 2.0) < CENTRE_BAND:
                return
            row = placed[n0:]
            for going_left in (False, True):
                # NEVER cap a side that already ENDS in a cut stone. Capping on
                # top of one the fill seated puts a second cross outside the
                # first, and the inner one is then stranded inland with its
                # ground corner against a flat neighbour — the seat the floor
                # rejects. It put seat 45 three stones deep in its row while
                # seat 47 sat correctly at the rim beside it.
                #
                # One mechanism or the other per side, never both.
                if row:
                    outer = (min(row, key=lambda p: p["x"]) if going_left
                             else max(row, key=lambda p: p["x"] + p["w"]))
                    if outer.get("irregular"):
                        continue
                best = scan(y, xl if going_left else xr, row_h, False,
                            going_left, want_cut=True, level=level)
                if not best:
                    continue
                _key, s2, deg, g = best
                record(s2, deg, g)
                if going_left:
                    xl = min(xl, g.bounds[0] - clear)
                else:
                    xr = max(xr, g.bounds[2] + clear)

        def fill_row():
            """Fill the current row until neither side takes another seat."""
            nonlocal row_h, xl, xr, level
            while True:
                # Level rows only. A taller seed is never admitted, even when the
                # row would otherwise end early: letting one in raises the row's
                # height, and every shorter seed already in it then sits under a
                # band that no later row can reach. Measured on a Ø80 plate that
                # traded 80.5% for 78.0% — the stretch of row left empty costs
                # less than the band.
                took = False
                for going_left in ((False, True) if fill == "centre" else (False,)):
                    # Once a side of the row has nothing left that fits, STOP
                    # scanning it. Re-scanning an exhausted side costs a full pass
                    # over the queue for every remaining seat on the other side,
                    # and that alone made a centre-out pack take 62 s against 5 s.
                    if dead[going_left]:
                        continue
                    # Cut stones compete for ordinary seats now, not just the
                    # leftovers cap_row() can reach. They are still admitted only
                    # at the outermost seat of a row with the cut facing out (see
                    # scan), so what this opens up is the chance to WIN that seat
                    # while the row is still being laid — after the fill has
                    # finished there is nothing left to win, which is why 0 of 56
                    # were being used once rows started packing to the chord.
                    best = scan(y, xl if going_left else xr, row_h, False,
                                going_left, level=level, want_cut=None)
                    if not best:
                        dead[going_left] = True
                        continue
                    _key, s, deg, g = best
                    record(s, deg, g)
                    # The stone that OPENS the row fixes the line every other
                    # stone in it must meet. The class picked from stock only
                    # narrowed the field; this is the real height.
                    if len(placed) == n0 + 1:
                        level = g.bounds[3] - y
                    # Leave the requested gap before the next seat, and before the
                    # next row. free() already enforces it against every
                    # neighbour; advancing the cursors too stops the search
                    # starting inside a gap it can never use, which would
                    # otherwise cost a slide attempt for every seed.
                    if going_left:
                        xl = min(xl, g.bounds[0] - clear)
                    else:
                        xr = max(xr, g.bounds[2] + clear)
                    row_h = max(row_h, g.bounds[3] - y)
                    took = True
                    # A cut stone CLOSES its side. It earns the seat because its
                    # ground corner follows the rim; build one more stone outside
                    # it and the cross is stranded mid-row against a flat
                    # neighbour — six such seats appeared on one plate before
                    # this rule, and the floor rejects every one.
                    if s.get("poly"):
                        dead[going_left] = True
                if not took:
                    break

        def row_gap():
            """Bare plate left at the two ENDS of this row, in mm of width."""
            if row_h <= 0.0 or len(placed) <= n0:
                return 0.0
            yout = max(abs(y), abs(y + row_h))
            if yout >= R:
                return 0.0
            ch = math.sqrt(R * R - yout * yout) - RIM_EPS
            row = placed[n0:]
            lo = min(p["x"] for p in row)
            hi = max(p["x"] + p["w"] for p in row)
            return max(lo + ch, 0.0) + max(ch - hi, 0.0)

        def lay_row(lvl):
            """Build this row from one height class. Returns the stone area seated."""
            nonlocal row_h, xl, xr, dead, level, anchored
            xr = half if fill == "centre" else -R
            xl = -half
            row_h = 0.0
            dead = {False: False, True: False}
            level = lvl
            # Cut stones claim both rims BEFORE the row is filled; whole stones
            # then fill inward between them.
            anchored = seed_row_ends()
            fill_row()
            squeeze()
            if row_h > 0.0:
                swap_row(n0)
            # Cut stones go on LAST, into the crescent whole stones cannot reach.
            cap_row()
            return sum(p["area"] for p in placed[n0:])

        def squeeze():
            """SQUEEZE: push the finished row against one rim and refill.

            Centre-out leaves a little space at BOTH ends of a row, and neither
            half is wide enough for another stone. Slid together against one side
            they often are — one extra seat per row is worth more than the
            symmetry, and the row is still a straight line either way.
            """
            nonlocal xl, xr, dead
            if not (row_h > 0.0 and fill == "centre"
                    and len(placed) > n0 and not anchored):
                return
            yout = max(abs(y), abs(y + row_h))
            if yout >= R:
                return
            chord = math.sqrt(R * R - yout * yout) - RIM_EPS
            lo = min(p["x"] for p in placed[n0:])
            hi = max(p["x"] + p["w"] for p in placed[n0:])
            lgap, rgap = lo + chord, chord - hi
            # Only worth doing when the two halves TOGETHER could seat a
            # stone that neither could alone.
            if not (min(lgap, rgap) > 0.01 and lgap + rgap >= narrowest):
                return
            shift = -lgap if lgap <= rgap else rgap
            for i in range(n0, len(placed)):
                occ[i] = affinity.translate(occ[i], shift, 0.0)
                p = placed[i]
                p["x"] += shift
                p["lx"] += shift
                p["poly"] = [(round(px + shift, 3), py) for px, py in p["poly"]]
            reindex()
            xl = min(p["x"] for p in placed[n0:]) - clear
            xr = max(p["x"] + p["w"] for p in placed[n0:]) + clear
            # Reopen both sides, EXCEPT one already closed by a cut stone —
            # refilling past it is exactly what stranded crosses in the middle
            # of a row.
            outer_l = min(placed[n0:], key=lambda p: p["x"])
            outer_r = max(placed[n0:], key=lambda p: p["x"] + p["w"])
            dead = {False: bool(outer_r.get("irregular")),
                    True: bool(outer_l.get("irregular"))}
            fill_row()

        # TRY EACH HEIGHT CLASS the stock offers for this row and keep whichever
        # covers the most plate. One class is a guess, however it is ranked: the
        # census can only count what a class HAS, not whether those stones nest
        # into this particular chord alongside what is already on the plate. A
        # class that looked best on paper left a 7.6 mm hole at the right rim of
        # one row; the runner-up filled the same row flush.
        #
        # This is the answer to "the inventory is large, so find a stone that
        # fits" — the row is not locked to the first class that looked good, it
        # is laid once per credible class and the best result kept.
        levels = pick_level(y)
        if not levels:
            break
        best_area, best_state = -1e18, None
        for lvl in levels[:LEVEL_TRIES]:
            mark_n, mark_used = len(placed), set(used)
            # A class is judged on the plate it WINS, not the stone it seats:
            # what it covers, less the bare strip it leaves at the row's ends.
            #
            # Ranking on seated area alone let a class that covered half the
            # chord with tall stones beat one that covered all of it with
            # shorter ones — on a thin pool that left a 30 mm strip at the end of
            # a row while every row stayed level, so nothing in the level rules
            # caught it. Gating on the gap instead was worse the other way: it
            # threw away a class that seated far more stone over a 7 mm end gap,
            # and cost 4 points on the full pool. Priced as area, the two are
            # comparable and neither can bully the other.
            area = lay_row(lvl) - row_gap() * row_h
            if area > best_area:
                # KEEP the winning row rather than laying it again. Re-running
                # the best class after the trials cost a fourth pass over every
                # row — a quarter of the packing time — to rebuild something
                # that had just been built.
                best_area = area
                best_state = (placed[mark_n:], occ[mark_n:], srcs[mark_n:], row_h)
            rollback(mark_n, mark_used)
        if best_state is None:
            break
        row_p, row_g, row_s, row_h = best_state
        for p_, g_, s_ in zip(row_p, row_g, row_s):
            used.add(id(s_))
            placed.append(p_)
            occ.append(g_)
            srcs.append(s_)
        reindex()
        y += (row_h + clear) if row_h > 0 else ROW_PROBE

    # ---- RECONSIDER: revisit finished rows and improve what greedy committed --
    def _row_groups():
        """placed indices grouped by row.

        Clustered on the baseline the stones actually sit on, not on `y` rounded
        to the millimetre. Rounding put stones from two different baselines in
        one group — a sweep-up seat sits up to SWEEP_ROW_TOL off the line — and
        the rebuild then re-laid the whole group at the LOWEST of them, which
        drove it into the row below: 158 of 217 rebuild placements were refused
        for want of space, and not one exchange survived.
        """
        rows, cur = [], []
        for i in sorted(range(len(placed)), key=lambda i: placed[i]["y"]):
            if cur and placed[i]["y"] - placed[cur[0]]["y"] > SWEEP_ROW_TOL:
                rows.append(cur)
                cur = []
            cur.append(i)
        if cur:
            rows.append(cur)
        return [r for r in rows if len(r) >= 2]

    def _reconsider_row(idxs):
        """One exchange on one row: take a stone out, put a different width in,
        and seat one more. Applied only when the stone gained is wider than the
        width given up, so the row can only improve."""
        row = [placed[i] for i in idxs]
        rh = max(p["h"] for p in row)
        ry = min(p["y"] for p in row)
        yout = max(abs(ry), abs(ry + rh))
        if yout >= R:
            return False
        chord = math.sqrt(R * R - yout * yout) - RIM_EPS
        lo = min(p["x"] for p in row)
        hi = max(p["x"] + p["w"] for p in row)
        gap = max(0.0, chord - hi) + max(0.0, lo + chord)
        # Whole stones only, as in swap_row: a cut stone re-laid mid-row has no
        # rim to face, and the notch check would drop it on the way back in.
        spare = [t for t in queue if id(t) not in used and not t.get("poly")]

        # The row's own height line, taken from the stones already in it, so a
        # reconsidered row stays as level as the fill left it.
        rlvl = min(p["h"] for p in row)

        def sides(t):
            # Capped at the row's OWN height, not at height + ROW_TOL. A
            # finished row has the next row laid flush on top of it, so the
            # vertical space is exactly what the row already occupies: a
            # replacement even a fraction taller cannot be seated at all. That
            # slack is why 136 of 196 rebuild placements were refused and no
            # exchange ever survived.
            L, W = float(t["L"]), float(t["W"])
            return [(a, b) for a, b in ((L, W), (W, L))
                    if rlvl - ROW_LEVEL_TOL <= b <= rh]

        best = None
        for k, b in enumerate(row):
            if b.get("irregular"):
                continue
            for a in spare:
                for wa, _h in sides(a):
                    freed = gap + (b["w"] - wa)
                    if freed <= 0.0:
                        continue
                    for c in spare:
                        if c is a:
                            continue
                        for wc, _h2 in sides(c):
                            if wc > freed:
                                continue
                            gain = wc - (b["w"] - wa)
                            if gain > 0.05 and (best is None or gain > best[0]):
                                best = (gain, k, a, wa, c, wc)
        if best is None:
            return False
        _g, k, a, wa, c, wc = best
        order = [(srcs[i], placed[i]["w"]) for i in idxs]
        order[k] = (a, wa)
        order.append((c, wc))
        cuts = [t for t in order if t[0].get("poly")]
        whole = [t for t in order if not t[0].get("poly")]
        if cuts:
            order = cuts[:1] + whole + cuts[1:]

        keep = [(placed[i], occ[i], srcs[i]) for i in idxs]
        for i in sorted(idxs, reverse=True):
            used.discard(id(srcs[i]))
            placed.pop(i)
            occ.pop(i)
            srcs.pop(i)
        reindex()

        def restore():
            """Put the lifted row back exactly as it was."""
            while len(placed) > n_before:
                used.discard(id(srcs[-1]))
                placed.pop()
                occ.pop()
                srcs.pop()
            for p_, g_, s_ in keep:
                used.add(id(s_))
                placed.append(p_)
                occ.append(g_)
                srcs.append(s_)
            reindex()
            return False

        n_before = len(placed)
        # RE-LAY, NESTED. Each stone is laid against the chord and then slid back
        # against its neighbour, exactly as fill_row() does.
        #
        # Abutting bounding boxes instead is why this pass never kept a single
        # exchange: the fill NESTS its stones, so a row re-laid box-to-box comes
        # out wider than the one it replaced, the last stone falls outside the
        # chord, and the atomic check rolls the whole thing back. 37 exchanges
        # were attempted on one plate and 0 survived.
        x = -chord
        seated = 0
        for t, w in order:
            pick = None
            for deg, rg, ww, hh in poses[id(t)]:
                if abs(ww - w) < 0.01 and hh <= rh + 1e-9:
                    pick = (deg, rg)
                    break
            if pick is None:
                continue
            deg, rg = pick
            g = affinity.translate(rg, x, ry)     # pose is origin-normalised
            if not free(g):
                continue
            g = slide_left(g)
            if t.get("poly") and not _cut_on_rim(g, R):
                continue
            record(t, deg, g)
            seated += 1
            x = g.bounds[2] + clear
        # A rebuild that did not actually GAIN a seat has made the plate worse;
        # put the original row back rather than keep the damage.
        if seated <= len(idxs):
            return restore()

        # BALANCED. The rebuilt row is nested tight against the left chord, so it
        # now has all of its slack on the right — slide it back to the middle.
        #
        # Leaving it left-flush is what made this pass unusable the first time: a
        # row the fill had balanced centre-out came back against the left rim, so
        # an improved row left a bare block down the right of the plate and the
        # whole plate read as shoved sideways. Every invariant passed and the
        # seat count went UP, and it was still the wrong plate.
        new = placed[n_before:]
        lo2 = min(p["x"] for p in new)
        hi2 = max(p["x"] + p["w"] for p in new)
        shift = ((chord - hi2) - (lo2 + chord)) / 2.0
        if abs(shift) > 0.005:
            # Lift the row out of the index first, or every stone in it would be
            # tested against where it used to be and refuse to move.
            row_geoms = occ[n_before:]
            moved = [affinity.translate(g, shift, 0.0) for g in row_geoms]
            del occ[n_before:]
            reindex()
            ok = all(free(g) for g in moved)
            occ.extend(moved if ok else row_geoms)
            reindex()
            if ok:
                for i in range(n_before, len(placed)):
                    p = placed[i]
                    p["x"] += shift
                    p["lx"] += shift
                    p["poly"] = [(round(px + shift, 3), py)
                                 for px, py in p["poly"]]
        return True

    if RECONSIDER:
        for _round in range(RECONSIDER_ROUNDS):
            changed = False
            tried = set()
            # The groups MUST be recomputed for every row. _reconsider_row pops
            # the row it is working on out of `placed` and appends the result —
            # and even a rolled-back rebuild puts the old stones back at the END
            # — so every index after the first row processed is stale. Taking
            # the groups once and looping over them therefore lifted stones from
            # the wrong rows, which left same-row neighbours standing in the way
            # of the rebuild: 140 of 204 placements refused, and no exchange
            # could ever survive. Rows are identified by their baseline instead,
            # so each is visited exactly once per round however `placed` moves.
            while True:
                groups = [g for g in _row_groups()
                          if round(placed[g[0]]["y"], 1) not in tried]
                if not groups:
                    break
                g = groups[0]
                tried.add(round(placed[g[0]]["y"], 1))
                if _reconsider_row(g):
                    changed = True
            if not changed:
                break

    # ---- SWEEP-UP: whole seeds into the pockets the row sweep walked past -----
    # The row sweep fills strictly left to right and never returns, and a row
    # ends the moment no remaining seed fits its height — so the rest of that
    # row is abandoned even where the circle is still wide open. That left space
    # a leftover seed would have fitted: on a Ø80 plate, six of six leftovers
    # still had a legal position, all in the same abandoned stretch.
    #
    # This pass only ever ADDS a seed to a spot that is already proven free, so
    # it cannot move or displace anything the rows placed — coverage can only go
    # up. Largest-first, so a big pocket is not wasted on a small seed.
    # Searched POCKET BY POCKET, not over a blanket grid. A seed dropped into a
    # gap comes to rest against its corners, so the pocket's own bounding-box
    # corners and outline vertices are the only anchors worth testing — a 1 mm
    # sweep of the whole plate per leftover seed took 38 s a plate for the same
    # result.
    smallest = min((shapes[id(s)].area for s in queue), default=0.0)
    # Sweep-up drops a stone wherever a pocket will take it, with no reference to
    # the rows — which is how a stone ends up sitting proud of its row at the
    # plate edge (always the highest seat number, because it is placed last).
    # Restrict it to the row baselines already established: a leftover seat is
    # only worth having if it sits ON a row.
    row_ys = sorted({round(p["y"], 1) for p in placed})
    gap_region = disc.difference(unary_union(occ)) if occ else disc
    for _ in range(len(queue)):
        if gap_region.is_empty:
            break
        pockets = [g for g in (list(gap_region.geoms)
                               if gap_region.geom_type != "Polygon" else [gap_region])
                   if g.geom_type == "Polygon" and g.area >= smallest]
        pockets.sort(key=lambda g: -g.area)
        # WHOLE stones only. This pass drops a leftover into any pocket that
        # will take it, with no notion of rims — so a cut stone landed wherever
        # it happened to fit, 17.30 mm inland on a Ø100 plate with its ground
        # corner facing a flat neighbour. Cut stones reach the plate through the
        # fill, cap_row and the row-end pass, every one of which holds them to
        # RIM_FLUSH of the boundary. They must not arrive by this door.
        spare = sorted((s for s in queue
                        if id(s) not in used and not s.get("poly")),
                       key=lambda s: -shapes[id(s)].area)
        chosen = None
        for part in pockets:
            px0, py0, px1, py1 = part.bounds
            pw, ph = px1 - px0, py1 - py0
            # A pocket that snakes between seeds carries hundreds of vertices,
            # and probing every one dominated the runtime. Simplify first and cap
            # the list: a seed comes to rest against a corner, and the corners
            # that matter survive simplification.
            corners = list(part.simplify(0.4).exterior.coords)[:-1][:SWEEP_ANCHORS]
            for s in spare:
                g0 = shapes[id(s)]
                if g0.area > part.area:
                    continue                   # cannot fit in this pocket at all
                for deg, g, w, h in poses[id(s)]:
                    if w > pw + 1e-9 or h > ph + 1e-9:
                        continue
                    gb = g.bounds
                    anchors = [(px0, py0), (px1 - w, py0), (px0, py1 - h), (px1 - w, py1 - h)]
                    anchors += [(vx, vy) for vx, vy in corners]
                    anchors += [(vx - w, vy) for vx, vy in corners]
                    for ax, ay in anchors:
                        # The anchor IS the seat's bottom edge, so the row test
                        # needs no geometry: asking the translated polygon for
                        # its bounds cost 342k shapely calls a plate, plus the
                        # translate itself for anchors that were never going to
                        # be used.
                        if row_ys and min(abs(ay - ry)
                                          for ry in row_ys) > SWEEP_ROW_TOL:
                            continue           # off the row line — skip it
                        # And it must not stand PROUD of the row it is joining.
                        # This pass never checked height, so it could seat a
                        # stone whose top finished 0.03 mm above the line the
                        # next row started from. That hairline is invisible and
                        # ruinous: any later seat laid on that baseline catches
                        # on it, and 219 mm2 of good space on a Ø100 plate became
                        # unreachable because of it.
                        _above = [ry for ry in row_ys if ry > ay + 0.05]
                        if _above and ay + h > min(_above) + 1e-9:
                            continue
                        cand = affinity.translate(g, ax - gb[0], ay - gb[1])
                        if free(cand):
                            chosen = (s, deg, slide_left(cand))
                            break
                    if chosen:
                        break
                if chosen:
                    break
            if chosen:
                break
        if not chosen:
            break
        record(*chosen)
        # Subtract just the seed that was placed. Rebuilding the free region from
        # the union of every seed each round was the single largest cost.
        gap_region = gap_region.difference(
            chosen[2].buffer(clear) if clear > 0.0 else chosen[2])

    # Coverage from the UNION of what is on the plate. Summing areas would
    # double-count anything that overlapped and could report over 100%.
    covered = unary_union([ShPoly(p["poly"]).buffer(0) for p in placed]).area if placed else 0.0
    return placed, 100 * covered / (math.pi * R * R)

def _fill_remaining_space(real, placed, fill, R):
    """
    Residual-space real-seed optimizer.

    This is a FINAL ADDITIVE pass. It does not move, remove, or rebuild any
    seed selected by the main Max-Coverage optimizer.

    Purpose:
        Find physically usable areas that remain after the normal row-based
        Max-Coverage packing and place additional UNUSED REAL seeds there.

    Key difference from the main row optimizer:
        The main optimizer keeps rows level. This pass does NOT require the
        additional seed to match an existing row height. A smaller seed is
        therefore allowed to occupy a physically valid pocket below/above a
        larger row.

    Rules:
        - Existing seeds never move.
        - Only unused real inventory is considered.
        - Smaller seeds are deliberately considered first.
        - Every ENHANCED_ANGLES orientation is tested.
        - Whole rectangular and irregular/cut seed geometry is respected.
        - CLEARANCE is respected.
        - Cut/irregular seeds are allowed only when they satisfy _cut_on_rim().
        - A seed is committed only when it produces additional coverage.
        - The free area is updated after every accepted placement.
        - Search is pocket/anchor based rather than a full millimetre grid,
          so the residual pass remains bounded for large inventories.
    """
    if not real:
        return placed, fill

    if not placed:
        return placed, fill

    from shapely import affinity
    from shapely.geometry import Point, Polygon as ShPoly
    from shapely.ops import unary_union

    # ---------------------------------------------------------------
    # Plate and clearance
    # ---------------------------------------------------------------
    disc = Point(0.0, 0.0).buffer(
        R,
        resolution=180,
    )

    clear = max(
        0.0,
        float(
            getattr(
                P,
                "CLEARANCE",
                0.0,
            ) or 0.0
        ),
    )

    # ---------------------------------------------------------------
    # Current occupied geometry
    # ---------------------------------------------------------------
    occupied_geoms = []

    for p in placed:
        poly = p.get("poly")
        if not poly:
            continue

        try:
            g = ShPoly(poly).buffer(0)
        except Exception:
            continue

        if not g.is_empty:
            occupied_geoms.append(g)

    if not occupied_geoms:
        return placed, fill

    occupied = unary_union(
        occupied_geoms
    )

    # This is the actual free area available to another seed.
    #
    # CLEARANCE is applied around existing seeds so a new seed cannot be
    # inserted closer than the user's requested distance.
    blocked = (
        occupied.buffer(clear)
        if clear > 0.0
        else occupied
    )

    gap_region = disc.difference(blocked)

    if gap_region.is_empty:
        return placed, fill

    # ---------------------------------------------------------------
    # Inventory already used on this plate
    # ---------------------------------------------------------------
    used_stocks = {
        str(p.get("stock", "")).strip()
        for p in placed
        if p.get("stock") is not None
    }

    spare = [
        s
        for s in real
        if str(s.get("stock", "")).strip()
        not in used_stocks
    ]

    if not spare:
        return placed, fill

    # ---------------------------------------------------------------
    # Build all four orientations once.
    #
    # We deliberately sort the candidate seeds SMALL-to-LARGE because
    # this pass exists specifically to exploit small seeds that the row
    # optimizer could not use.
    # ---------------------------------------------------------------
    poses = []

    for s in spare:
        g0 = _seed_footprint(s)

        if g0 is None or g0.is_empty:
            continue

        try:
            base_cut = _cut_direction(g0)
        except Exception:
            base_cut = None

        seed_area = float(
            getattr(g0, "area", 0.0)
        )

        if seed_area <= 0.0:
            continue

        for deg in ENHANCED_ANGLES:
            try:
                rg = (
                    affinity.rotate(
                        g0,
                        deg,
                        origin="centroid",
                    )
                    if deg
                    else g0
                )

                if rg.is_empty:
                    continue

                bx0, by0, bx1, by1 = rg.bounds

                # Normalize the pose so its bounding-box lower-left is (0,0).
                rg = affinity.translate(
                    rg,
                    -bx0,
                    -by0,
                )

                poses.append({
                    "seed": s,
                    "stock": str(
                        s.get("stock", "")
                    ).strip(),
                    "angle": int(deg),
                    "geom": rg,
                    "w": bx1 - bx0,
                    "h": by1 - by0,
                    "area": seed_area,
                    "cut": base_cut,
                })

            except Exception:
                continue

    if not poses:
        return placed, fill

    # Smallest first is intentional.
    poses.sort(
        key=lambda x: (
            x["area"],
            max(x["w"], x["h"]),
            min(x["w"], x["h"]),
        )
    )

    # ---------------------------------------------------------------
    # Candidate-pocket controls
    #
    # These prevent an inventory of 1600+ seeds from causing a huge
    # full-plate x/y grid search.
    # ---------------------------------------------------------------
    MAX_POCKETS = 20
    MAX_VERTICES = 28
    MAX_ROUNDS = 120

    # A candidate is allowed a tiny numerical tolerance.
    EPS = 1e-7

    def _polygon_parts(region):
        if region.is_empty:
            return []

        if region.geom_type == "Polygon":
            return [region]

        if hasattr(region, "geoms"):
            return [
                g
                for g in region.geoms
                if g.geom_type == "Polygon"
                and not g.is_empty
            ]

        return []

    def _candidate_anchors(part, w, h):
        """
        Generate a bounded set of useful candidate lower-left positions.

        A rectangle/seed that improves a pocket normally comes to rest against
        one or more pocket edges/corners. We therefore test:
            - four pocket bounding-box corners
            - pocket centroid/representative point
            - simplified polygon vertices
            - offsets from simplified vertices

        No full plate grid is used.
        """
        bx0, by0, bx1, by1 = part.bounds

        anchors = [
            # Bounding-box corners.
            (bx0, by0),
            (bx1 - w, by0),
            (bx0, by1 - h),
            (bx1 - w, by1 - h),

            # Bounding-box centre.
            (
                (bx0 + bx1 - w) / 2.0,
                (by0 + by1 - h) / 2.0,
            ),
        ]

        # Representative point is often better than centroid for concave
        # pockets because it is guaranteed to be inside the polygon.
        try:
            rp = part.representative_point()
            anchors.extend([
                (rp.x, rp.y),
                (rp.x - w / 2.0, rp.y - h / 2.0),
            ])
        except Exception:
            pass

        # Simplify the pocket before taking vertices. This is the same basic
        # strategy used by the existing sweep-up implementation.
        try:
            simple = part.simplify(
                0.25,
                preserve_topology=True,
            )

            if simple.geom_type == "Polygon":
                vertices = list(
                    simple.exterior.coords
                )[:-1]
            else:
                vertices = []

            if vertices:
                if len(vertices) > MAX_VERTICES:
                    step = max(
                        1,
                        len(vertices) // MAX_VERTICES,
                    )
                    vertices = vertices[::step][
                        :MAX_VERTICES
                    ]

                for vx, vy in vertices:
                    anchors.extend([
                        # Treat the vertex as each possible corner.
                        (vx, vy),
                        (vx - w, vy),
                        (vx, vy - h),
                        (vx - w, vy - h),

                        # Small inward offsets help when exact vertex
                        # contact is rejected by numerical geometry.
                        (
                            vx + EPS,
                            vy + EPS,
                        ),
                        (
                            vx - w - EPS,
                            vy + EPS,
                        ),
                        (
                            vx + EPS,
                            vy - h - EPS,
                        ),
                    ])

        except Exception:
            pass

        # Remove duplicate anchors.
        unique = []
        seen = set()

        for ax, ay in anchors:
            key = (
                round(ax, 4),
                round(ay, 4),
            )

            if key in seen:
                continue

            seen.add(key)
            unique.append(
                (float(ax), float(ay))
            )

        return unique

    def _place_pose_in_anchor(pose, ax, ay):
        """
        Translate a normalized pose so its bounding-box lower-left is the
        requested anchor.
        """
        try:
            return affinity.translate(
                pose["geom"],
                ax,
                ay,
            )
        except Exception:
            return None

    def _is_legal(candidate, pose):
        """
        Full physical validity check.
        """
        if candidate is None or candidate.is_empty:
            return False

        # Must be fully inside the usable circle.
        try:
            if not disc.covers(candidate):
                return False
        except Exception:
            return False

        # No overlap with an already placed seed.
        try:
            if clear > 0.0:
                if candidate.distance(
                    occupied
                ) < clear - 1e-9:
                    return False
            else:
                if (
                    candidate.intersection(
                        occupied
                    ).area
                    > EPS
                ):
                    return False
        except Exception:
            return False

        # Irregular/cut seed:
        # preserve the existing rule that the missing corner belongs at
        # the plate rim, not in the middle of the plate.
        if pose["cut"] is not None:
            try:
                if not _cut_on_rim(
                    candidate,
                    R,
                ):
                    return False
            except Exception:
                return False

        return True

    def _placement_record(pose, candidate):
        """
        Convert the Shapely candidate to the same dictionary shape used by
        the existing Max-Coverage renderer/Excel/database code.
        """
        bx0, by0, bx1, by1 = candidate.bounds

        try:
            rp = candidate.representative_point()
        except Exception:
            rp = Point(
                (bx0 + bx1) / 2.0,
                (by0 + by1) / 2.0,
            )

        return {
            "stock": pose["seed"]["stock"],
            "cts": pose["seed"].get(
                "cts",
                0.0,
            ),

            # Actual placed bounding dimensions.
            "L": round(
                bx1 - bx0,
                1,
            ),
            "W": round(
                by1 - by0,
                1,
            ),

            "H": pose["seed"].get(
                "H",
                round(
                    (
                        P.T_LO
                        + P.T_HI
                    ) / 2.0,
                    3,
                ),
            ),

            # Original imported dimensions.
            "rawL": pose["seed"].get("L"),
            "rawW": pose["seed"].get("W"),

            "x": bx0,
            "y": by0,
            "w": bx1 - bx0,
            "h": by1 - by0,

            # IMPORTANT:
            # This preserves the exact rotation selected by the residual pass.
            "angle": pose["angle"],

            "kind": "real",

            "poly": [
                (
                    round(px, 3),
                    round(py, 3),
                )
                for px, py in candidate.exterior.coords
            ],

            "area": candidate.area,

            "irregular": bool(
                pose["seed"].get("poly")
            ),

            "lx": rp.x,
            "ly": rp.y,
        }

    # ---------------------------------------------------------------
    # Residual optimization loop
    # ---------------------------------------------------------------
    placed_now = list(placed)
    fill_now = float(fill)

    for _round in range(
        MAX_ROUNDS
    ):
        if gap_region.is_empty:
            break

        pockets = [
            g
            for g in _polygon_parts(
                gap_region
            )
            if g.area > 0.01
        ]

        if not pockets:
            break

        # Largest pockets first. Within each pocket, SMALL seeds are
        # considered first.
        pockets.sort(
            key=lambda g: -g.area
        )

        pockets = pockets[
            :MAX_POCKETS
        ]

        best = None

        for pocket in pockets:
            p0, p1, p2, p3 = pocket.bounds
            pw = p2 - p0
            ph = p3 - p1

            if pw <= 0.0 or ph <= 0.0:
                continue

            pocket_area = float(
                pocket.area
            )

            # -------------------------------------------------------
            # Try unused small seeds first.
            # -------------------------------------------------------
            for pose in poses:
                stock = pose["stock"]

                if stock in used_stocks:
                    continue

                w = pose["w"]
                h = pose["h"]

                # A bounding-box test is only a cheap filter.
                # The exact Shapely test follows.
                if w > pw + 1e-6:
                    continue

                if h > ph + 1e-6:
                    continue

                if pose["area"] > pocket_area + 1e-6:
                    continue

                anchors = _candidate_anchors(
                    pocket,
                    w,
                    h,
                )

                for ax, ay in anchors:
                    candidate = _place_pose_in_anchor(
                        pose,
                        ax,
                        ay,
                    )

                    if not _is_legal(
                        candidate,
                        pose,
                    ):
                        continue

                    gain = float(
                        candidate.area
                    )

                    if gain <= 0.01:
                        continue

                    # ------------------------------------------------
                    # Score the candidate.
                    #
                    # We want to use the small stones, but we should
                    # still prefer the candidate that provides the
                    # best useful coverage when several small seeds
                    # can fit.
                    #
                    # Primary:
                    #   additional covered area
                    #
                    # Secondary:
                    #   how much of the candidate's bounding box is
                    #   actually used by the pocket geometry
                    #
                    # Tertiary:
                    #   smaller seed size, so small leftover pockets
                    #   are actively exploited.
                    # ------------------------------------------------
                    try:
                        pocket_overlap = (
                            candidate.intersection(
                                pocket
                            ).area
                        )
                    except Exception:
                        pocket_overlap = gain

                    if pocket_overlap <= 0.01:
                        continue

                    # Higher = better.
                    #
                    # Area remains dominant. The small-size bonus is
                    # deliberately tiny so we never sacrifice meaningful
                    # coverage merely to use a smaller seed.
                    compactness = (
                        pocket_overlap
                        / max(
                            gain,
                            1e-9,
                        )
                    )

                    small_bonus = 1.0 / max(
                        pose["area"],
                        1.0,
                    )

                    score = (
                        gain
                        + 0.01
                        * compactness
                        + 0.05
                        * small_bonus
                    )

                    candidate_info = {
                        "score": score,
                        "gain": gain,
                        "pose": pose,
                        "geometry": candidate,
                    }

                    if (
                        best is None
                        or score
                        > best["score"]
                    ):
                        best = candidate_info

        # Nothing in the remaining pockets can be added.
        if best is None:
            break

        # -----------------------------------------------------------
        # Commit the best new real seed.
        # -----------------------------------------------------------
        pose = best["pose"]
        candidate = best["geometry"]

        new_seed = _placement_record(
            pose,
            candidate,
        )

        placed_now.append(
            new_seed
        )

        stock = pose["stock"]

        used_stocks.add(
            stock
        )

        # Update occupied geometry.
        if clear > 0.0:
            block_for_gap = candidate.buffer(
                clear
            )
        else:
            block_for_gap = candidate

        occupied = unary_union([
            occupied,
            block_for_gap,
        ])

        # Recalculate the available region only from the newly placed
        # seed. This avoids rebuilding the union of every seed every round.
        gap_region = gap_region.difference(
            block_for_gap
        )

        # Recalculate actual coverage from the placed seed outlines.
        #
        # UNION is essential; summing areas could double-count if numerical
        # geometry ever produced a tiny overlap.
        try:
            covered = unary_union([
                ShPoly(
                    p["poly"]
                ).buffer(0)
                for p in placed_now
                if p.get("poly")
            ]).area

            fill_now = (
                100.0
                * covered
                / (
                    math.pi
                    * R
                    * R
                )
            )
        except Exception:
            # Keep the previous valid coverage if a pathological polygon
            # prevents the reporting union.
            pass

    return placed_now, fill_now


def enhanced_plate_job(args):
    """MAX COVERAGE — pack the plate at several row phases and keep the best.

    A single pack is one greedy sweep, and a greedy sweep is only ever as good as
    where it starts. The baseline of the first row decides which chords all the
    rows above it land on, and that alone swung the reference plate from 76.3% to
    84.0%. Rather than betting on one phase, this packs the plate once per phase
    in ROW_PHASES, refines either side of the winner, and renders the best plate.

    Both cut-seed policies are tried and the better BUILDABLE result is kept —
    coverage on its own would pick a layout whose seats the shop floor rejects.

    Every run is independent and non-destructive — nothing is carried between
    them — so the result can only be at least as good as the single run this
    replaces, at the cost of packing the plate a handful of times.

    `args` = (real, pi, plate_d, R, min_size, path).
    """
    real, pi, plate_d, R, min_size, path = args
    if not real:
        render_enhanced_circle([], [], pi, R, 0.0, path)
        return (pi, [], 0.0, round(2 * R, 4), 0.0)

    from shapely.geometry import Point
    disc = Point(0.0, 0.0).buffer(R, resolution=180)

    best = {"placed": None, "fill": -1.0, "score": -1.0, "phase": 0.0,
            "policy": CUT_POLICIES[0], "dir": FILL_DIRECTIONS[0],
            "seat": SEAT_SCORES[0], "waste": 0.0, "nwaste": 0}
    tried = set()

    def consider(ph, policy, direction, seat, result):
        """Score one finished pack and keep it if it beats the running best.

        Split out of `attempt` so a batch of packs computed in parallel is
        scored here, in the parent, in exactly the order the sequential sweep
        would have scored them. `score > best` is strict, so order decides ties
        and must not change.
        """
        pl, f = result
        waste, nw, stuck = _interior_waste(pl, disc)
        # SCORE = area that can actually be BUILT. Coverage on its own prefers
        # the layout that squeezes in more stones by putting crosses against flat
        # neighbours — seats the shop floor has to reject. A rejected seat costs
        # both the notch beside it and the stone sitting in it, so both come off.
        score = (f / 100.0) * math.pi * R * R - waste - stuck
        if score > best["score"]:
            best.update(placed=pl, fill=f, score=score, phase=ph,
                        policy=policy, waste=waste, nwaste=nw)
            best["dir"] = direction
            best["seat"] = seat

    # EVERY policy gets EVERY phase. It is tempting to sweep the phases once and
    # test the other policies only near the winner — on one pool both policies
    # peaked at the same phase, so it looked free. It is not. On a 90-seed pool
    # they peak in different places, and probing the second policy at the first
    # one's phase found 82.85% with an unbuildable seat when that policy's own
    # best was 83.50% with none. The pruned search returned a plate that was
    # neither policy's best. Coverage is worth more than the seconds saved.
    #
    # These combinations share no state, so they are computed ACROSS CORES and
    # scored here in the parent in the original order. Same packs, same scoring,
    # same tie-breaking — only the wall-clock changes. Measured on a 202-stone
    # Ø90 pool: 46.8 s sequential against 13.5 s over 8 cores, with an identical
    # winning fill of 84.43%. (Threads were tried and are WORSE than sequential,
    # 68.5 s: two fifths of a pack is shapely's Python-level wrapper code, which
    # holds the GIL.)
    combos = [(ph, policy, direction, seat)
              for direction in FILL_DIRECTIONS
              for seat in SEAT_SCORES
              for policy in CUT_POLICIES
              for ph in ROW_PHASES]
    for (ph, policy, direction, seat), result in zip(
            combos, _parallel.run_packs(args, [(-R + ph, pol, d, st)
                                               for ph, pol, d, st in combos])):
        key = (round(ph, 3), policy, direction, seat)
        if key in tried or ph < 0.0:
            continue
        tried.add(key)
        consider(ph, policy, direction, seat, result)

    # Refine around the winner — the coarse step is wider than the difference
    # between a good phase and the best one. Only the winning combination is
    # refined; re-sweeping the losers costs a pack each for nothing.
    #
    # SEQUENTIAL on purpose, unlike the sweep above: each refine reads
    # best["phase"], which the previous one may just have moved, so they are a
    # dependent chain rather than independent work.
    for d in (-ROW_PHASE_REFINE, ROW_PHASE_REFINE, -ROW_PHASE_REFINE / 2,
              ROW_PHASE_REFINE / 2):
        ph = best["phase"] + d
        policy, direction, seat = best["policy"], best["dir"], best["seat"]
        key = (round(ph, 3), policy, direction, seat)
        if key in tried or ph < 0.0:
            continue
        tried.add(key)
        consider(ph, policy, direction, seat,
                 _pack_once(args, -R + ph, policy, direction, seat))

    placed, fill = best["placed"] or [], max(0.0, best["fill"])

    # Existing final row-end improvement.
    placed, fill = _fill_row_ends(real, placed, fill, R)

    # Residual pocket search. OFF by default — see RESIDUAL_FILL for the
    # measurements: it is the largest single cost in a run AND the source of the
    # scattered rim stones the floor rejected.
    if RESIDUAL_FILL:
        placed, fill = _fill_remaining_space(real, placed, fill, R)

    # Rim pockets: the crescent left between the outermost stones and the edge.
    # Additive, attached-only — see _fill_rim_pockets.
    placed, fill = _fill_rim_pockets(real, placed, fill, R)

    render_enhanced_circle(placed, placed, pi, R, fill, path)
    return (pi, placed, round(fill, 4), round(2 * R, 4), 0.0)




def _fill_row_ends(real, placed, fill, R):
    """Drop leftover stones into the corners the rows could not reach.

    Runs ONCE, on the winning plate, and only ADDS: no stone already on the
    plate is moved, swapped or removed, so the arrangement the search chose is
    exactly the arrangement that ships.

    The space it goes after is at the ENDS of the top and bottom rows. Those
    bands are trapezoids — on a Ø80 plate the top row spans ±14.60 mm at its top
    edge but ±28.95 mm at its base — so a stone SHORTER than the row reaches into
    the corner even though the row itself cannot be extended. The stone sits
    lower than its neighbours as a result; that step is the price of the seat and
    is why this is a deliberate last pass rather than part of the fill.

    The row sweep will not find these seats: it holds every stone in a row to one
    height class, which is what keeps rows level. The sweep-up pass will not
    either — it probes pocket CORNERS, and these seats are only reachable by
    sliding along the row's baseline.

    Cut stones are still held to the rim rule, so a cross may only land here if
    it is genuinely against the curve.
    """
    if not placed or not real:
        return placed, fill

    from shapely import affinity
    from shapely.geometry import Point, Polygon
    from shapely.ops import unary_union

    disc = Point(0.0, 0.0).buffer(R, resolution=180)
    # ONE PREPARED UNION, not a per-candidate spatial index.
    #
    # Replacing this with an STRtree over the individual stones was tried and is
    # MEASURABLY WORSE: 844.8 s -> 1433.9 s on the reference plate, for
    # byte-identical output. GEOS already indexes a MultiPolygon internally, so
    # `g.intersects(taken)` is one cheap native call, while the index version
    # pays a Python-level box construction, a query returning a numpy array, and
    # a Python loop with two shapely calls per neighbour returned. At the handful
    # of neighbours a seat actually has, that overhead dwarfs what it saves.
    taken = unary_union([Polygon(p["poly"]).buffer(0) for p in placed])
    # "Distance between seeds" from the criteria form applies to THESE seats too.
    # Without it this pass seated stones against their neighbours whatever the
    # form asked for, and a run at 1.5 mm came back with pairs 0.0 mm apart.
    clear = max(0.0, float(getattr(P, "CLEARANCE", 0.0) or 0.0))
    used = {p["stock"] for p in placed}
    spare = [s for s in real if s["stock"] not in used]
    if not spare:
        return placed, fill

    # Collapse candidates that are the SAME SIZE. The leftover inventory holds
    # hundreds of stones, and probing each separately rediscovers the same answer
    # — every row is re-scanned against every pose on every round, so this loop
    # is where the pass spends its time.
    #
    # Matched EXACTLY, never to a tolerance. Bucketing by ~0.1 mm places the
    # representative's outline and then labels it with a different stone's
    # number, so the plate would show a seat the chosen stone does not have.
    # Cut stones are not collapsed at all: their outlines differ in ways a
    # bounding box cannot capture.
    _shapes = {}
    for s in spare:
        g0 = _seed_footprint(s)
        if g0 is None:
            continue
        cut0 = _cut_direction(g0)
        is_cut = cut0 is not None
        for deg in ENHANCED_ANGLES:
            rg = affinity.rotate(g0, deg, origin="centroid") if deg else g0
            b = rg.bounds
            w, h = b[2] - b[0], b[3] - b[1]
            key = (s["stock"] if is_cut else (round(w, 6), round(h, 6)), deg, is_cut)
            entry = _shapes.get(key)
            if entry is None:
                _shapes[key] = [deg, is_cut, affinity.translate(rg, -b[0], -b[1]),
                                w, h, [s]]
            else:
                entry[5].append(s)
    if not _shapes:
        return placed, fill
    # (supply, deg, is_cut, geom, w, h) — `supply` is every stone that can fill
    # this shape; the stock number is drawn from it only once a seat is won.
    poses = [(e[5], e[0], e[1], e[2], e[3], e[4]) for e in _shapes.values()]
    narrow = min(p[4] for p in poses)

    rows = {}
    for p in placed:
        rows.setdefault(round(p["y"], 1), []).append(p)

    for _round in range(ROW_END_ROUNDS):
        # Only rows with a genuine opening are worth sliding along — this keeps
        # the pass off the hot path of a plate that has nothing to gain.
        live = []
        for y, row in rows.items():
            rh = max(p["h"] for p in row)
            # Measure the opening at the row's WIDE edge, not its narrow one.
            # A row's own chord is taken at whichever edge is further from the
            # centre, because a full-height stone is bound by that; but the seat
            # this pass is looking for is taken by a SHORTER stone, which is
            # bound by the wide edge instead. Testing the narrow edge here made
            # the top row look full — 1.90 mm of opening against a 14.60 mm
            # chord — when the same row has 28.95 mm of chord at its base and a
            # 10 mm stone fits in the corner.
            yin = min(abs(y), abs(y + rh))
            if yin >= R:
                continue
            chord = math.sqrt(R * R - yin * yin)
            lo = min(p["x"] for p in row)
            hi = max(p["x"] + p["w"] for p in row)
            if max(lo + chord, chord - hi) >= narrow:
                # Cap the seat at the space that is REALLY free above this row,
                # not at the row's nominal height. By the time this pass runs the
                # next row is already laid on top, and a row's tallest stone can
                # stand a fraction proud of the line the next row started from.
                # A candidate allowed that full height pokes into the row above —
                # only by hundredths of a millimetre, but enough to catch on it
                # and stop the seat nesting back into its own row, which is what
                # left stones 10 and 15 mm out on a Ø100 plate.
                above = [b for b in rows if b > y + 0.05]
                head = (min(above) - y) if above else (R - y)
                # Carry the ROW's own height alongside the space actually free
                # above it: the first bounds how tall a candidate may be, the
                # second how SHORT it may be and still belong to this row.
                live.append((y, min(rh, head), rh))
        if not live:
            break

        pick = None
        for y, rh, row_h in live:
            # Where the row already reaches, so a seat can be scored on how
            # close to it the stone ends up.
            row_lo = min(p["x"] for p in rows[y])
            row_hi = max(p["x"] + p["w"] for p in rows[y])
            for supply, deg, is_cut, rg, w, h in poses:
                if h > rh + 1e-9:
                    continue
                s = next((c for c in supply if c["stock"] not in used), None)
                if s is None:
                    continue               # every stone of this size is spent
                # ...and not so much SHORTER than its row that the seat stops
                # reading as part of the row. This pass legitimately seats a
                # stone below full row height — that is how it reaches the
                # trapezoid corner — but a stone under
                # ROW_END_MIN_HEIGHT_FRAC of the row leaves a step the floor
                # sees as a gap with a chip in it rather than a filled corner.
                if h < ROW_END_MIN_HEIGHT_FRAC * row_h:
                    continue
                # NEAREST THE ROW, not the first position that happens to fit.
                # The sweep runs from -R upward, so for a seat at the LEFT end of
                # a row the first fit is the one furthest out — the stone ends up
                # against the rim with a channel between it and the row. Sliding
                # it back afterwards is unreliable: it catches on anything it
                # touches on the way. Choosing the innermost fitting position
                # instead puts it where it belongs to begin with.
                near = None
                # A few hundredths of clearance above the baseline. A row's
                # tallest stone can finish a whisker proud of the line the next
                # row starts from — 0.03 mm measured on a Ø100 plate — and that
                # hairline is enough to block a 94 mm2 seat completely, because a
                # candidate laid exactly on the baseline catches on it. Lifting
                # the seat by hundredths of a millimetre clears the protrusion
                # and stays well inside ROW_LEVEL_TOL, so the row still reads
                # level. Without this, two seats worth 219 mm2 were unreachable.
                for lift in _row_end_lifts(rh, h):
                    if h + lift > rh + 1e-9:
                        continue
                    yl = y + lift
                    # Bound the sweep ANALYTICALLY instead of walking the whole
                    # diameter and asking shapely at every step. A stone
                    # occupying the band [yl, yl+h] can only be inside the disc
                    # where |x| clears the narrower of the two edge half-chords,
                    # so everything outside that is a guaranteed reject. The old
                    # loop ran from -R to R at 0.25 mm — 320 positions, each
                    # paying a disc.contains() against a 180-gon and an
                    # intersects() against the union of every placed stone. On a
                    # 20-round pass over 8 rows and ~1200 poses that is where the
                    # time went.
                    _yin = min(abs(yl), abs(yl + h))
                    _hw = math.sqrt(R * R - _yin * _yin) if _yin < R else 0.0
                    if 2.0 * _hw < w:
                        continue                  # band is narrower than the stone
                    x = -_hw
                    x_stop = _hw - w
                    # The seat must sit hard against the row it is filling
                    # beside. A stone that lands with a channel behind it holds
                    # the seat open, leaves a gap nothing can fill, and on the
                    # plate reads as debris against the rim rather than a filled
                    # corner — which is precisely why the previous wide-band
                    # plate was rejected.
                    #
                    # A CUT stone is pinned to the rim and cannot be slid in
                    # afterwards, so it has always been held to this. A whole
                    # stone CAN nest, but only if it started close enough to
                    # reach; letting it take a distant seat produced the scatter.
                    _gap_max = ROW_END_CUT_GAP if is_cut else ROW_END_MAX_GAP
                    while x <= x_stop + 1e-9:
                        # DISTANCE FIRST. It is pure arithmetic, while everything
                        # below builds a polygon and asks GEOS about it. Since
                        # the seat is only acceptable within _gap_max of the row,
                        # and we only want the CLOSEST one, most x positions can
                        # be dismissed before any geometry exists at all — which
                        # is what makes a 20-round pass affordable.
                        if x + w <= row_lo:
                            d = row_lo - (x + w)       # sits left of the row
                        elif x >= row_hi:
                            d = x - row_hi             # sits right of it
                        else:
                            d = 0.0                    # inside the row's span
                        if d > _gap_max or (near is not None and d >= near[0]):
                            x += ROW_END_STEP
                            continue
                        g = affinity.translate(rg, x, yl)
                        if (disc.contains(g) and not g.intersects(taken)
                                and g.distance(taken) >= clear - 1e-9):
                            ok = True
                            if is_cut:
                                # Was an inline copy of the rim + notch tests,
                                # which drifted from _cut_on_rim the moment that
                                # gained the outward-direction rule — and this is
                                # one of the paths that let an inward-facing
                                # cross onto a finished plate. Call the one test.
                                ok = _cut_on_rim(g, R)
                            if ok:
                                if near is None or d < near[0]:
                                    near = (d, g)
                        x += ROW_END_STEP
                    if near is not None and near[0] <= ROW_END_STEP:
                        break      # already tight against the row; no need to lift
                if near is not None:
                    # LEVEL WITH ITS ROW FIRST, size second.
                    #
                    # This pass exists to reach the trapezoid corner beyond a
                    # row's end, and a stone SHORTER than the row is how it gets
                    # there — but only when nothing of the row's own height will
                    # fit. Ranking on area alone ignored that completely, so a
                    # 7.94 mm stone could win a seat at the end of a 9.3 mm row
                    # purely for being wider, leaving a 1.4 mm step along the top
                    # of an otherwise level row. On the plate that reads as a
                    # broken row, which is the one thing the row rules exist to
                    # prevent.
                    #
                    # `level` uses the SAME band the row packer holds a row to
                    # (ROW_LEVEL_TOL), so "matches its row" means here exactly
                    # what it means there. A level stone now beats any shorter
                    # one whatever its area; among equals, the bigger still wins.
                    level = 1 if (row_h - h) <= ROW_LEVEL_TOL else 0
                    key = (level, near[1].area)
                    if pick is None or key > pick[0]:
                        pick = (key, s, deg, near[1], y, is_cut)
        if pick is None:
            break

        _a, s, deg, g, y, _is_cut = pick
        # NEST it, all the way back to the row. The scan above steps in
        # ROW_END_STEP and keeps the position where the stone happened to fit,
        # which leaves it standing off its neighbour — a 1.05 mm slot mid-row on
        # a Ø90 plate, and far worse on a bigger one, where the wedge is deep and
        # the stone can end up marooned out by the rim with a visible channel
        # between it and the row it belongs to.
        #
        # The slide is bounded by the plate, not by a fixed distance: it runs
        # until the stone touches its neighbour or the rim. Capping it at four
        # steps was enough for a Ø90 plate and left stones stranded on a Ø100.
        # OVERLAP, not mere contact. `intersects` is true when two stones merely
        # touch, and this packer lays them edge to edge — so a seat already
        # touching anything, including the row beneath it, failed on the first
        # step and never moved at all. That left stones sitting 10 to 15 mm out
        # from their row on a Ø100 plate while the metric happily reported them
        # as belonging to it. Sliding must stop on real overlap, which has area;
        # contact does not.
        _dir = -1.0 if g.centroid.x > 0 else 1.0
        _d, _limit = 0.0, 2.0 * R
        while _d < _limit:
            _try = affinity.translate(g, _dir * (_d + ROW_END_NEST), 0.0)
            if not disc.contains(_try):
                break
            # A CUT STONE MUST NOT BE NESTED OFF THE RIM. The scan checks the rim
            # rule, and then this slide used to drag the stone inward to meet its
            # row — 18.52 mm on a Ø100 plate, which put a cross 17.30 mm inland
            # with its ground corner against a flat neighbour. The two fixes were
            # working against each other. A cross may close up only as far as the
            # rim rule still holds; the gap that leaves beside it is the curve of
            # the plate, and it belongs there.
            if _is_cut and not _cut_on_rim(_try, R):
                break
            if clear > 0.0:
                if _try.distance(taken) < clear - 1e-9:
                    break          # keep the gap the form asked for
            elif _try.intersection(taken).area > 1e-9:
                break              # edge to edge, but never overlapping
            _d += ROW_END_NEST
        if _d > 0.0:
            g = affinity.translate(g, _dir * _d, 0.0)
        used.add(s["stock"])
        taken = unary_union([taken, g])
        bx0, by0, bx1, by1 = g.bounds
        rp = g.representative_point()
        placed.append({
            "stock": s["stock"], "cts": s.get("cts", 0.0),
            "L": round(bx1 - bx0, 1), "W": round(by1 - by0, 1),
            "H": s.get("H", round((P.T_LO + P.T_HI) / 2.0, 3)),
            "rawL": s.get("L"), "rawW": s.get("W"),
            "x": bx0, "y": by0, "w": bx1 - bx0, "h": by1 - by0, "angle": deg,
            "kind": "real",
            "poly": [(round(px, 3), round(py, 3)) for px, py in g.exterior.coords],
            "area": g.area,
            "irregular": bool(s.get("poly")),
            "lx": rp.x, "ly": rp.y,
        })
        rows.setdefault(round(by0, 1), []).append(placed[-1])

    covered = unary_union([Polygon(p["poly"]).buffer(0) for p in placed]).area
    return placed, 100.0 * covered / (math.pi * R * R)


def render_circle(placed, real, pi, R, fill, path):
    """Render a whole-plate Demo Max-Fill plate: real seeds (viridis) + HYBRID
    fillers — big variable-size filler seeds (blue) and fixed 2×2 dummy seeds (red) —
    across the FULL circle. Filler kind read from p['kind'] ('big' | 'dummy')."""
    PLATE = P.PLATE_D
    big = [p for p in placed if p.get("kind") == "big"]
    small = [p for p in placed if p.get("kind") == "dummy"]
    nr = len(real)
    fig, ax = plt.subplots(figsize=(7.5, 7.5))
    ax.add_patch(Circle((0, 0), PLATE / 2, fc="#e9e9ec", ec="#888", lw=1.5, zorder=0))
    ax.add_patch(Circle((0, 0), R, fc="none", ec="#c0392b", lw=1.3, ls="--", zorder=1))
    cmap = plt.cm.viridis
    # one PatchCollection per group -> 100x fewer draw calls than per-patch add_patch
    if small:
        ax.add_collection(PatchCollection([Rect((p["x"], p["y"]), p["w"], p["h"]) for p in small],
                          facecolor="#f4a3a3", edgecolor="#c0142c", linewidth=0.35, zorder=2))
    if big:
        ax.add_collection(PatchCollection([Rect((p["x"], p["y"]), p["w"], p["h"]) for p in big],
                          facecolor="#bcd6f0", edgecolor="#1f6fb2", linewidth=0.7, zorder=2))
    if real:
        ax.add_collection(PatchCollection([Rect((p["x"], p["y"]), p["w"], p["h"]) for p in real],
                          facecolor=[cmap(i / max(1, nr - 1)) for i in range(nr)],
                          edgecolor="white", linewidth=0.6, zorder=2))
    for p in big:                                      # label EVERY big filler with its EXACT W×H
        fs = min(5.2, max(2.6, min(p["w"], p["h"]) * 0.95))  # scale text to the smaller side
        ax.text(p["x"] + p["w"] / 2, p["y"] + p["h"] / 2, f"{p['w']:.1f}×{p['h']:.1f}",
                ha="center", va="center", fontsize=fs, color="#13507f", zorder=3)
    for i, p in enumerate(real):                       # real-seed labels
        ax.text(p["x"] + p["w"] / 2, p["y"] + p["h"] / 2,
                f"{p['stock']}\n{p['L']:.1f}×{p['W']:.1f}\nH {p['H']:.2f}",
                ha="center", va="center", fontsize=6, color="white", zorder=3)
    lim = PLATE / 2 + 4
    ax.set_xlim(-lim, lim); ax.set_ylim(-lim, lim + 8); ax.set_aspect("equal"); ax.axis("off")
    ds = small[0]["w"] if small else 2.0                  # uniform dummy side (= min filler size)
    ax.legend(handles=[Patch(fc="#2e86c1", label=f"REAL seeds ({nr})"),
                       Patch(fc="#bcd6f0", label=f"big fillers ({len(big)})"),
                       Patch(fc="#f4a3a3", label=f"{ds:g}×{ds:g} dummies ({len(small)})")],
              loc="lower center", bbox_to_anchor=(0.5, -0.03), fontsize=8, frameon=True, ncol=3)
    circle_area = math.pi * R * R
    covered = fill / 100.0 * circle_area
    ax.set_title(f"Mixed + hybrid fill · Plate {pi:02d} · area covered "
                 f"{covered:.0f} mm² of {circle_area:.0f} mm² ({fill:.1f}%)\n"
                 f"Ø{PLATE:g} plate · blue = big fillers (W×H) · red = {ds:g}×{ds:g} mm dummies", fontsize=10)
    fig.savefig(path, dpi=85)              # fixed limits set -> no bbox_inches='tight' rescan
    plt.close(fig)


def render_cross_circle(placed, real, pi, R, fill, path):
    """Render a CROSS-fill plate: real seeds (viridis) + big rectangle fillers (blue)
    + cross/plus seeds (orange, drawn as their two bars) + 2×2 dummies (red)."""
    PLATE = P.PLATE_D
    big = [p for p in placed if p.get("kind") == "big"]
    crosses = [p for p in placed if p.get("kind") == "cross"]
    tris = [p for p in placed if p.get("kind") == "tri"]
    irregs = [p for p in placed if p.get("kind") == "irregular"]
    small = [p for p in placed if p.get("kind") == "dummy"]
    nr = len(real)
    cmap = plt.cm.viridis
    faces = [cmap(i / max(1, nr - 1)) for i in range(nr)]
    DCOL = {"big": "#bcd6f0", "cross": "#f6c177", "tri": "#9ed99b", "irregular": "#c9a0dc", "dummy": "#f4a3a3"}
    DEDGE = {"big": "#1f6fb2", "cross": "#b9770f", "tri": "#2e8b3d", "irregular": "#7d3c98", "dummy": "#c0142c"}

    # Widen the panel for the seed list rather than squeezing the list into a
    # fixed one — see LEGEND_MIN_ROW_IN. The plate panel keeps its 9.6 inches,
    # so the plate itself renders exactly as before at every diameter.
    _lw = _legend_panel_in(len(placed), 9.6 * 0.92)
    fig = plt.figure(figsize=(9.6 + _lw, 9.6))
    gs = fig.add_gridspec(1, 2, width_ratios=[9.6, _lw], wspace=0.02)
    ax = fig.add_subplot(gs[0, 0])
    axl = fig.add_subplot(gs[0, 1]); axl.axis("off")

    ax.add_patch(Circle((0, 0), PLATE / 2, fc="#e9e9ec", ec="#888", lw=1.5, zorder=0))
    ax.add_patch(Circle((0, 0), R, fc="none", ec="#c0392b", lw=1.2, ls="--", zorder=1))
    if small:
        ax.add_collection(PatchCollection([Rect((p["x"], p["y"]), p["w"], p["h"]) for p in small],
                          facecolor="#f4a3a3", edgecolor="#c0142c", linewidth=0.35, zorder=2))
    if big:
        ax.add_collection(PatchCollection([Rect((p["x"], p["y"]), p["w"], p["h"]) for p in big],
                          facecolor="#bcd6f0", edgecolor="#1f6fb2", linewidth=0.7, zorder=2))
    cross_bars = [Rect((bx, by), bw, bh) for p in crosses for (bx, by, bw, bh) in p["bars"]]
    if cross_bars:
        ax.add_collection(PatchCollection(cross_bars, facecolor="#f6c177", edgecolor="#b9770f",
                          linewidth=0.5, zorder=2))
    if tris:                                            # right-triangle seeds (green)
        ax.add_collection(PatchCollection([MplPoly(p["tri"], closed=True) for p in tris],
                          facecolor="#9ed99b", edgecolor="#2e8b3d", linewidth=0.5, zorder=2))
    if irregs:                                          # irregular boundary polygons (purple)
        ax.add_collection(PatchCollection([MplPoly(p["poly"], closed=True) for p in irregs],
                          facecolor="#c9a0dc", edgecolor="#7d3c98", linewidth=0.5, zorder=2))
    if real:                                            # seamless zero-gap real seeds
        ax.add_collection(PatchCollection([Rect((p["x"], p["y"]), p["w"], p["h"]) for p in real],
                          facecolor=faces, edgecolor=faces, linewidth=1.0, zorder=2))

    # NUMBER real seeds 1..N (white); dummy fillers D1..Dk (white too — the "D" distinguishes).
    real_ids = set(id(p) for p in real)
    dummies = [p for p in placed if id(p) not in real_ids]
    nums = [(str(i + 1), p["x"] + p["w"] / 2, p["y"] + p["h"] / 2, p["w"] * p["h"], "white")
            for i, p in enumerate(real)]
    for k, p in enumerate(dummies, 1):
        lx = p.get("lx", p["x"] + p["w"] / 2); ly = p.get("ly", p["y"] + p["h"] / 2)
        nums.append((f"D{k}", lx, ly, p.get("area", p["w"] * p["h"]), "white"))
    _draw_plate_numbers(ax, nums)
    lim = PLATE / 2 + 3
    ax.set_xlim(-lim, lim); ax.set_ylim(-lim, lim); ax.set_aspect("equal"); ax.axis("off")
    circle_area = math.pi * R * R
    covered = fill / 100.0 * circle_area
    ax.set_title(f"Machine-Cut Fill · Plate {pi:02d} · {covered:.0f} of {circle_area:.0f} mm² covered "
                 f"({fill:.1f}%) · Ø{PLATE:g} · {nr} real + {len(dummies)} dummy fillers", fontsize=10.5)

    # ---- seed list on the right: real seeds, then dummy fillers ----
    entries = [(faces[i], "#555", f"{i + 1}.",
                f"{p['stock']}   {p['L']:.1f}×{p['W']:.1f}   H {p['H']:.2f}", "#111")
               for i, p in enumerate(real)]
    for k, p in enumerate(dummies, 1):
        kind = p.get("kind", "dummy")
        entries.append((DCOL.get(kind, "#c9a0dc"), DEDGE.get(kind, "#7d3c98"), f"D{k}",
                        f"filler   {p['w']:.0f}×{p['h']:.0f}   {p.get('area', p['w'] * p['h']):.0f} mm²", "#555"))
    _draw_legend_list(axl, f"Seeds on this plate ({nr})",
                      f"D = dummy filler ({len(dummies)})", entries)

    fig.savefig(path, dpi=125, bbox_inches="tight")
    plt.close(fig)


def render_real_circle(real, pi, R, fill, path):
    """Render the ARRANGE plate: ONLY the real seeds (no dummy fill), each labelled with a
    NUMBER, plus a full seed LIST beside the plate (colour swatch + stock + size + thickness).

    Shows the two settings that shaped the layout — the MARGIN kept clear around
    the plate (shaded between the plate edge and the dashed usable circle) and the
    DISTANCE BETWEEN SEEDS — so a plate can be judged without going back to the
    criteria form. Same treatment as the Max Coverage plate.
    """
    PLATE = P.PLATE_D
    margin = max(0.0, PLATE / 2.0 - R)
    seed_gap = max(0.0, float(getattr(P, "CLEARANCE", 0.0) or 0.0))
    nr = len(real)
    cmap = plt.cm.viridis
    faces = [cmap(i / max(1, nr - 1)) for i in range(nr)]

    # Widen the panel for the seed list rather than squeezing the list into a
    # fixed one — see LEGEND_MIN_ROW_IN. The plate panel keeps its 9.6 inches,
    # so the plate itself renders exactly as before at every diameter.
    _lw = _legend_panel_in(len(placed), 9.6 * 0.92)
    fig = plt.figure(figsize=(9.6 + _lw, 9.6))
    gs = fig.add_gridspec(1, 2, width_ratios=[9.6, _lw], wspace=0.02)
    ax = fig.add_subplot(gs[0, 0])
    axl = fig.add_subplot(gs[0, 1]); axl.axis("off")

    ax.add_patch(Circle((0, 0), PLATE / 2, fc="#e9e9ec", ec="#888", lw=1.5, zorder=0))
    # Shade the MARGIN ring so the clear band reads as a deliberate setting rather
    # than space the packer failed to use.
    if margin > 0.01:
        ax.add_patch(Circle((0, 0), PLATE / 2, fc="#f6dcd7", ec="none", zorder=0.4))
        ax.add_patch(Circle((0, 0), R, fc="#e9e9ec", ec="none", zorder=0.5))
    ax.add_patch(Circle((0, 0), R, fc="none", ec="#c0392b", lw=1.2, ls="--", zorder=1))
    if margin > 0.01:
        d = 0.7071          # 45° diagonal, where seeds rarely reach
        ax.annotate(
            f"margin {margin:g} mm",
            xy=(-(R + margin * 0.5) * d, (R + margin * 0.5) * d),
            xytext=(-(PLATE / 2 + 1) * d - 12, (PLATE / 2 + 1) * d + 6),
            fontsize=9, color="#a03a2c", ha="center", va="bottom", zorder=6,
            arrowprops=dict(arrowstyle="-", color="#a03a2c", lw=0.9,
                            shrinkA=0, shrinkB=1))
    if real:
        # Seamless zero-gap look: each seat's edge painted in its own fill colour.
        ax.add_collection(PatchCollection([Rect((p["x"], p["y"]), p["w"], p["h"]) for p in real],
                          facecolor=faces, edgecolor=faces, linewidth=1.0, zorder=2))
    _cw = [display_turn(p) for p in real]
    _draw_plate_numbers(ax, [(str(i + 1), p["x"] + p["w"] / 2, p["y"] + p["h"] / 2,
                              p["w"] * p["h"], "white") for i, p in enumerate(real)],
                        angles=_cw)
    lim = PLATE / 2 + 3
    ax.set_xlim(-lim, lim); ax.set_ylim(-lim, lim); ax.set_aspect("equal"); ax.axis("off")
    circle_area = math.pi * R * R
    covered = fill / 100.0 * circle_area
    ax.set_title(
        f"Arrange · Plate {pi:02d} · {nr} seeds · {covered:.0f} of {circle_area:.0f} mm² "
        f"covered ({fill:.1f}%)\n"
        f"plate Ø{PLATE:g} · margin {margin:g} mm → usable Ø{2 * R:g} · "
        f"distance between seeds {seed_gap:g} mm\n"
        f"{band_caption()}",
        fontsize=10.5)

    entries = [(faces[i], "#555", f"{i + 1}.",
                f"{p['stock']}   {p['L']:.1f}×{p['W']:.1f}   H {p['H']:.2f}   ↻{_cw[i]}°", "#111")
               for i, p in enumerate(real)]
    _draw_legend_list(
        axl, f"Seeds on this plate ({nr})",
        ("real seeds only · "
         + (f"{seed_gap:g} mm between seeds" if seed_gap > 0 else "seeds touching")
         + "\n↻ = turn CLOCKWISE from the seed as measured"),
        entries)

    fig.savefig(path, dpi=125, bbox_inches="tight")
    plt.close(fig)


def real_only_job(args):
    """Picklable per-plate worker: render ONLY the real seeds (NO dummy/filler fill). Same
    signature/return as `plate_job` so it is a drop-in for the Arrange view.
    `args` = (real, pi, plate_d, R, min_size, path). Returns (pi, placed, fill, span, gap)."""
    real, pi, plate_d, R, min_size, path = args
    P.PLATE_D = plate_d
    placed = list(real)                                    # real seeds only — no fillers appended
    fill = 100 * sum(p["w"] * p["h"] for p in placed) / (math.pi * R * R)
    span = P.span_mm(placed); gap = P.inner_gap(placed)
    render_real_circle(real, pi, R, fill, path)
    return (pi, placed, round(fill, 4), round(span, 4), round(gap, 4))


def main():
    ap = argparse.ArgumentParser(description="Demo max-fill with synthetic fillers (saved separately).")
    ap.add_argument("file", nargs="?", default="BLOCK.xlsx")
    ap.add_argument("--t-lo", dest="t_lo", type=float, default=0.67)
    ap.add_argument("--t-hi", dest="t_hi", type=float, default=0.73)
    ap.add_argument("--plate", type=float, default=90.0)
    ap.add_argument("--margin", type=float, default=3.0)
    ap.add_argument("--shape", default="all", choices=["all", "square", "rectangle"])
    ap.add_argument("--min-seed", dest="min_seed", type=float, default=2.0)
    a = ap.parse_args()

    P.T_LO, P.T_HI = a.t_lo, a.t_hi
    P.PLATE_D, P.USABLE_D = a.plate, a.plate - a.margin
    P.R = P.USABLE_D / 2.0
    P.S2 = (P.USABLE_D / math.sqrt(2)) / 2.0
    P.INSCRIBED = P.USABLE_D / math.sqrt(2)
    P.MINSEED = a.min_seed
    R = P.R
    circle_area = math.pi * R * R
    allowed = {"square", "rectangle"} if a.shape == "all" else {a.shape}
    blocks = P.load_blocks(a.file, shapes=allowed, square_tol=0.05)
    print(f"real eligible seeds: {len(blocks)} (thickness {a.t_lo}-{a.t_hi} mm, shape {a.shape})")

    queue = P._mixed_landscape(blocks)                  # MIXED real packing (square+rect, whole plate)
    img_dir = "images_demofill"; os.makedirs(img_dir, exist_ok=True)
    for f in glob.glob(f"{img_dir}/*.png"):
        os.remove(f)
    wb = Workbook(); sm = wb.active; sm.title = "Summary"
    sm.append(["PLATE", "REAL_SEEDS", "BIG_FILLERS", "DUMMY_2x2", "PLATE_FILL_%"])
    pi = total_real = 0; fills = []
    while queue:
        real_placed = P._mixed_one_plate(queue)         # real seeds, mixed rows, whole circle
        if not real_placed:
            break
        pi += 1
        big, small = hybrid_fill(real_placed, R, min_size=a.min_seed, step=0.5)
        placed = list(real_placed)
        for k, (x, y, w, h) in enumerate(big, 1):
            placed.append({"stock": f"FILL-{pi:02d}-F{k}", "cts": 0.0, "L": w, "W": h, "H": 0.0,
                           "x": x, "y": y, "w": w, "h": h, "angle": 0, "filler": True, "kind": "big"})
        for k, (x, y) in enumerate(small, 1):
            placed.append({"stock": f"FILL-{pi:02d}-D{k}", "cts": 0.0, "L": a.min_seed, "W": a.min_seed,
                           "H": 0.0, "x": x, "y": y, "w": a.min_seed, "h": a.min_seed, "angle": 0,
                           "filler": True, "kind": "dummy"})
        real = [p for p in placed if not p.get("filler")]
        fill = 100 * sum(p["w"] * p["h"] for p in placed) / circle_area
        fills.append(fill); total_real += len(real)
        print(f"  plate {pi:02d}: {len(real)} real + {len(big)} big + {len(small)} dummy(2×2) "
              f"-> {fill:.2f}% plate-fill")
        render_circle(placed, real, pi, R, fill, f"{img_dir}/plate_{pi:02d}.png")
        ws = wb.create_sheet(f"Plate_{pi:02d}")
        ws.append(["TYPE", "STOCK_NO", "W_mm", "H_mm", "SHAPE", "CENTER_X", "CENTER_Y", "ANGLE"])
        for p in placed:
            shp = "SQUARE" if abs(p["w"] - p["h"]) <= 0.05 * max(p["w"], p["h"]) else "RECT"
            typ = "REAL" if not p.get("filler") else ("BIG_FILLER" if p.get("kind") == "big" else "DUMMY_2x2")
            ws.append([typ, p["stock"], round(p["w"], 2), round(p["h"], 2), shp,
                       round(p["x"] + p["w"] / 2, 2), round(p["y"] + p["h"] / 2, 2), p["angle"]])
        sm.append([pi, len(real), len(big), len(small), round(fill, 2)])
    if pi:
        sm.append([]); sm.append(["AVG", "", "", "", round(sum(fills) / pi, 2)])
    for c in range(1, 6):
        sm.cell(1, c).fill = PatternFill("solid", fgColor="1F4E78")
        sm.cell(1, c).font = Font(bold=True, color="FFFFFF")
    P.safe_save(wb, "arrangement_demofill.xlsx")
    print(f"-> {pi} plates, avg {sum(fills) / max(1, pi):.2f}% plate-fill  "
          f"(images_demofill\\, arrangement_demofill.xlsx)")


if __name__ == "__main__":
    main()