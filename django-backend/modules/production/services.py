# ===================== APPLICATION LAYER: services =====================
# Seed import: read the uploaded Excel datasheet, auto-create any new batches in
# TRN_Batch, skip duplicate stock numbers, and insert the new rows into
# TRN_SeedData. Mirrors the original engine's reader — columns are POSITIONAL
# (row 1 = header): BatchNo, StockNo, Pcs, Cts, Length, Width, Height.
#
# Nothing is filtered by size/shape here (unlike the packer's gate) — every data
# row in the sheet that isn't a duplicate is inserted, so the table mirrors the
# datasheet.

import json
import uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Count, Q
from openpyxl import load_workbook

from modules.core.exceptions import ConflictError

from .models import Batch, DomainError, SeedData
from .repository import BatchRepository, SeedRepository
from .shapes import (
    CUT_CORNER_DEFAULT, CUT_EPS, parse_cut_corner, sides_to_corners,
    validate_corners,
)


# Max ids per IN clause. SQL Server caps a statement at 2100 parameters, and
# mssql-django's overflow path spills to a temp table that TRUNCATES 36-char
# UUIDs to char(32) — so going over does not raise, it silently returns the
# wrong rows. 900 leaves room for the query's other parameters.
IN_CHUNK = 900


def chunked_in(ids, fetch):
    """Run `fetch(chunk)` over `ids` in safe-sized batches and chain the results.

    Every `__in` clause keyed on arrangement or seed ids has to go through here.
    The history screen passes EVERY active arrangement's id in one clause, which
    is fine at today's 296 runs and breaks silently somewhere past 2100 — the
    kind of fault that appears years later as "the history screen started
    showing the wrong plate counts" with nothing in the logs.
    """
    ids = list(ids)
    out = []
    for i in range(0, len(ids), IN_CHUNK):
        out.extend(fetch(ids[i:i + IN_CHUNK]))
    return out


def _uuid_or_error(value, field="arrangeId"):
    """Coerce an id to UUID, or raise DomainError.

    The arrange/plate tables key on `uniqueidentifier`, so handing the ORM a
    non-UUID string makes Django's field raise ValidationError — which no view
    catches, so the caller got a bare HTTP 500 with a stack trace. A DomainError
    renders as a clean {"detail": ...} through the existing handler instead.
    (The ORM parameterises either way; this is robustness, not injection.)
    """
    try:
        return uuid.UUID(str(value).strip())
    except (AttributeError, TypeError, ValueError):
        raise DomainError(f"{field} is not a valid id.")


class BatchService:
    @staticmethod
    def list_with_counts():
        """Batches + seed counts, for the Batch Selection screen."""
        return BatchRepository.list_with_counts()


class ArrangementService:
    """Read-only history of every arrangement run (TRN_SeedArrange header + its
    TRN_SeedPlate rows), for the Arrangement History screen."""

    @staticmethod
    def _num(v):
        return float(v) if v is not None else None

    @staticmethod
    def _method(has_arrange_img, has_machine_img, has_enhanced_img, has_excel):
        """The processing option isn't stored in its own column, but each option writes a
        distinct set of artifacts, so it is recoverable. A run that produced more than one
        method's image is a Compare. Runs saved before EnhancedImagePath existed have no
        Max Coverage image, so they fall back to the Excel-only signal."""
        images = sum((bool(has_arrange_img), bool(has_machine_img), bool(has_enhanced_img)))
        if images > 1:
            return "Compare"
        if has_excel and images == 1 and not has_enhanced_img:
            return "Compare"          # Excel + a single non-enhanced image = Compare
        if has_enhanced_img:
            return "Max Coverage"
        if has_excel and images == 0:
            return "Max Coverage"     # legacy Max Coverage: Excel only, image never stored
        if has_machine_img:
            return "Machine-Cut Fill"
        if has_arrange_img:
            return "Arrange"
        return "—"

    @classmethod
    def _methods_by_run(cls, arrange_ids):
        """arrange_id → method label, from each run's stored plate artifacts."""
        from .models import SeedArrangePlate

        flags = {}
        rows = chunked_in(arrange_ids, lambda c: SeedArrangePlate.objects.filter(
            arrange_id__in=c).values_list(
            "arrange_id", "arrange_image_path", "machine_cut_image_path",
            "enhanced_image_path", "excel_path",
        ))
        for aid, a_img, m_img, e_img, xls in rows:
            f = flags.setdefault(aid, [False, False, False, False])
            f[0] |= bool(a_img)
            f[1] |= bool(m_img)
            f[2] |= bool(e_img)
            f[3] |= bool(xls)
        return {aid: cls._method(*f) for aid, f in flags.items()}

    @classmethod
    def list(cls):
        """Every arrange run, newest first, with its plate count and method.

        TRN_SeedArrange.EntryDate is a DATE (no time), so it cannot order runs made on the
        same day — sorting on it alone falls back to comparing random UUIDs and buries the
        newest run. TRN_SeedPlate.EntryDate IS a datetime, so we sort on each run's latest
        plate timestamp and only fall back to the header's date when a run has no plates.
        """
        from datetime import datetime, time, timezone

        from django.db.models import Count, Max

        from .models import SeedArrange, SeedArrangePlate

        rows = list(SeedArrange.objects.filter(is_active=True))
        ids = [r.arrange_id for r in rows]
        agg = {
            a["arrange_id"]: a
            for a in chunked_in(ids, lambda c: SeedArrangePlate.objects.filter(
                arrange_id__in=c).values("arrange_id")
                .annotate(n=Count("arrange_plate_id"), t=Max("entry_date")))
        }
        methods = cls._methods_by_run(ids)

        def run_at(r):
            t = (agg.get(r.arrange_id) or {}).get("t")
            if t:
                return t
            if r.entry_date:  # date-only header → midnight UTC, so it sorts below real runs
                return datetime.combine(r.entry_date, time.min, tzinfo=timezone.utc)
            return datetime.min.replace(tzinfo=timezone.utc)

        rows.sort(key=run_at, reverse=True)
        # How much inventory each run is holding (TRN_SeedData.Used_ID). Shown in
        # the history so a run's seeds can be released from a screen that always
        # works — Finalization needs a live in-memory job, which is gone after a
        # backend restart, and the release action went with it.
        from .models import SeedData
        held = {
            r["used_id"]: r["n"]
            for r in chunked_in(ids, lambda c: SeedData.objects.filter(used_id__in=c)
                                .values("used_id").annotate(n=Count("seed_id")))
        }
        return [
            {
                "arrangeId": str(r.arrange_id),
                "method": methods.get(r.arrange_id, "—"),
                "mode": r.mode,
                "shape": r.shape,
                "plateCount": (agg.get(r.arrange_id) or {}).get("n", r.plate_no or 0),
                "average": cls._num(r.average),
                "plateDiameter": cls._num(r.plate_diameter),
                "thicknessMin": cls._num(r.thickness_min),
                "thicknessMax": cls._num(r.thickness_max),
                # Seed-width band the run used. null at either end = unbounded
                # there, which is also every run made before the field existed.
                "widthMin": cls._num(r.width_min),
                "widthMax": cls._num(r.width_max),
                # Stored as a comma-joined string; hand the UI a real list.
                "batches": [b for b in (r.batches or "").split(",") if b.strip()],
                "entryDate": r.entry_date.isoformat() if r.entry_date else None,
                # Real timestamp (from the plate rows) — what the list is actually sorted by.
                "runAt": run_at(r).isoformat(),
                "isFinalized": bool(r.is_finalized),
                "seedsHeld": held.get(r.arrange_id, 0),
            }
            for r in rows
        ]

    # detail.Method → the label used on the plate images, so the UI can pair list to image.
    _METHOD_KEY_LABEL = {"arrange": "Arrange", "machinefill": "Machine-Cut", "enhanced": "Max Coverage"}

    # Max ids per IN clause — well under SQL Server's 2100-parameter statement limit.
    _IN_CHUNK = 900

    @classmethod
    def _fetch_seeds(cls, model, ids):
        """seed_id → (id, stock, L, W, H, cts) for `ids`, fetched in safe-sized chunks."""
        ids = list(ids)
        out = {}
        for i in range(0, len(ids), cls._IN_CHUNK):
            out.update({
                r[0]: r
                for r in model.objects.filter(seed_id__in=ids[i:i + cls._IN_CHUNK])
                .values_list("seed_id", "stock_no", "length", "width", "height", "cts")
            })
        return out

    @classmethod
    def _seeds_by_plate(cls, arrange_id, fallback_method):
        """(plate_no, method_label) → the seeds that method placed on that plate, from
        TRN_SeedArrangeDetails joined to the seed tables. `Plate_ID` on a detail row holds
        the PLATE NUMBER (not an MST_SeedPlate id). Rows written before per-method saving
        have Method NULL and are attributed to `fallback_method`."""
        from .models import DummySeedData, SeedArrangeDetail, SeedData

        rows = list(
            SeedArrangeDetail.objects.filter(arrange_id=arrange_id)
            .values_list("plate_id", "seed_type", "seed_id", "method", "cut_area_mm2", "cut_pct")
        )
        if not rows:
            return {}
        # A Compare run stores a row per seed PER METHOD, so the same Seed_ID repeats.
        # De-duplicate, and chunk the IN clause: SQL Server caps a statement at 2100
        # parameters, and the driver's overflow path spills to a temp table that truncates
        # 36-char UUIDs to char(32).
        real_ids = {r[2] for r in rows if not r[1] and r[2]}
        dummy_ids = {r[2] for r in rows if r[1] and r[2]}
        real = cls._fetch_seeds(SeedData, real_ids)
        dummy = cls._fetch_seeds(DummySeedData, dummy_ids)

        out = {}
        for plate_no, is_dummy, sid, method, cut_area, cut_pct in rows:
            src = (dummy if is_dummy else real).get(sid)
            if plate_no is None or src is None:
                continue
            label = cls._METHOD_KEY_LABEL.get(method) if method else fallback_method
            _, stock, length, width, height, cts = src
            out.setdefault((plate_no, label or "—"), []).append({
                "stock": stock or "—",
                "length": cls._num(length),
                "width": cls._num(width),
                "height": cls._num(height),
                "cts": cls._num(cts),
                # 0.00 is stored for whole (untrimmed) seeds — surface it as "no trim".
                "cutArea": (cls._num(cut_area) or None) if cut_area else None,
                "cutPct": (cls._num(cut_pct) or None) if cut_pct else None,
                "real": not is_dummy,
            })
        for seeds in out.values():
            seeds.sort(key=lambda s: (not s["real"], s["stock"]))
        return out

    @classmethod
    def detail(cls, arrange_id):
        """One run's header plus a row per plate (fill %s, counts, image/Excel paths, seeds)."""
        from .models import SeedArrange, SeedArrangeDetail, SeedArrangePlate

        arrange_id = _uuid_or_error(arrange_id)
        header = SeedArrange.objects.filter(arrange_id=arrange_id).first()
        if header is None:
            raise DomainError("Arrangement not found.")
        plates = list(SeedArrangePlate.objects.filter(arrange_id=arrange_id).order_by("plate_no"))
        method = cls._methods_by_run([header.arrange_id]).get(header.arrange_id, "—")
        # Legacy rows (Method NULL) came from the single method that fed per_plate_real:
        # Arrange won over the others, so a legacy Compare's list is its Arrange list.
        fallback = {"Compare": "Arrange", "Machine-Cut Fill": "Machine-Cut"}.get(method, method)
        seeds_by_plate = cls._seeds_by_plate(arrange_id, fallback)
        # Count the PRIMARY method's real seeds only — a Compare run now stores a row per
        # seed per method, so a raw count would multiply by the number of methods.
        seed_count = sum(
            sum(1 for s in seeds if s["real"])
            for (_, label), seeds in seeds_by_plate.items()
            if label == fallback
        ) or SeedArrangeDetail.objects.filter(arrange_id=arrange_id, seed_type=False).count()
        return {
            "arrangeId": str(header.arrange_id),
            "method": method,
            "mode": header.mode,
            "shape": header.shape,
            "average": cls._num(header.average),
            "plateDiameter": cls._num(header.plate_diameter),
            "thicknessMin": cls._num(header.thickness_min),
            "thicknessMax": cls._num(header.thickness_max),
            "widthMin": cls._num(header.width_min),
            "widthMax": cls._num(header.width_max),
            "batches": [b for b in (header.batches or "").split(",") if b.strip()],
            "entryDate": header.entry_date.isoformat() if header.entry_date else None,
            "isFinalized": bool(header.is_finalized),
            "seedCount": seed_count,
            "plates": [
                {
                    "plateNo": p.plate_no,
                    "plateName": p.plate_name,
                    "arrangeFillPct": cls._num(p.arrange_fill_pct),
                    "machineFillPct": cls._num(p.machine_fill_pct),
                    "enhancedFillPct": cls._num(p.enhanced_fill_pct),
                    "finalizedFillPct": cls._num(p.finalized_fill_pct),
                    "realSeedCount": p.real_seed_count,
                    "dummyCount": p.dummy_count,
                    "arrangeImageUrl": p.arrange_image_path,
                    "machineImageUrl": p.machine_cut_image_path,
                    "enhancedImageUrl": p.enhanced_image_path,
                    "finalizedImageUrl": p.finalized_image_path,
                    "excelUrl": p.excel_path,
                    # One seed list per method, keyed by the same label as the plate image.
                    "seedsByMethod": {
                        label: seeds
                        for (plate_no, label), seeds in seeds_by_plate.items()
                        if plate_no == p.plate_no
                    },
                }
                for p in plates
            ],
        }


class InventoryService:
    """Consuming seeds when an arrangement is finalized, and giving them back.

    A finalized arrangement's stones are physically committed to plates, so they
    must not be offered to the next run — that is what TRN_SeedData.ISUsed and
    Used_ID are for. Both columns have existed since the schema was written and
    were never populated: an earlier attempt wrote ISUsed at GENERATION time,
    which consumed inventory merely for previewing a layout, and was rolled back
    (see the note in engine_runner.write_arrangement). Tying it to the user's
    explicit finalize instead is what makes it safe.

    Used_ID records WHICH arrangement took each seed. That is what lets
    unfinalize give back exactly the seeds this run consumed and no others —
    without it, releasing one arrangement could free stones another one is
    still using.
    """

    # Which layout's seed list a plate consumes. A Compare run records BOTH an
    # `arrange` and an `enhanced` placement for the same stones, and the two put
    # 36-67% of seeds on DIFFERENT plate numbers (measured across the live
    # history) — so "the seeds on plate 2" only has an answer once a layout is
    # named. Max Coverage is the layout that gets physically built, so it wins.
    BUILT_METHOD = "enhanced"

    @staticmethod
    def seed_ids_for(arrange_id, plate_no=None):
        """Real seeds this arrangement placed — all of them, or one plate's.

        Dummy/filler seeds carry SeedType=True and are excluded: they are not
        inventory. Rows are restricted to the built layout (see BUILT_METHOD),
        falling back to whatever methods exist for runs that never produced a
        Max Coverage layout — otherwise an Arrange-only run would consume
        nothing at all.
        """
        from .models import SeedArrangeDetail

        qs = (SeedArrangeDetail.objects
              .filter(arrange_id=arrange_id, seed_type=False)
              .exclude(seed_id__isnull=True))
        if plate_no is not None:
            qs = qs.filter(plate_id=plate_no)
        built = qs.filter(method=InventoryService.BUILT_METHOD)
        return set((built if built.exists() else qs).values_list("seed_id", flat=True))

    @staticmethod
    @transaction.atomic
    def consume_plate(arrange_id, plate_no, user_id=None):
        """Take one plate's seeds out of the pool — called when the plate is
        assigned a name, which is what "this plate is finalized" means here.

        Only this plate's stones are touched, so the rest of the arrangement
        stays available and can still be re-generated against what is left.
        """
        from .models import SeedData

        seed_ids = InventoryService.seed_ids_for(arrange_id, plate_no)
        if not seed_ids:
            return {"placed": 0, "consumed": 0}
        # A stone already committed to a DIFFERENT arrangement must not be
        # silently re-taken — that would be two physical plates claiming one
        # seed. Happens when this run was generated before that one was
        # assigned, so its seed list is stale.
        clash = list(
            SeedData.objects
            .filter(seed_id__in=seed_ids, is_used=True)
            .exclude(used_id=arrange_id)
            .values_list("stock_no", flat=True)[:10]
        )
        if clash:
            raise ConflictError(
                "These seeds are already used by another finalized plate: "
                + ", ".join(str(s) for s in clash)
                + ". Re-generate this arrangement so it excludes them."
            )
        now = datetime.now(timezone.utc)
        consumed = SeedData.objects.filter(seed_id__in=seed_ids).update(
            is_used=True, used_id=arrange_id, update_date=now, update_by=user_id,
        )
        # `placed` counts what the plate recorded; `consumed` counts the rows
        # actually flipped. They differ when a re-import replaced the seed rows,
        # leaving old details pointing at Seed_IDs no longer in inventory.
        return {"placed": len(seed_ids), "consumed": consumed}

    @staticmethod
    @transaction.atomic
    def return_plate(arrange_id, plate_no, user_id=None):
        """Give one plate's seeds back — called when its name is released.

        Matched on Used_ID as well as the seed list, so a stone another
        arrangement is holding is never freed by releasing this plate.
        """
        from .models import SeedData

        seed_ids = InventoryService.seed_ids_for(arrange_id, plate_no)
        if not seed_ids:
            return {"returned": 0}
        now = datetime.now(timezone.utc)
        returned = SeedData.objects.filter(
            seed_id__in=seed_ids, used_id=arrange_id,
        ).update(is_used=False, used_id=None, update_date=now, update_by=user_id)
        return {"returned": returned}

    @staticmethod
    @transaction.atomic
    def unfinalize(arrange_id, user_id=None):
        """Return EVERY seed this arrangement is holding, whichever plate took it.

        A recovery action, not part of the normal flow — releasing a plate is
        the per-plate route. Kept because it is the only way back from a bulk
        consume, including one done by an earlier version of this feature.

        Filters on Used_ID, never on the detail list, so a seed another run took
        stays taken even if both runs happened to place it.
        """
        from .models import SeedArrange, SeedData

        arrange_id = _uuid_or_error(arrange_id)
        header = SeedArrange.objects.filter(arrange_id=arrange_id).first()
        if header is None:
            raise DomainError("Arrangement not found.")
        now = datetime.now(timezone.utc)
        returned = SeedData.objects.filter(used_id=arrange_id).update(
            is_used=False, used_id=None, update_date=now, update_by=user_id,
        )
        header.finalized_by_user = False
        header.finalized_by = None
        header.finalized_date = None
        header.save(update_fields=["finalized_by_user", "finalized_by", "finalized_date"])
        return {"finalized": False, "returned": returned}

    @staticmethod
    def sync_header(arrange_id, user_id=None):
        """TRN_SeedArrange.FinalizedByUser = "every plate of this run is named".

        IsFinalized is set to 1 automatically on every generated run, so it
        cannot carry this meaning; FinalizedByUser is the column that can.
        """
        from .models import SeedArrange, SeedArrangePlate

        header = SeedArrange.objects.filter(arrange_id=arrange_id).first()
        if header is None:
            return False
        rows = list(SeedArrangePlate.objects.filter(arrange_id=arrange_id)
                    .values_list("plate_name", flat=True))
        done = bool(rows) and all((n or "").strip() for n in rows)
        if bool(header.finalized_by_user) == done:
            return done
        now = datetime.now(timezone.utc)
        header.finalized_by_user = done
        header.finalized_by = user_id if done else None
        header.finalized_date = now if done else None
        header.save(update_fields=["finalized_by_user", "finalized_by", "finalized_date"])
        return done

    @staticmethod
    def status(arrange_id):
        """What the Finalization screen shows: which plates are committed, and
        how much inventory is left for the next run."""
        from .models import SeedArrange, SeedArrangePlate, SeedData

        arrange_id = _uuid_or_error(arrange_id)
        header = SeedArrange.objects.filter(arrange_id=arrange_id).first()
        if header is None:
            raise DomainError("Arrangement not found.")
        named = {
            pn: (nm or "").strip()
            for pn, nm in SeedArrangePlate.objects.filter(arrange_id=arrange_id)
            .values_list("plate_no", "plate_name")
        }
        # Seeds this run placed that some OTHER run has since committed. Once a
        # plate has any, it can no longer be built as drawn — assigning it would
        # be refused, so the screen must say so BEFORE the user tries rather
        # than failing at the click. Happens when a second arrangement is
        # generated from the leftovers and assigned first.
        all_ids = InventoryService.seed_ids_for(arrange_id)
        # Chunked: a wide-band run places 60+ seeds per plate, so a multi-plate
        # arrangement's seed list runs to several hundred ids and a big one would
        # reach the 2100-parameter cap here first.
        taken_elsewhere = set(chunked_in(all_ids, lambda c: SeedData.objects
                                         .filter(seed_id__in=c, is_used=True)
                                         .exclude(used_id=arrange_id)
                                         .values_list("seed_id", flat=True))) if all_ids else set()
        plates = []
        for pn in sorted(p for p in named if p is not None):
            ids = InventoryService.seed_ids_for(arrange_id, pn)
            lost = len(ids & taken_elsewhere)
            plates.append({
                "plateNo": pn,
                "plateName": named.get(pn) or None,
                "seeds": len(ids),
                "consumed": bool(named.get(pn)),
                # Non-zero = this plate is stale; re-generate before assigning.
                "takenElsewhere": lost,
                "canAssign": lost == 0,
            })
        return {
            "arrangeId": str(arrange_id),
            "isFinalized": bool(header.finalized_by_user),
            "plates": plates,
            "seedsInArrangement": len(InventoryService.seed_ids_for(arrange_id)),
            "seedsConsumedByThisRun": SeedData.objects.filter(used_id=arrange_id).count(),
            "seedsAvailable": SeedData.objects.exclude(is_used=True).count(),
            "seedsUsedTotal": SeedData.objects.filter(is_used=True).count(),
        }


class PlateService:
    """Per-plate name assignment against the MST_SeedPlate master (finalize step).
    Free-text names are supported — a master row is created on demand."""

    @staticmethod
    @transaction.atomic
    def assign(arrange_id, plate_no, plate_name, user_id=None):
        from .models import SeedArrangePlate, SeedPlate

        name = (plate_name or "").strip()
        if not name or arrange_id is None or plate_no is None:
            raise DomainError("arrangeId, plateNo and plateName are required.")
        arrange_id = _uuid_or_error(arrange_id)
        row = SeedArrangePlate.objects.filter(arrange_id=arrange_id, plate_no=plate_no).first()
        if row is None:
            raise DomainError("Plate not found.")
        now = datetime.now(timezone.utc)
        old = (row.plate_name or "").strip()
        if old and old != name:  # release the plate's previous name back to the pool
            SeedPlate.objects.filter(plate_name=old).update(is_used=False, is_released=True, update_date=now)
        master, _ = SeedPlate.objects.get_or_create(plate_name=name[:50], defaults={"is_active": True})
        master.is_used = True
        master.is_released = False
        master.update_date = now
        master.save(update_fields=["is_used", "is_released", "update_date"])
        row.plate_name = master.plate_name
        row.plate_id = master.plate_id
        row.update_date = now
        row.save(update_fields=["plate_name", "plate_id", "update_date"])
        # Naming a plate IS finalizing it: its stones are committed, so they
        # leave the pool now — and only this plate's, so the rest of the run
        # stays available for a re-generate. Raises ConflictError if any of them
        # already belongs to another plate, which aborts the whole assign.
        taken = InventoryService.consume_plate(arrange_id, plate_no, user_id)
        InventoryService.sync_header(arrange_id, user_id)
        return {"assigned": True, "plateName": master.plate_name,
                "seedsConsumed": taken["consumed"]}

    @staticmethod
    @transaction.atomic
    def release(arrange_id, plate_no, user_id=None):
        from .models import SeedArrangePlate, SeedPlate

        arrange_id = _uuid_or_error(arrange_id)
        row = SeedArrangePlate.objects.filter(arrange_id=arrange_id, plate_no=plate_no).first()
        if not row or not row.plate_name:
            return {"released": False, "plateName": None, "seedsReturned": 0}
        name = row.plate_name
        now = datetime.now(timezone.utc)
        SeedPlate.objects.filter(plate_name=name).update(is_used=False, is_released=True, update_date=now)
        row.plate_name = None
        row.plate_id = None
        row.update_date = now
        row.save(update_fields=["plate_name", "plate_id", "update_date"])
        # The plate is no longer committed, so its stones go back in the pool —
        # only the ones THIS arrangement took (matched on Used_ID).
        back = InventoryService.return_plate(arrange_id, plate_no, user_id)
        InventoryService.sync_header(arrange_id, user_id)
        return {"released": True, "plateName": name, "seedsReturned": back["returned"]}

    @staticmethod
    @transaction.atomic
    def release_plate(plate_id):
        """Free a plate from the PLATE MASTER side, given only its Plate_ID.

        `release` above is keyed on (arrange_id, plate_no) because Finalization
        works one arrangement at a time. Plate Master has no arrangement in hand
        — it lists the master rows — so it could not call it, and ISUsed is
        read-only on the serializer. A plate left reading "in use" was therefore
        unreachable: the only way to free it was to walk back into the exact
        arrangement that took it.

        This clears BOTH sides, which `release` also does but keyed the other
        way round: the master row goes back to the pool, and any arrangement
        still naming this plate forgets it. Clearing the arrangement matters —
        leaving it pointing at a freed plate is how two arrangements end up
        claiming the same physical plate.
        """
        from .models import SeedArrangePlate, SeedPlate

        try:
            pid = int(plate_id)
        except (TypeError, ValueError):
            raise DomainError("A valid plateId is required.")
        master = SeedPlate.objects.filter(plate_id=pid).first()
        if master is None:
            raise DomainError("Plate not found.")
        now = datetime.now(timezone.utc)
        # Match on the id AND on the name: rows written before Plate_ID was
        # stored carry only the name, and they must be cleared too or the plate
        # reads free here while an old arrangement still lists it.
        used_by = SeedArrangePlate.objects.filter(
            Q(plate_id=pid) | Q(plate_name=master.plate_name)
        ) if master.plate_name else SeedArrangePlate.objects.filter(plate_id=pid)
        cleared = used_by.update(plate_name=None, plate_id=None, update_date=now)
        master.is_used = False
        master.is_released = True
        master.update_date = now
        master.save(update_fields=["is_used", "is_released", "update_date"])
        return {"released": True, "plateName": master.plate_name, "clearedFrom": cleared}

    @staticmethod
    @transaction.atomic
    def set_active(plate_id, active, user_id=None):
        """Take a plate out of circulation, or put it back. Touches ISActive ONLY.

        `is_active` is not a display flag: the plate names Finalization offers are
        `SeedPlate.objects.filter(is_active=True)`, so clearing it is what removes
        a plate from the dropdown. That makes it the SOFT DELETE this app wants —
        the row stays, the arrangements that named it keep their history, and the
        same call puts it back.

        WHY THIS IS A POST AND NOT A DELETE OR PUT. Live runs under IIS, whose
        WebDAV module answers DELETE and PUT with 405 before the request reaches
        Python — the same reason Plate Master's edit and delete and Finalization's
        "Return all" do not work there while every POST does. A hard delete would
        also be the wrong shape regardless: MST_SeedPlate carries no foreign keys
        (see sql/create_arrange_tables.sql), so removing a row would silently
        orphan every arrangement still pointing at it instead of being refused.

        Deliberately NOT touched: is_used, is_released, the arrangement link, and
        any seed. Releasing inventory is InventoryService.release's job and stays
        exactly as it was — a plate can be deactivated while free, and taking it
        out of the dropdown must never move stock on its own.
        """
        from .models import SeedPlate

        try:
            pid = int(plate_id)
        except (TypeError, ValueError):
            raise DomainError("A valid plateId is required.")
        master = SeedPlate.objects.filter(plate_id=pid).first()
        if master is None:
            raise DomainError("Plate not found.")
        # A plate a run is still holding must not vanish from the dropdown while
        # its stones are committed — release it first, which is the action that
        # hands the inventory back.
        if not active and master.is_used:
            raise DomainError(
                "This plate is still assigned to an arrangement. Release it "
                "first, then deactivate it.")
        master.is_active = bool(active)
        master.update_date = datetime.now(timezone.utc)
        master.update_by = user_id
        master.save(update_fields=["is_active", "update_date", "update_by"])
        return {"plateId": pid, "plateName": master.plate_name,
                "active": bool(master.is_active)}

    @staticmethod
    def finalized_list():
        """Every plate that carries an assigned name — i.e. every finalized plate.

        Reads TRN_SeedPlate, not the job cache, so it survives a backend restart
        and does not need the run that produced it to still be open. Newest
        first. Seed counts come from the built (Max Coverage) layout, matching
        what was actually consumed.
        """
        from .models import SeedArrange, SeedArrangeDetail, SeedArrangePlate

        rows = list(
            SeedArrangePlate.objects
            .exclude(plate_name__isnull=True).exclude(plate_name="")
        )
        if not rows:
            return []
        aids = {r.arrange_id for r in rows}
        # Seeds per (run, plate) in one grouped query rather than one per row.
        counts = {}
        for d in chunked_in(aids, lambda c: SeedArrangeDetail.objects
                            .filter(arrange_id__in=c, seed_type=False)
                            .values("arrange_id", "plate_id", "method")
                            .annotate(n=Count("seed_id", distinct=True))):
            counts.setdefault((d["arrange_id"], d["plate_id"]), {})[d["method"] or ""] = d["n"]
        runs = {
            a.arrange_id: a
            for a in chunked_in(aids, lambda c: SeedArrange.objects.filter(arrange_id__in=c))
        }

        def seeds_of(r):
            by_method = counts.get((r.arrange_id, r.plate_no), {})
            if not by_method:
                return r.real_seed_count or 0
            return by_method.get(InventoryService.BUILT_METHOD) or max(by_method.values())

        def when(r):
            run = runs.get(r.arrange_id)
            return r.update_date or r.entry_date or getattr(run, "finalized_date", None)

        out = [
            {
                "plateName": r.plate_name,
                "plateId": r.plate_id,
                "plateNo": r.plate_no,
                "arrangeId": str(r.arrange_id) if r.arrange_id else None,
                "seeds": seeds_of(r),
                # The fill of the layout that was built, falling back for older rows.
                "fillPct": ArrangementService._num(
                    r.finalized_fill_pct or r.enhanced_fill_pct or r.arrange_fill_pct),
                "plateDiameter": ArrangementService._num(
                    getattr(runs.get(r.arrange_id), "plate_diameter", None)),
                "imageUrl": (r.finalized_image_path or r.enhanced_image_path
                             or r.arrange_image_path),
                "finalizedAt": when(r).isoformat() if when(r) else None,
            }
            for r in rows
        ]
        out.sort(key=lambda x: (x["finalizedAt"] or "", x["plateName"] or ""), reverse=True)
        return out

    @staticmethod
    def names(arrange_id):
        from .models import SeedArrangePlate

        arrange_id = _uuid_or_error(arrange_id)
        return {
            str(pn): nm
            for pn, nm in SeedArrangePlate.objects.filter(arrange_id=arrange_id).values_list("plate_no", "plate_name")
        }


# ---- Implausible-measurement guard (import time) -----------------------------
# A seed whose measurement is nonsense must never reach TRN_SeedData. One that
# does is not merely an unusable row: `pack_v2._mixed_one_plate` opens every
# plate with a centre row taken from the head of the queue, and a seed too wide
# for any chord makes that row come back empty, which returns None, which breaks
# the caller's `while queue` loop. Every seed still behind it is dropped from the
# arrangement with nothing reported. On the live inventory ONE such row
# (DOMI002328, stored 1285.00 x 9.03 mm) stranded 190 of 634 eligible seeds.
#
# The engine now gates these out too (engine_runner._fits_the_plate), but that
# gate depends on the plate size and only runs at packing time. Catching them at
# IMPORT is what stops the bad row existing in the first place, and it puts the
# problem in front of the person holding the datasheet, who can still fix it.
#
# The ceiling is set from the data, not picked round. Across 1766 sound rows the
# longest side measured is 23.98 mm and nothing exceeds 30 mm; the three corrupt
# rows are 1285-1288 mm. Anything from about 30 to 1000 separates them, so 60 mm
# is 2.5x the largest real seed and 20x below the corrupt ones — it cannot reject
# a plausible seed, and a seed longer than 60 mm could not be arranged on any
# plate this shop runs (Ø90-Ø110) in any case.
MAX_SEED_MM = 60.0
# Below this a "measurement" is noise rather than a seed. The smallest real short
# side in stock is 2.01 mm, so this floor is four times clear of live data and
# only catches zeros and stray marks.
MIN_SEED_MM = 0.5


def _implausible(length, width, height):
    """Reason this row's measurements cannot be a real seed, or None if fine."""
    for name, v in (("length", length), ("width", width)):
        f = float(v)
        if f > MAX_SEED_MM:
            return ("%s %.2f mm is above the %.0f mm limit — no seed is that "
                    "large, so this is a mis-typed measurement (check the "
                    "decimal point)" % (name, f, MAX_SEED_MM))
        if f < MIN_SEED_MM:
            return "%s %.2f mm is below the %.1f mm minimum" % (name, f, MIN_SEED_MM)
    if float(height) <= 0:
        return "thickness %.2f mm must be greater than zero" % float(height)
    return None


def _blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def _dec(v):
    """Excel cell → Decimal(2dp) or None (for the decimal(18,2) columns)."""
    if _blank(v):
        return None
    try:
        return Decimal(str(v)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _int(v):
    if _blank(v):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _text(v, limit=50):
    if _blank(v):
        return None
    return str(v).strip()[:limit]


def _norm_hdr(v):
    return "" if v is None else str(v).strip().lower().replace(" ", "").replace("_", "")


def _header_row(ws):
    """The first row, normalised. Empty list if the sheet has no rows."""
    try:
        return [_norm_hdr(c) for c in
                next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    except StopIteration:
        return []


def _find_col(hdr, *prefixes):
    """Index of the first header starting with any of these, or None.

    By NAME, never by position. A datasheet that carries extra columns in front
    — a serial number, a department, anything — shifts every field along, and a
    positional reader then takes the wrong cell for every seed. That is not
    hypothetical: a sheet with four leading columns imported 35 stones with L3
    stored as the thickness, which put them all outside the packing gate.
    """
    for i, h in enumerate(hdr):
        if h and any(h.startswith(p) for p in prefixes):
            return i
    return None


def _column_map(ws):
    """Locate every field this importer understands, by header name.

    Returns a dict of role -> column index. Roles absent from the sheet are
    simply missing from the dict.
    """
    hdr = _header_row(ws)
    return {
        "batch": _find_col(hdr, "batchno", "batch"),
        "stock": _find_col(hdr, "stockno", "stokeno", "stokno", "stock", "stoke"),
        "pcs": _find_col(hdr, "pcs", "pes"),
        "cts": _find_col(hdr, "cts", "ct"),
        "l1": _find_col(hdr, "l1"),
        "w2": _find_col(hdr, "w2"),
        "l3": _find_col(hdr, "l3"),
        "w4": _find_col(hdr, "w4"),
        "length": _find_col(hdr, "length"),
        "width": _find_col(hdr, "width"),
        "height": _find_col(hdr, "height", "heigth", "thick"),
        "corner": _find_col(hdr, "crossangle", "crosscorner", "cutfacing", "crossfacing"),
        "cross": _find_col(hdr, "cross"),
    }


def _detect_layout(ws):
    """Which datasheet layout is this?

      "legacy" — BatchNo StockNo Pcs Cts Length Width Height  (+ optional Corners)
      "sides"  — BatchNo StockNo Pcs Cts  L1 W2 L3 W4  Height  (+ optional corner)

    The "sides" sheet measures a seed as FOUR AXIS-PARALLEL EDGES instead of one
    Length and one Width. Read as the legacy layout it silently stores L3 as the
    thickness — every seed then lands with a ~5-11 mm thickness and is dropped by
    the packing gate, so the import reports success and the seeds never reach a
    plate.

    Detection is by HEADER NAME and does not care where the columns sit.
    """
    cols = _column_map(ws)
    if all(cols.get(k) is not None for k in ("l1", "w2", "l3", "w4")):
        return "sides"

    # No four sides, but a thickness column that is NOT where the legacy layout
    # keeps it means this is a shape we do not understand. Reading it positionally
    # would put some other measurement into a field, so refuse and say what was
    # found rather than import something quietly wrong.
    hdr = _header_row(ws)
    h = cols.get("height")
    if h is not None and h != 6 and cols.get("length") is None:
        raise DomainError(
            "This datasheet's columns were not recognised. The thickness column "
            "%r is at position %d, and no Length/Width or L1/W2/L3/W4 columns "
            "were found. Importing it would store the wrong measurement in the "
            "wrong field, so it has been stopped. Headers found: %s."
            % (hdr[h], h + 1, ", ".join('"%s"' % (x or "(blank)") for x in hdr)))
    return "legacy"


def _sides_are_cut(sides):
    """Do these four edges describe a stone with a corner ground off?

    A matched pair means that axis is uncut; both matched means a plain
    rectangle, which needs no corner declared.
    """
    if not sides or len(sides) != 4 or any(s is None for s in sides):
        return False
    l1, w2, l3, w4 = (float(s) for s in sides)
    return abs(l1 - l3) > CUT_EPS and abs(w2 - w4) > CUT_EPS


def _header_cell(ws, idx):
    """The idx-th (0-based) header cell, or None if the sheet is that short."""
    try:
        hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return None
    return hdr[idx] if idx < len(hdr) else None


def _is_thickness(h):
    """Does this normalised header name a thickness column? Tolerates the
    'HEIGTH' misspelling that the measuring machine emits."""
    return h.startswith("height") or h.startswith("heigth") or h.startswith("thick")


class SeedImportService:
    @staticmethod
    def read_rows(file_obj):
        """Parse the datasheet into a list of dicts (no DB access). Blank rows are
        skipped. openpyxl reads a Django UploadedFile directly (it's file-like)."""
        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as exc:  # corrupt bytes / not really an .xlsx
            raise DomainError(f"Could not read the Excel file: {exc}")
        ws = wb.active
        if _detect_layout(ws) == "sides":
            return SeedImportService._read_sides(ws)
        # Column 8 has two possible jobs, told apart by its HEADER, never by
        # position. Historically it is the optional Corners outline; a datasheet
        # that instead marks cut stones with CROSS/NOCROSS puts the marker there.
        # Read positionally, a marker is parsed as coordinates and every row
        # reports "corner list could not be read" — 80 warnings on an 80-row
        # sheet, burying anything real.
        col8_is_cross = _norm_hdr(_header_cell(ws, 7)).startswith("cross")
        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            # Column 8 (Corners) is OPTIONAL and additive: a sheet with only the
            # original 7 columns pads to None here and behaves exactly as before.
            # It is deliberately NOT part of the blank-row tests below — a row
            # carrying only an outline, with no batch and no stock number, is not
            # a seed and must still be skipped.
            cells = list(raw) + [None] * 8  # pad short rows
            batch, stock, pcs, cts, length, width, height, corners = cells[:8]
            if all(_blank(c) for c in (batch, stock, pcs, cts, length, width, height)):
                continue  # entirely empty row
            if _blank(batch) and _blank(stock):
                # Not a seed: no batch AND no stock number. Skips trailing totals/
                # summary rows (which carry PCS/CTS sums but no identity) so they are
                # never imported and can't break the insert with their blank cells.
                continue
            rows.append({
                "batch_no": _text(batch),
                "stock_no": _text(stock),
                "pcs": _int(pcs),
                "cts": _dec(cts),
                "length": _dec(length),
                "width": _dec(width),
                "height": _dec(height),
                # Raw text, validated later in import_seeds (which knows L/W).
                "corners_raw": (None if col8_is_cross or _blank(corners)
                                else str(corners).strip()),
                # True when the sheet marks this stone as cut/trimmed. The marker
                # says THAT a seed is cut, never HOW MUCH, so it cannot rebuild an
                # outline — a marked seed is still packed at its full Length x
                # Width. Carried so the import can report the count.
                "is_cross": (bool(col8_is_cross) and not _blank(corners)
                             and str(corners).strip().upper() == "CROSS"),
            })
        wb.close()
        return rows

    @staticmethod
    def _read_sides(ws):
        """Read the four-edge layout:
        BatchNo StockNo Pcs Cts L1 W2 L3 W4 Height [CrossAngle].

        Length and Width are the LONGER edge of each pair (the seed's bounding
        box); the shorter edge of each pair is what the corner cut took away, and
        `sides` carries all four through so import_seeds can rebuild the outline.

        Column 10, when present, DECLARES which corner the cross is on — L1/L2/
        R1/R2 for left-top, left-bottom, right-top, right-bottom. It is the only
        source of that: the machine normalises its four sides, so they give the
        cut's size but never its corner. Sheets without the column still import
        exactly as before.
        """
        cols = _column_map(ws)

        def pick(cells, role):
            i = cols.get(role)
            return cells[i] if i is not None and i < len(cells) else None

        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            cells = list(raw) + [None] * 4
            # BY NAME, not position — the sheet may carry leading columns.
            batch, stock = pick(cells, "batch"), pick(cells, "stock")
            pcs, cts = pick(cells, "pcs"), pick(cells, "cts")
            l1, w2 = pick(cells, "l1"), pick(cells, "w2")
            l3, w4 = pick(cells, "l3"), pick(cells, "w4")
            height = pick(cells, "height")
            xangle = pick(cells, "corner")
            if xangle is None:
                xangle = pick(cells, "cross")
            if all(_blank(c) for c in (batch, stock, pcs, cts, l1, w2, l3, w4, height)):
                continue
            if _blank(batch) and _blank(stock):
                continue                      # totals / summary row, same rule as legacy
            sides = tuple(_dec(v) for v in (l1, w2, l3, w4))
            have = [s for s in sides if s is not None]
            # A PLAIN stone has one length and one width, so the sheet may give
            # LENGTH/WIDTH and leave the four sides blank — writing the same
            # number into L1 and L3 just to fill the columns would be pointless
            # and invites the mistake of putting the WIDTH into L3, which reads
            # as a 2 mm cut on a stone that has none.
            #
            # So each row chooses: four sides when they are all there (a cut
            # stone), otherwise LENGTH and WIDTH (a plain one).
            plain_l, plain_w = _dec(pick(cells, "length")), _dec(pick(cells, "width"))
            if len(have) == 4:
                box_l = max(sides[0], sides[2])
                box_w = max(sides[1], sides[3])
            else:
                box_l, box_w = plain_l, plain_w
            rows.append({
                "batch_no": _text(batch),
                "stock_no": _text(stock),
                "pcs": _int(pcs),
                "cts": _dec(cts),
                # Bounding box: the longer edge of each pair for a cut stone,
                # the plain LENGTH/WIDTH otherwise.
                "length": box_l,
                "width": box_w,
                "height": _dec(height),
                "corners_raw": None,
                "sides": sides if len(have) == 4 else None,
                # Declared cross corner, normalised. None means the cell was
                # blank or unreadable — import_seeds warns rather than guessing.
                "cut_corner": parse_cut_corner(xangle),
                "cut_corner_raw": None if _blank(xangle) else str(xangle).strip(),
            })
        return rows

    @staticmethod
    def _ensure_batches(batch_nos, entry_by=None):
        """Insert a TRN_Batch row for every BatchNo that doesn't exist yet (each
        with a new Batch_ID, ISActive=1); leave existing batches untouched.
        Returns the list of BatchNos newly inserted."""
        wanted = {b.strip() for b in batch_nos if b and b.strip()}
        if not wanted:
            return []
        existing = BatchRepository.existing_nos(wanted)
        new = sorted(wanted - existing)
        if new:
            now = datetime.now(timezone.utc)
            BatchRepository.bulk_create([
                Batch(batch_id=uuid.uuid4(), batch_no=b, is_active=True, entry_date=now, entry_by=entry_by)
                for b in new
            ])
        return new

    @staticmethod
    def import_seeds(file_obj, entry_by=None):
        """Read the uploaded file and insert only the NEW rows into TRN_SeedData.

        Duplicates are detected by StockNo: a row is skipped if its StockNo already
        exists in the table, or repeats earlier in the same sheet. Returns:
            {imported, skipped_count, skipped:[{stock_no,batch_no,reason}],
             batches_created, batches, warnings:[{stock_no,batch_no,reason}]}

        `warnings` lists rows that WERE imported but whose optional corner list
        failed validation, so they fall back to a plain Length x Width rectangle.
        """
        rows = SeedImportService.read_rows(file_obj)

        # Ensure every batch referenced in the sheet exists in TRN_Batch, then map
        # BatchNo → Batch_ID so each seed can store its Batch_ID.
        batches_created = SeedImportService._ensure_batches(
            [r["batch_no"] for r in rows], entry_by=entry_by,
        )
        batch_id_by_no = BatchRepository.id_by_no()
        existing = SeedRepository.existing_stock_nos()

        now = datetime.now(timezone.utc)
        to_create = []
        skipped = []
        warnings = []
        cross_count = 0
        seen_in_sheet = set()
        imported_by_batch = {}  # batch_no (stripped) -> how many seeds imported

        for r in rows:
            stock = (r["stock_no"] or "").strip()
            if stock:
                if stock in existing:
                    skipped.append({"stock_no": stock, "batch_no": r["batch_no"], "reason": "Entry already exists"})
                    continue
                if stock in seen_in_sheet:
                    skipped.append({"stock_no": stock, "batch_no": r["batch_no"], "reason": "Duplicate row in sheet"})
                    continue
                seen_in_sheet.add(stock)
            # A row that cannot yield all three dimensions must be REPORTED, not
            # sent to the database. Passing a NULL into a numeric column fails
            # the whole INSERT with "arithmetic overflow converting nvarchar to
            # numeric" — every other row is lost with it and the screen shows a
            # bare 500. One unusable row should cost that row and nothing else.
            missing = [n for n, v in (("length", r["length"]), ("width", r["width"]),
                                      ("height", r["height"])) if v is None]
            if missing:
                skipped.append({
                    "stock_no": stock or None, "batch_no": r["batch_no"],
                    "reason": "no %s — check the sheet has LENGTH and WIDTH for a "
                              "plain seed, or all four of L1 W2 L3 W4 for a cut one"
                              % " or ".join(missing),
                })
                continue
            # Measurements that cannot describe a real seed. Reported like any
            # other skip, so the row is visible on the import screen rather than
            # quietly entering the table and breaking a packing run later.
            nonsense = _implausible(r["length"], r["width"], r["height"])
            if nonsense:
                skipped.append({
                    "stock_no": stock or None, "batch_no": r["batch_no"],
                    "reason": nonsense,
                })
                continue
            bno = (r["batch_no"] or "").strip()
            imported_by_batch[bno] = imported_by_batch.get(bno, 0) + 1
            # Optional outline for an irregular seed. A corner list that fails
            # validation NEVER costs the seed: it is imported as a plain
            # Length x Width rectangle (the pre-existing behaviour) and the reason
            # is reported as a warning. Blank is the normal case and is silent.
            assumed = False
            if r.get("sides"):
                # Four-edge layout: rebuild the cut-corner outline from
                # L1/W2/L3/W4 for the size, and the declared column for the
                # corner. A cut stone with no corner declared CANNOT be built
                # correctly — the sides do not say which way it faces — so it is
                # imported as a plain rectangle and reported. That is safe (the
                # real stone is smaller than its rectangle and still drops into
                # the seat) but it forfeits nesting and reads coverage high.
                corner = r.get("cut_corner")
                raw_corner = r.get("cut_corner_raw")
                if corner is None and raw_corner:
                    corners_json, warn = None, (
                        "cross corner %r not recognised — expected LT, LB, RT or RB"
                        % raw_corner)
                elif corner is None and _sides_are_cut(r["sides"]):
                    # No corner declared. Assume LEFT-TOP and MARK it, so Max
                    # Coverage spends the declared stones first and falls back to
                    # these only to fill what is left. The assumption is right
                    # about two thirds of the time: rotation turns LEFT-TOP into
                    # RIGHT-BOTTOM, but LEFT-BOTTOM and RIGHT-TOP are mirror
                    # images no rotation can reach.
                    corners_json, warn = sides_to_corners(
                        *r["sides"], corner=CUT_CORNER_DEFAULT)
                    assumed = corners_json is not None
                    if warn is None:
                        warn = ("no cross corner given — assumed %s"
                                % CUT_CORNER_DEFAULT)
                else:
                    corners_json, warn = sides_to_corners(*r["sides"], corner=corner)
            else:
                corners_json, warn = validate_corners(
                    r.get("corners_raw"), r["length"], r["width"],
                )
            if warn:
                warnings.append({"stock_no": stock or None, "batch_no": r["batch_no"], "reason": warn})
            if r.get("is_cross"):
                cross_count += 1
            to_create.append(SeedData(
                seed_id=uuid.uuid4(),
                batch_id=batch_id_by_no.get((r["batch_no"] or "").strip()),
                stock_no=r["stock_no"],
                pcs=r["pcs"],
                cts=r["cts"],
                length=r["length"],
                width=r["width"],
                height=r["height"],
                # A declared corner stores the bare outline, as before. An
                # ASSUMED one stores {"pts": ..., "assumed": true} so the packer
                # can tell them apart; _poly_from_seed reads both.
                corners_json=(json.dumps({"pts": corners_json, "assumed": True})
                              if corners_json and assumed
                              else json.dumps(corners_json) if corners_json else None),
                entry_date=now,
                entry_by=entry_by,
            ))

        if to_create:
            SeedRepository.bulk_create(to_create)

        # Per-batch detail for the Form 1 result grid: every batch that received
        # seeds in this import, its imported count, and whether it was just created.
        new_set = set(batches_created)
        batches = [
            {"batch_no": (bno or None), "imported_count": n, "is_new": bno in new_set}
            for bno, n in sorted(imported_by_batch.items())
        ]

        return {
            "imported": len(to_create),
            "skipped_count": len(skipped),
            "skipped": skipped,
            "batches_created": batches_created,
            "batches": batches,
            # Rows imported as rectangles because their outline failed validation.
            # Additive key: existing clients that don't read it are unaffected.
            "warnings": warnings,
            # How many imported stones the sheet marked CROSS. Reported so the
            # count is visible at import: with no cut measurements on the sheet
            # these are packed at their full Length x Width, which makes the
            # coverage figure read high by however much was ground off them.
            # 0 for a sheet with no CROSS column.
            "cross_count": cross_count,
        }
