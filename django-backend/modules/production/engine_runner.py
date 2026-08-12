"""Thin wrapper around the existing Seed → Plate packing engine.

`arranger/engine/` holds the unmodified `pack_v2.py` + `demo_fill.py` from the
original desktop app. We add that folder to sys.path so `demo_fill`'s top-level
`import pack_v2` resolves, then drive the same pipeline the approved app uses:

    load_blocks → _mixed_landscape → _mixed_one_plate (per plate)
        → real_only_job        (Arrange: real seeds only)
        → guillo_plate_job     (Machine-Cut: recursive-guillotine dummy fill)

Each job writes its plate PNGs (and, for Compare, a per-plate .xlsx with both
images + dimension tables) into a private directory under MEDIA_ROOT and returns
a dict shaped exactly like the frontend `Job` contract.
"""
import json
import math
import os
import sys

# Headless matplotlib — MUST be set before the engine imports matplotlib.
os.environ.setdefault("MPLBACKEND", "Agg")
os.environ.setdefault("PYTHONIOENCODING", "utf-8")

_ENGINE_DIR = os.path.join(os.path.dirname(__file__), "engine")
if _ENGINE_DIR not in sys.path:
    sys.path.insert(0, _ENGINE_DIR)

import pack_v2 as P  # noqa: E402
import demo_fill as D  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.drawing.image import Image as XLImage  # noqa: E402
from openpyxl.styles import Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
import matplotlib.pyplot as _plt  # noqa: E402
from matplotlib.patches import Circle as _Circle, Patch as _Patch, Rectangle as _Rect  # noqa: E402

SHAPE_SETS = {
    "all": {"square", "rectangle"},
    "square": {"square"},
    "rectangle": {"rectangle"},
}

_HDR_FILL = PatternFill("solid", fgColor="1F4E78")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_TH = Side(style="thin", color="E3E8EE")
_BORDER = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
# "Cut off" removed: seeds arrive already cut and the engine never cuts one, so
# the column was "—" on every row. dim_rows() still returns `cut` — the stored
# TRN_SeedArrangeDetails columns and older exports read it — it is simply no
# longer a column in the sheet.
_COLS = ["#", "Type", "ID / Stock", "W × H (mm)", "Thick (mm)", "Shape", "Center"]


def _num(params, key):
    """Read a required numeric param straight from the user's form value — no
    default. Missing/blank/non-numeric raises (the API validates first, so this
    only fires on a genuinely bad request)."""
    return float(params[key])


def _apply_globals(params):
    """Set the engine's module globals from the request params (every value is
    the one the user entered in the form). Returns the usable-circle radius R."""
    plate_d = _num(params, "plateD")
    margin = _num(params, "margin")
    # Margin is the clear space kept AROUND the whole plate — the same margin on every
    # side. Seeds may only go inside a circle that is `margin` smaller on each side, so
    # the diameter loses 2 × margin. e.g. plateD 90, margin 5 → usable Ø80, R 40. This
    # matches how the margin is physically left on the real plate (a ring all around).
    usable = plate_d - 2 * margin
    if usable <= 0:
        raise ValueError("2 × margin must be smaller than the plate diameter")
    P.PLATE_D = plate_d
    P.USABLE_D = usable
    P.R = usable / 2.0
    P.S2 = (usable / math.sqrt(2)) / 2.0
    P.INSCRIBED = usable / math.sqrt(2)
    P.T_LO = _num(params, "tLo")
    P.T_HI = _num(params, "tHi")
    P.GRID = _num(params, "grid")
    P.CLEARANCE = _num(params, "clearance")
    P.MINSEED = _num(params, "minSeed")
    return P.R


def _fmt_dim(v):
    """Exact stored measurement as text — no round-off. Values are decimal(18,2),
    so two places captures them exactly; trailing zeros are trimmed (18.40 → 18.4,
    27.00 → 27) so the number reads the same as it does in the source sheet."""
    return f"{float(v):.2f}".rstrip("0").rstrip(".")


def seed_cut(p):
    """How much of THIS SEED was cut off, in mm2, and as a share of the seed.

    Max Coverage tiles the plate into seats sized from the pool's MEDIAN stone,
    then drops a stone into each. Two areas are therefore in play and they are
    not interchangeable:

        p["nomarea"]  the seat's full box before the plate edge clipped it
        p["area"]     what is left of that seat
        rawL x rawW   the stone's own measured size

    The cut reported against a stock number has to be about the STONE: a stone
    is only cut when its seat is smaller than the stone itself, and then only by
    the difference. Measuring `nomarea - area` instead describes the SEAT, which
    is why two completely different stones in mirror-image seats reported the
    identical cut, and why a 53.7 mm2 stone could be shown losing 90 mm2 — more
    than the whole stone.

    A stone smaller than its seat is NOT trimmed: it fits whole and the seat
    simply has slack, so the cut is zero.
    """
    area = p.get("area")
    if area is None:
        return 0.0, 0.0
    if p.get("irregular"):
        # The stone was placed from its MEASURED outline, and an outline is only
        # ever placed if it fits the seat whole — so nothing was cut off here.
        # Its rawL x rawW is the bounding BOX, which is larger than the stone by
        # the corner the manufacturer already ground away; measuring against that
        # box would report that pre-existing corner as if the plate edge had just
        # trimmed it, on seeds sitting nowhere near the edge.
        return 0.0, 0.0
    dw = p.get("rawL") if p.get("rawL") is not None else p.get("w")
    dh = p.get("rawW") if p.get("rawW") is not None else p.get("h")
    try:
        seed_area = float(dw) * float(dh)
    except (TypeError, ValueError):
        return 0.0, 0.0
    if seed_area <= 0:
        return 0.0, 0.0
    cut = max(0.0, seed_area - float(area))
    return cut, 100.0 * cut / seed_area


def dim_rows(placed):
    """Per-seed dimension rows in the frontend `DimRow` shape."""
    rows = []
    for p in placed:
        is_real = not p.get("filler")
        w, h = p["w"], p["h"]                            # seat box (position/plate geometry)
        # Size shown = the seed's REAL stone measurement from the import when available
        # (Max Coverage stores it as rawL/rawW); other methods place seeds at their true
        # size, so w/h already IS the real size. Shape follows the size that is shown.
        dw = p.get("rawL") if p.get("rawL") is not None else w
        dh = p.get("rawW") if p.get("rawW") is not None else h
        size = f"{_fmt_dim(dw)} × {_fmt_dim(dh)}"
        cut = "—"                                       # "how much was cut off" column
        if is_real and "clipped" in p:                  # Enhanced Version seat (whole or trimmed)
            shape = "square" if abs(dw - dh) <= 0.05 * max(dw, dh) else "rect"
            # "Trimmed" follows whether the STONE actually lost material, not
            # whether its seat happened to touch the plate edge — a stone that
            # fits inside a clipped seat is still whole.
            cut_area, pct = seed_cut(p)
            if cut_area > 0.005:
                typ = "TRIMMED seed"
                cut = f"{cut_area:.0f} mm² ({pct:.0f}%)"
            else:
                typ = "Real seed (whole)"
        elif is_real:
            shape = "square" if abs(w - h) <= 0.05 * max(w, h) else "rect"
            typ = "Real seed"
        elif p.get("kind") == "enhanced":               # Enhanced Version: placeholder dummy (Ø only)
            shape = "irregular"
            typ = "Placeholder dummy"
            size = f"Ø {p.get('diam', 0):.1f} mm"
        else:
            shape, typ = "irregular", "Irregular dummy"
        rows.append({
            "type": typ,
            "stock": str(p["stock"]),
            "size": size,
            "thick": f"{p.get('H', 0.0):.2f}",
            "shape": shape,
            "cut": cut,
            "center": f"{p['x'] + w / 2:.1f}, {p['y'] + h / 2:.1f}",
            "real": is_real,
        })
    return rows


# Type label per finalize outcome, for the FINALIZED detail table.
_FINAL_TYPE = {"kept": "Real seed", "alt": "Replacement", "dummy": "Dummy fill", "empty": "Empty"}


def _final_dim_rows(final_placed):
    """Detail rows for the FINALIZED plate — same 7-column shape as dim_rows, but the
    Type reflects the finalize outcome: kept real / replacement real / dummy fill / empty."""
    rows = []
    for p in final_placed:
        kind = p.get("kind", "kept")
        w, h = p["w"], p["h"]
        if kind == "dummy":
            shape = "irregular"
        elif kind == "empty":
            shape = "-"
        else:
            shape = "square" if abs(w - h) <= 0.05 * max(w, h) else "rect"
        rows.append({
            "type": _FINAL_TYPE.get(kind, "Real seed"),
            "stock": str(p.get("stock", "")),
            "size": f"{w:.1f} × {h:.1f}",
            "thick": f"{p.get('H', 0.0):.2f}",
            "shape": shape,
            "center": f"{p['x'] + w / 2:.1f}, {p['y'] + h / 2:.1f}",
            "real": kind in ("kept", "alt"),
        })
    return rows


def _write_table(ws, r0, c0, title, rows):
    ws.cell(r0, c0, title).font = Font(bold=True)
    for j, col in enumerate(_COLS):
        cell = ws.cell(r0 + 1, c0 + j, col)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.border = _BORDER
    for k, row in enumerate(rows, 1):
        vals = [k, row["type"], row["stock"], row["size"], row["thick"], row["shape"], row["center"]]
        for j, val in enumerate(vals):
            cell = ws.cell(r0 + 1 + k, c0 + j, val)
            cell.border = _BORDER
            if j == 3:
                cell.font = Font(bold=True)


def _write_single_xlsx(path, plate_no, heading, table_title, img_path, rows):
    """Per-plate workbook for a SINGLE-stage result (Arrange / Machine-Cut / Enhanced):
    the plate image + one seed detail table (same columns as the on-screen table)."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Plate_{plate_no:02d}"
    ws.cell(1, 1, heading).font = Font(bold=True, size=13, color="1F4E78")
    if img_path and os.path.exists(img_path):
        im = XLImage(img_path)
        # Keep the image's real proportions (the enhanced plate is wide: plate + seed list).
        ratio = (im.height / im.width) if im.width else 1.0
        im.width = 720
        im.height = int(round(720 * ratio))
        ws.add_image(im, "A3")
    _write_table(ws, 20, 1, f"{table_title} — {len(rows)} seeds", rows)
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 14
    wb.save(path)


def _write_compare_xlsx(path, plate_no, panels):
    """Compare workbook: for each SELECTED method a heading + plate image + full seed table,
    laid out left-to-right. panels = list of (label, image_path, rows)."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Plate_{plate_no:02d}"
    ws.cell(1, 1, f"Plate {plate_no:02d} — Comparison ({len(panels)} methods)").font = Font(bold=True, size=13, color="1F4E78")
    col = 1
    for label, img_path, rows in panels:
        ws.cell(2, col, label).font = Font(bold=True, size=11, color="1F4E78")
        if img_path and os.path.exists(img_path):
            im = XLImage(img_path)
            ratio = (im.height / im.width) if im.width else 1.0
            im.width = 430
            im.height = int(round(430 * ratio))
            ws.add_image(im, ws.cell(3, col).coordinate)
        _write_table(ws, 20, col, f"{label} — {len(rows or [])} seeds", rows or [])
        for j in range(8):
            ws.column_dimensions[get_column_letter(col + j)].width = 13
        col += 10   # 8-col table + a 1-col gap
    wb.save(path)


def _poly_from_seed(s):
    """TRN_SeedData.CornersJSON -> ([(x, y), ...], assumed) or (None, False).

    Two stored shapes are accepted:

      [[x, y], ...]                      an outline built from a DECLARED corner
      {"pts": [[x, y], ...], "assumed": true}
                                         the datasheet left the cross corner
                                         blank, so LEFT-TOP was assumed

    The flag matters because an assumed corner is right only about two thirds of
    the time — rotation turns LEFT-TOP into RIGHT-BOTTOM, but LEFT-BOTTOM and
    RIGHT-TOP are mirror images it can never reach. Marking them lets the packer
    spend the known-good stones first and fall back to the assumed ones only to
    fill what is left.

    Stored data is trusted only as far as it parses: the outline was validated by
    shapes.validate_corners() at import time, but a hand-edited row must degrade
    to the plain rectangle rather than break a packing run.
    """
    raw = getattr(s, "corners_json", None)
    if not raw:
        return None, False
    try:
        blob = json.loads(raw)
        assumed = False
        if isinstance(blob, dict):
            assumed = bool(blob.get("assumed"))
            blob = blob.get("pts") or []
        pts = [(float(x), float(y)) for x, y in blob]
    except (TypeError, ValueError, AttributeError, json.JSONDecodeError):
        return None, False
    return (pts, assumed) if len(pts) >= 3 else (None, False)


def _blocks_from_seeds(seeds, shapes, square_tol):
    """Apply the SAME thickness-gate + shape filter as pack_v2.load_blocks, but to
    SeedData rows (objects with .length/.width/.height/.stock_no/.cts) instead of an
    Excel file. Reads the P.T_LO/P.T_HI globals set by _apply_globals."""
    out = []
    for s in seeds:
        L = float(s.length) if s.length is not None else None
        W = float(s.width) if s.width is not None else None
        H = float(s.height) if s.height is not None else None
        if None in (L, W, H):
            continue
        if not (P.T_LO <= H <= P.T_HI):          # thickness first
            continue
        sh = "square" if abs(L - W) <= square_tol * max(L, W) else "rectangle"
        if sh not in shapes:                      # shape filter (secondary)
            continue
        blk = {"stock": s.stock_no, "cts": float(s.cts or 0.0), "L": L, "W": W, "H": H, "shape": sh}
        # Optional true outline for an irregular seed, as [(x, y), ...] in mm.
        # Additive and nullable: only Max Coverage reads it. The Arrange packer
        # (engine/pack_v2.py) reads L/W/H/stock/cts/shape only, so it is unaffected.
        poly, assumed = _poly_from_seed(s)
        if poly:
            blk["poly"] = poly
            if assumed:
                # Cross corner was blank on the datasheet and LEFT-TOP was
                # assumed. Max Coverage places these AFTER the stones whose
                # corner was declared.
                blk["corner_assumed"] = True
        out.append(blk)
    return out


def load_blocks_from_db(shapes, square_tol, batches=None):
    """Read the seeds to arrange straight from the TRN_SeedData table (optionally
    filtered to one or more Batch_IDs). This is the source of truth — no Excel file
    needed. `batches` is a list of Batch_IDs (uuids); empty/None = all seeds."""
    from .models import SeedData

    qs = SeedData.objects.all()
    if batches:
        qs = qs.filter(batch_id__in=list(batches))
    return _blocks_from_seeds(qs.iterator(), shapes, square_tol)


def run(action, params, out_dir, media_base, progress=lambda p: None):
    """Run one job end to end, sourcing seeds from the TRN_SeedData table.

    action     — "arrange" | "machinefill" | "compare" | "enhanced"
    params     — the JobParams dict (may include `batch` to filter by BatchNo)
    out_dir    — private filesystem dir for this job's artifacts (under MEDIA_ROOT)
    media_base — URL prefix that maps to out_dir (e.g. /media/jobs/job-abc123)
    progress   — callback(int 0..100) invoked after each plate

    Returns a dict: {plates, pairs, arrangeAvg, machineAvg, enhancedAvg}.
    """
    R = _apply_globals(params)
    shape_set = SHAPE_SETS[params["shape"]]
    square_tol = _num(params, "squareTol")
    min_seed = _num(params, "minSeed")
    plate_d = _num(params, "plateD")
    raw_batches = params.get("batches") or []
    if isinstance(raw_batches, str):  # tolerate a single id sent as a string
        raw_batches = [raw_batches]
    batches = [str(b).strip() for b in raw_batches if str(b).strip()]

    blocks = load_blocks_from_db(shape_set, square_tol, batches)

    # Exclude the user's stock numbers AND the occupied (finalized) seeds frozen in
    # occupiedExclude — both kept in the stored params so a re-generate matches exactly.
    exclude_raw = ((params.get("exclude") or "") + " " + (params.get("occupiedExclude") or "")).replace(",", " ").split()
    if exclude_raw:
        excluded = {s.strip() for s in exclude_raw if s.strip()}
        blocks = [b for b in blocks if str(b["stock"]).strip() not in excluded]

    queue = P._mixed_landscape(blocks)
    real_plates = []
    while queue:
        rp = P._mixed_one_plate(queue)
        if not rp:
            break
        real_plates.append(rp)

    total = len(real_plates) or 1
    # Compare = user picks WHICH methods to compare (2 or 3 of arrange/machinefill/enhanced).
    # Kept in canonical order so the panels always read arrange → machine → max-coverage.
    if action == "compare":
        raw_methods = params.get("methods") or ["arrange", "machinefill"]
        if isinstance(raw_methods, str):
            raw_methods = [raw_methods]
        chosen = set(raw_methods)
        methods = [m for m in ("arrange", "machinefill", "enhanced") if m in chosen]
        if len(methods) < 2:
            methods = ["arrange", "machinefill"]
    else:
        methods = [action]
    do_arrange = "arrange" in methods
    do_machine = "machinefill" in methods
    do_enhanced = "enhanced" in methods

    arrange_dir = os.path.join(out_dir, "arrange")
    machine_dir = os.path.join(out_dir, "machinefill")
    enhanced_dir = os.path.join(out_dir, "enhanced")
    excel_dir = os.path.join(out_dir, "excel")
    for d in (arrange_dir, machine_dir, enhanced_dir, excel_dir):
        os.makedirs(d, exist_ok=True)

    # ---- Max Coverage: ONE shared stone pool for the whole run ------------------
    # Max Coverage tiles far denser than Arrange packs, so a plate's Arrange-sized
    # allocation cannot fill its grid. Every Max Coverage plate therefore draws from
    # one ordered pool of all eligible stones and takes exactly what its grid needs;
    # what it places is removed before the next plate, so a stone is still used once
    # in the whole arrangement. The plates come out full and the run needs fewer.
    #
    # This applies to EVERY run — standalone AND Compare — so Max Coverage has one
    # algorithm and one output no matter which screen launched it. Consequence for
    # Compare: its panel for plate N holds the stones Max Coverage chose for that
    # plate, which are not the same stones as Arrange's panel. The comparison is
    # therefore method-vs-method over the whole run (how full, how many plates),
    # not a per-plate like-for-like of identical stones.
    enh_pool = [s for p in real_plates for s in p] if do_enhanced else []

    arrange_plates, machine_plates, enhanced_plates, pairs = [], [], [], []
    per_plate_real = []   # (plate_no, [real seed dicts])  → TRN_SeedArrangeDetails SeedType 0
    per_plate_dummy = []  # (plate_no, [dummy filler dicts]) → TRN_DummySeedData + SeedType 1
    # EVERY method's real placements, so a Compare run stores a seed list per method
    # (they differ: Max Coverage packs more, trimmed, seats than Arrange).
    per_method_real = {"arrange": [], "machinefill": [], "enhanced": []}
    plate_meta = []       # per-plate summary + image paths → TRN_SeedPlate

    for idx, real in enumerate(real_plates, 1):
        arows = mrows = erows = None
        afill = mfill = efill = 0.0
        e_made = False   # did Max Coverage produce a plate (pool not yet exhausted)?
        m_dummy = 0
        real_count = dummy_count = 0
        apath = os.path.join(arrange_dir, f"plate_{idx:02d}.png")
        mpath = os.path.join(machine_dir, f"plate_{idx:02d}.png")
        epath = os.path.join(enhanced_dir, f"plate_{idx:02d}.png")

        if do_arrange:
            _, aplaced, afill, _, _ = D.real_only_job((real, idx, plate_d, R, min_seed, apath))
            arows = dim_rows(aplaced)
            per_plate_real.append((idx, aplaced))
            per_method_real["arrange"].append((idx, aplaced))
            real_count = len(aplaced)
            arrange_plates.append({
                "plateNo": idx,
                "fillPct": round(afill, 1),
                "dummyCount": 0,
                "imageUrl": f"{media_base}/arrange/plate_{idx:02d}.png",
                "seeds": arows,
            })

        if do_machine:
            _, mplaced, mfill, _, _ = D.guillo_plate_job((real, idx, plate_d, R, min_seed, mpath))
            mrows = dim_rows(mplaced)
            reals_only = [p for p in mplaced if not p.get("filler")]
            per_method_real["machinefill"].append((idx, reals_only))
            if not do_arrange:  # machinefill-only: take the real seeds (skip fillers)
                per_plate_real.append((idx, reals_only))
                real_count = len(reals_only)
            per_plate_dummy.append((idx, [p for p in mplaced if p.get("filler")]))
            m_dummy = sum(1 for p in mplaced if p.get("filler"))
            dummy_count = m_dummy
            machine_plates.append({
                "plateNo": idx,
                "fillPct": round(mfill, 1),
                "dummyCount": sum(1 for p in mplaced if p.get("filler")),
                "imageUrl": f"{media_base}/machinefill/plate_{idx:02d}.png",
                "seeds": mrows,
            })

        if do_enhanced and enh_pool:
            # Feed the whole remaining pool; the job fills its grid and stops. Remove
            # what it placed so no later plate can reuse any of it.
            _, eplaced, efill, _, _ = D.enhanced_plate_job((list(enh_pool), idx, plate_d, R, min_seed, epath))
            _spent = {p["stock"] for p in eplaced}
            enh_pool = [s for s in enh_pool if s["stock"] not in _spent]
            e_made = True
            erows = dim_rows(eplaced)
            ereals = [p for p in eplaced if not p.get("filler")]
            per_method_real["enhanced"].append((idx, ereals))
            if not (do_arrange or do_machine):   # only the SAVED method feeds per_plate_real
                per_plate_real.append((idx, ereals))
                real_count = len(ereals)
                dummy_count = 0
            if action != "compare":              # standalone Max Coverage → its own per-plate xlsx
                exlsx_name = f"plate_{idx:02d}.xlsx"
                _write_single_xlsx(
                    os.path.join(excel_dir, exlsx_name), idx,
                    f"Plate {idx:02d} — Max Coverage", "MAX COVERAGE", epath, erows,
                )
                enhanced_plates.append({
                    "plateNo": idx, "fillPct": round(efill, 1), "dummyCount": 0,
                    "imageUrl": f"{media_base}/enhanced/plate_{idx:02d}.png",
                    "seeds": erows, "exportUrl": f"{media_base}/excel/{exlsx_name}",
                })
            else:
                enhanced_plates.append({
                    "plateNo": idx, "fillPct": round(efill, 1), "dummyCount": 0,
                    "imageUrl": f"{media_base}/enhanced/plate_{idx:02d}.png", "seeds": erows,
                })

        if action == "compare":
            # One panel per SELECTED method (canonical order); combined per-plate Excel.
            specs = []  # (method, label, fill, url, local_path, rows, dummyCount)
            if do_arrange:
                specs.append(("arrange", "Arrange", afill, f"{media_base}/arrange/plate_{idx:02d}.png", apath, arows, 0))
            if do_machine:
                specs.append(("machinefill", "Machine-Cut Fill", mfill, f"{media_base}/machinefill/plate_{idx:02d}.png", mpath, mrows, m_dummy))
            if e_made:
                specs.append(("enhanced", "Max Coverage", efill, f"{media_base}/enhanced/plate_{idx:02d}.png", epath, erows, 0))
            xlsx_name = f"plate_{idx:02d}.xlsx"
            _write_compare_xlsx(os.path.join(excel_dir, xlsx_name), idx, [(s[1], s[4], s[5]) for s in specs])
            pairs.append({
                "plateNo": idx,
                "methods": [s[0] for s in specs],
                "panels": [{"method": s[0], "label": s[1], "fillPct": round(s[2], 1),
                            "imageUrl": s[3], "seeds": s[5], "dummyCount": s[6]} for s in specs],
                "exportUrl": f"{media_base}/excel/{xlsx_name}",
            })

        # Coverage area from the plate image titles: total = usable circle area (π·R²),
        # covered = fill% × total. R is the usable radius applied in _apply_globals.
        total_area = round(math.pi * R * R, 2)
        plate_meta.append({
            "plateNo": idx,
            "arrangeImg": f"{media_base}/arrange/plate_{idx:02d}.png" if do_arrange else None,
            "machineImg": f"{media_base}/machinefill/plate_{idx:02d}.png" if do_machine else None,
            "enhancedImg": f"{media_base}/enhanced/plate_{idx:02d}.png" if e_made else None,
            "excelPath": f"{media_base}/excel/plate_{idx:02d}.xlsx" if action in ("compare", "enhanced") else None,
            "arrangeFill": round(afill, 2) if do_arrange else None,
            "machineFill": round(mfill, 2) if do_machine else None,
            "enhancedFill": round(efill, 2) if e_made else None,
            "totalArea": total_area,
            "arrangeCovered": round(total_area * afill / 100.0, 2) if do_arrange else None,
            "machineCovered": round(total_area * mfill / 100.0, 2) if do_machine else None,
            "realCount": real_count,
            "dummyCount": dummy_count,
        })
        progress(min(99, round(idx / total * 100)))

        # Standalone Max Coverage: once every stone is placed there is nothing left
        # for further plates, so stop rather than emitting empty ones.
        if action == "enhanced" and not enh_pool:
            break

    arrange_avg = round(sum(p["fillPct"] for p in arrange_plates) / len(arrange_plates), 1) if arrange_plates else 0.0
    machine_avg = round(sum(p["fillPct"] for p in machine_plates) / len(machine_plates), 1) if machine_plates else 0.0
    enhanced_avg = round(sum(p["fillPct"] for p in enhanced_plates) / len(enhanced_plates), 1) if enhanced_plates else 0.0

    if action == "arrange":
        plates = arrange_plates
    elif action == "machinefill":
        plates = machine_plates
    elif action == "enhanced":
        plates = enhanced_plates
    else:
        plates = []

    saved = _save_arrangement(action, params, per_plate_real, per_plate_dummy, batches,
                              arrange_avg if do_arrange else (enhanced_avg if do_enhanced else machine_avg),
                              per_method_real=per_method_real)
    if saved.get("arrangeId"):
        # The plate NAME is assigned later, at finalize time (not at generation) — the
        # user confirms the selection first, then names the plate. So rows are created
        # with PlateName NULL and updated by finalize_arrangement.
        _save_plate_rows(saved["arrangeId"], plate_meta)

    return {
        "plates": plates,
        "pairs": pairs,
        "arrangeAvg": arrange_avg,
        "machineAvg": machine_avg,
        "enhancedAvg": enhanced_avg,
        "arrangeId": saved.get("arrangeId"),
        "seedsStored": saved.get("seedsStored", 0),
        "dummiesStored": saved.get("dummiesStored", 0),
        # How many seeds passed the selected criteria (shape / thickness / batch /
        # excludes). 0 → nothing matched, so no plate was generated → UI shows a notice.
        "seedsMatched": len(blocks),
    }


def _save_arrangement(action, params, per_plate_real, per_plate_dummy, batches, avg, per_method_real=None):
    """Persist an arrangement:
      * TRN_SeedArrange           — 1 header row.
      * TRN_SeedArrangeDetails    — 1 row per placed seed (SeedType 0 = real, 1 = dummy).
      * TRN_SeedData.ISUsed       — set 1 for each placed real seed.
      * TRN_DummySeedData         — 1 row per dummy filler (Machine-Cut).
    Returns {arrangeId, seedsStored, dummiesStored}."""
    import uuid
    from datetime import date, datetime, timezone
    from decimal import Decimal

    from .models import DummySeedData, SeedArrange, SeedArrangeDetail, SeedData

    _ZERO = Decimal("0.00")
    # SQL Server caps a statement at 2100 parameters. SeedArrangeDetail now binds 11 columns
    # per row, so 500-row batches (5500 params) spill onto a temp-table path that truncates.
    _DETAIL_BATCH = 150

    if not per_plate_real and not per_plate_dummy:
        return {"arrangeId": None, "seedsStored": 0, "dummiesStored": 0}

    # The input criteria this arrangement was generated with, each in its own column so
    # we can track/query which seeds were arranged under which parameters.
    def _f(key):
        try:
            return float(params.get(key))
        except (TypeError, ValueError):
            return None

    # StockNo → Seed_ID (the placed seeds carry 'stock' = StockNo).
    stock_ids = {
        (s or "").strip(): sid
        for sid, s in SeedData.objects.values_list("seed_id", "stock_no")
        if s and s.strip()
    }

    header = SeedArrange.objects.create(
        arrange_id=uuid.uuid4(),
        is_active=True,
        plate_no=len(per_plate_real),          # number of plates in the arrangement
        average=round(float(avg), 2),           # average fill %, e.g. 95.50
        # Input criteria used for this run (own columns for easy querying):
        mode=(params.get("mode") or None),
        shape=(params.get("shape") or None),
        square_tol=_f("squareTol"),
        thickness_min=_f("tLo"),
        thickness_max=_f("tHi"),
        plate_diameter=_f("plateD"),
        margin=_f("margin"),
        min_filler_size=_f("minSeed"),
        exclude_stocks=(params.get("exclude") or "") or None,
        batches=(",".join(batches) if batches else None),
        entry_date=date.today(),
        # Auto-finalize the system's arrangement by default (user can override).
        is_finalized=True,
        finalized_by_user=False,   # 0 = system/auto; set 1 when the user finalizes
        finalized_date=datetime.now(timezone.utc),
    )

    # One detail row per placed seed PER METHOD, so Arrange and Max Coverage each keep their
    # own seed list (their placements differ). Falls back to the single legacy list.
    by_method = per_method_real or {}
    if not any(by_method.values()):
        by_method = {(params.get("_method") or action): per_plate_real}

    details = []
    for method, plates in by_method.items():
        for plate_no, seeds in plates:
            for sd in seeds:
                sid = stock_ids.get(str(sd.get("stock", "")).strip())
                if sid is None:
                    continue
                # Max Coverage trims boundary seats; record how much came off. 0.00 = whole
                # seed. NOTE: never leave these NULL — the SQL Server driver rejects a
                # bulk_create batch that mixes NULL and non-NULL decimals for one column.
                cut = pct = _ZERO
                cut_f, pct_f = seed_cut(sd)
                if cut_f > 0.005:
                    cut = Decimal(str(round(cut_f, 2)))
                    pct = Decimal(str(round(pct_f, 2)))
                details.append(SeedArrangeDetail(
                    detail_id=uuid.uuid4(),
                    arrange_id=header.arrange_id,
                    seed_type=False,        # 0 = real seed
                    is_recommended=True,    # system placed it
                    is_final=True,          # auto-final by default
                    seed_id=sid,
                    plate_id=plate_no,
                    method=method,
                    cut_area_mm2=cut,
                    cut_pct=pct,
                ))
    if details:
        SeedArrangeDetail.objects.bulk_create(details, batch_size=_DETAIL_BATCH)

    now = datetime.now(timezone.utc)

    # NOTE: TRN_SeedData.ISUsed is intentionally NOT written here (feature rolled
    # back per request). Real placements are recorded only in TRN_SeedArrangeDetails.

    # Dummy (filler) seeds → TRN_DummySeedData + a SeedType=1 detail each.
    dummy_rows, dummy_details = [], []
    # A dummy gets a Batch_ID only when the arrangement was for a single batch.
    batch_id = batches[0] if batches and len(batches) == 1 else None
    for plate_no, dummies in per_plate_dummy:
        for d in dummies:
            did = uuid.uuid4()
            dummy_rows.append(DummySeedData(
                seed_id=did,
                batch_id=batch_id,
                stock_no=(str(d.get("stock", "")).strip()[:50] or None),
                pcs=1,
                cts=0,
                length=round(float(d.get("w", 0) or 0), 2),
                width=round(float(d.get("h", 0) or 0), 2),
                height=round(float(d.get("H", 0) or 0), 2),
                entry_date=now,
            ))
            dummy_details.append(SeedArrangeDetail(
                detail_id=uuid.uuid4(),
                arrange_id=header.arrange_id,
                seed_type=True,     # 1 = dummy
                is_recommended=True,
                is_final=True,
                seed_id=did,
                plate_id=plate_no,
                method="machinefill",   # only Machine-Cut Fill produces dummy fillers
                cut_area_mm2=_ZERO,
                cut_pct=_ZERO,
            ))
    if dummy_rows:
        DummySeedData.objects.bulk_create(dummy_rows, batch_size=_DETAIL_BATCH)
        SeedArrangeDetail.objects.bulk_create(dummy_details, batch_size=_DETAIL_BATCH)

    return {
        "arrangeId": str(header.arrange_id),
        "seedsStored": len(details),
        "dummiesStored": len(dummy_rows),
    }


def _save_plate_rows(arrange_id, plate_meta, plate_name=None, plate_id=None):
    """Insert one TRN_SeedPlate row per plate (image paths, fill %s + covered/total
    area, counts). `plate_name` is the name the USER picked from the MST_SeedPlate master
    (stored in PlateName), and `plate_id` its matching master Plate_ID. Every plate row
    for this run is stamped with them. FinalizedImagePath is filled later by
    generate_final."""
    import uuid
    from datetime import datetime, timezone

    from .models import SeedArrangePlate

    now = datetime.now(timezone.utc)
    rows = []
    for m in plate_meta:
        rows.append(SeedArrangePlate(
            arrange_plate_id=uuid.uuid4(),
            arrange_id=arrange_id,
            plate_no=m["plateNo"],
            plate_id=plate_id,
            plate_name=plate_name,
            arrange_image_path=m["arrangeImg"],
            machine_cut_image_path=m["machineImg"],
            enhanced_image_path=m.get("enhancedImg"),
            excel_path=m["excelPath"],
            arrange_fill_pct=m["arrangeFill"],
            machine_fill_pct=m["machineFill"],
            enhanced_fill_pct=m.get("enhancedFill"),
            total_area_mm2=m.get("totalArea"),
            arrange_covered_mm2=m.get("arrangeCovered"),
            machine_covered_mm2=m.get("machineCovered"),
            real_seed_count=m["realCount"],
            dummy_count=m["dummyCount"],
            entry_date=now,
        ))
    # Insert in groups that share the same NULL pattern. mssql-django builds one
    # multi-row INSERT per batch and infers each parameter's type from the FIRST
    # row, so a batch mixing NULL and non-NULL for the same column misbinds the
    # rest — a decimal column ends up receiving the neighbouring path string and
    # SQL Server raises "Arithmetic overflow converting nvarchar to numeric".
    # A Compare run hits this whenever Max Coverage finishes in fewer plates than
    # Arrange: the trailing plates have no enhanced image/fill while the leading
    # ones do. Grouping keeps every batch internally consistent.
    if rows:
        _NULLABLE = ("arrange_image_path", "machine_cut_image_path", "enhanced_image_path",
                     "excel_path", "arrange_fill_pct", "machine_fill_pct", "enhanced_fill_pct",
                     "total_area_mm2", "arrange_covered_mm2", "machine_covered_mm2",
                     "plate_id", "plate_name")

        def _null_sig(r):
            return tuple(getattr(r, f, None) is None for f in _NULLABLE)

        groups = {}
        for r in rows:
            groups.setdefault(_null_sig(r), []).append(r)
        for group in groups.values():
            SeedArrangePlate.objects.bulk_create(group, batch_size=500)


# ---------------------------------------------------------------------------
# Finalized-arrangement generation: re-derive the ORIGINAL layout (deterministic
# re-pack keeps every seed in its exact position), drop the alternate seeds into
# the removed slots, and render new plate images.
# ---------------------------------------------------------------------------

def _render_final_circle(placed, pi, R, fill, path):
    """Render a finalized plate: kept real seeds (blue), user-placed replacement seeds
    (gold, orange edge), and empty seats the user left unfilled (dashed outline)."""
    plate = P.PLATE_D
    fig, ax = _plt.subplots(figsize=(7.5, 7.5))
    ax.add_patch(_Circle((0, 0), plate / 2, fc="#e9e9ec", ec="#888", lw=1.5, zorder=0))
    ax.add_patch(_Circle((0, 0), R, fc="none", ec="#c0392b", lw=1.3, ls="--", zorder=1))
    for p in placed:
        kind = p.get("kind", "kept")
        if kind == "empty":
            ax.add_patch(_Rect((p["x"], p["y"]), p["w"], p["h"], facecolor="none",
                               edgecolor="#999", linewidth=1.0, linestyle="--", zorder=2))
            ax.text(p["x"] + p["w"] / 2, p["y"] + p["h"] / 2, "empty", ha="center", va="center",
                    fontsize=6, color="#777", style="italic", zorder=3)
            continue
        style = {"alt": ("#f1c40f", "#d35400", None), "dummy": ("#e74c3c", "#922b21", "///")}.get(
            kind, ("#2e86c1", "white", None))
        fc, ec, hatch = style
        ax.add_patch(_Rect((p["x"], p["y"]), p["w"], p["h"], facecolor=fc, edgecolor=ec,
                           linewidth=0.8, hatch=hatch, zorder=2))
        ax.text(p["x"] + p["w"] / 2, p["y"] + p["h"] / 2, f"{p['stock']}\n{p['w']:.1f}×{p['h']:.1f}",
                ha="center", va="center", fontsize=5.5,
                color="black" if kind == "alt" else "white", zorder=3)
    lim = plate / 2 + 4
    ax.set_xlim(-lim, lim); ax.set_ylim(-lim, lim + 8); ax.set_aspect("equal"); ax.axis("off")
    n_alt = sum(1 for p in placed if p.get("kind") == "alt")
    n_dummy = sum(1 for p in placed if p.get("kind") == "dummy")
    n_empty = sum(1 for p in placed if p.get("kind") == "empty")
    ax.legend(handles=[
        _Patch(fc="#2e86c1", label="kept real"),
        _Patch(fc="#f1c40f", ec="#d35400", label=f"replacement real ({n_alt})"),
        _Patch(fc="#e74c3c", hatch="///", label=f"dummy fill ({n_dummy})"),
        _Patch(fc="none", ec="#999", label=f"empty ({n_empty})"),
    ], loc="lower center", bbox_to_anchor=(0.5, -0.03), fontsize=7.5, frameon=True, ncol=4)
    ax.set_title(f"Finalized arrangement · Plate {pi:02d} · {fill:.1f}% filled\n"
                 f"{n_alt} replacement real · {n_dummy} dummy fill · {n_empty} empty seat", fontsize=10)
    fig.savefig(path, dpi=85)
    _plt.close(fig)


def generate_final(params, arrange_id, out_dir, media_base):
    """Regenerate the arrangement images for a finalized selection: kept seeds stay in
    place; each removed seat shows the seed the USER chose to place (via
    ReplacesSeed_ID) or is drawn EMPTY if the user left it unfilled. No auto-placement.
    Returns a list of {plateNo, fillPct, imageUrl, altCount, emptyCount, plateName}."""
    from .models import SeedArrangeDetail, SeedArrangePlate, SeedData

    R = _apply_globals(params)
    shape_set = SHAPE_SETS[params["shape"]]
    square_tol = _num(params, "squareTol")
    plate_d_val = _num(params, "plateD")
    # Plate names are assigned at finalize (continuous P{n}-{Ø}); read them back to show
    # the user. Falls back to a computed name if a plate row has none yet.
    name_by_plate = dict(
        SeedArrangePlate.objects.filter(arrange_id=arrange_id).values_list("plate_no", "plate_name")
    )
    raw_batches = params.get("batches") or []
    if isinstance(raw_batches, str):
        raw_batches = [raw_batches]
    batches = [str(b).strip() for b in raw_batches if str(b).strip()]

    # Re-derive the ORIGINAL layout (deterministic given the same seeds + params).
    # Apply the SAME exclude filter the run used (user excludes + the occupied seeds
    # frozen into occupiedExclude at job creation) so the re-pack reproduces it exactly.
    blocks = load_blocks_from_db(shape_set, square_tol, batches)
    exclude_raw = ((params.get("exclude") or "") + " " + (params.get("occupiedExclude") or "")).replace(",", " ").split()
    if exclude_raw:
        excluded = {s.strip() for s in exclude_raw if s.strip()}
        blocks = [b for b in blocks if str(b["stock"]).strip() not in excluded]
    queue = P._mixed_landscape(blocks)
    real_plates = []
    while queue:
        rp = P._mixed_one_plate(queue)
        if not rp:
            break
        real_plates.append(rp)

    # Only reconstruct the plates this arrangement ACTUALLY has. The re-pack runs against
    # the live TRN_SeedData, which can change between the original run and now (the
    # DiamondElement DB is reset externally) — without this bound a grown inventory would
    # re-pack far more plates and render dozens of images, hanging the request.
    from django.db.models import Max

    from .models import SeedArrange
    hdr = SeedArrange.objects.filter(arrange_id=arrange_id).first()
    n_plates = hdr.plate_no if (hdr and hdr.plate_no) else None
    if not n_plates:
        n_plates = (SeedArrangeDetail.objects.filter(arrange_id=arrange_id)
                    .aggregate(m=Max("plate_id"))["m"]) or len(real_plates)
    real_plates = real_plates[:n_plates]

    # Recommended reals: which are removed (is_final=0), keyed by StockNo → Seed_ID.
    rec = list(
        SeedArrangeDetail.objects.filter(arrange_id=arrange_id, seed_type=False, is_recommended=True)
        .values_list("seed_id", "is_final")
    )
    id2stock = {s.seed_id: str(s.stock_no or "").strip()
                for s in SeedData.objects.filter(seed_id__in=[sid for sid, _ in rec])}
    stock2id = {v: k for k, v in id2stock.items()}
    removed_ids = {sid for sid, fin in rec if not fin}

    # Replacements (user-placed or auto-filled): removed Seed_ID (ReplacesSeed_ID) →
    # ("alt", SeedData) for a real seed or ("dummy", DummySeedData) for a dummy fill.
    from .models import DummySeedData

    repl_real = list(
        SeedArrangeDetail.objects.filter(arrange_id=arrange_id, seed_type=False,
                                         is_recommended=False, is_final=True)
        .exclude(replaces_seed_id=None).values_list("replaces_seed_id", "seed_id")
    )
    repl_dummy = list(
        SeedArrangeDetail.objects.filter(arrange_id=arrange_id, seed_type=True,
                                         is_recommended=False, is_final=True)
        .exclude(replaces_seed_id=None).values_list("replaces_seed_id", "seed_id")
    )
    real_seed = {s.seed_id: s for s in SeedData.objects.filter(seed_id__in=[rid for _, rid in repl_real])}
    dummy_seed = {s.seed_id: s for s in DummySeedData.objects.filter(seed_id__in=[rid for _, rid in repl_dummy])}
    replaces_map = {}
    for removed_id, sid in repl_real:
        if sid in real_seed:
            replaces_map[removed_id] = ("alt", real_seed[sid])
    for removed_id, sid in repl_dummy:
        if sid in dummy_seed:
            replaces_map[removed_id] = ("dummy", dummy_seed[sid])

    final_dir = os.path.join(out_dir, "final")
    os.makedirs(final_dir, exist_ok=True)
    for f in os.listdir(final_dir):
        if f.endswith(".png"):
            os.remove(os.path.join(final_dir, f))

    circle_area = math.pi * R * R
    plates = []
    for idx, placed in enumerate(real_plates, 1):
        final_placed, n_alt, n_dummy, n_empty = [], 0, 0, 0
        for p in placed:
            sid = stock2id.get(str(p.get("stock", "")).strip())
            if sid in removed_ids:
                entry = replaces_map.get(sid)
                if entry is not None:  # a seed (real or dummy) fills this seat
                    kind, rs = entry
                    aw, ah = float(rs.length or 0), float(rs.width or 0)
                    if aw > p["w"] or ah > p["h"]:
                        aw, ah = ah, aw
                    aw, ah = min(aw, p["w"]), min(ah, p["h"])
                    final_placed.append({
                        "x": p["x"] + (p["w"] - aw) / 2, "y": p["y"] + (p["h"] - ah) / 2,
                        "w": aw, "h": ah, "stock": (str(rs.stock_no) if kind == "alt" else "DUMMY"),
                        "H": float(rs.height or 0.0), "kind": kind,
                    })
                    n_alt += kind == "alt"
                    n_dummy += kind == "dummy"
                else:  # left empty
                    final_placed.append({**p, "kind": "empty"})
                    n_empty += 1
            else:
                final_placed.append({**p, "kind": "kept"})

        fill = 100.0 * sum(s["w"] * s["h"] for s in final_placed if s.get("kind") != "empty") / circle_area
        path = os.path.join(final_dir, f"plate_{idx:02d}.png")
        _render_final_circle(final_placed, idx, R, fill, path)
        final_rows = _final_dim_rows(final_placed)
        plates.append({
            "plateNo": idx, "fillPct": round(fill, 1),
            "imageUrl": f"{media_base}/final/plate_{idx:02d}.png",
            "altCount": n_alt, "dummyCount": n_dummy, "emptyCount": n_empty,
            "seeds": final_rows,
            "plateName": name_by_plate.get(idx) or f"P{idx}-{plate_d_val:g}",
        })

        # Record the finalized image path/fill/covered area on the plate row (TRN_SeedPlate).
        from .models import SeedArrangePlate
        SeedArrangePlate.objects.filter(arrange_id=arrange_id, plate_no=idx).update(
            finalized_image_path=f"{media_base}/final/plate_{idx:02d}.png",
            finalized_fill_pct=round(fill, 2),
            finalized_covered_mm2=round(circle_area * fill / 100.0, 2),
            total_area_mm2=round(circle_area, 2),
            update_date=__import__("datetime").datetime.now(__import__("datetime").timezone.utc),
        )
        # Add the finalized image + detail table to the per-plate Excel — ONLY if it exists.
        _add_final_to_excel(os.path.join(out_dir, "excel", f"plate_{idx:02d}.xlsx"),
                            path, final_rows)

    return plates


def _add_final_to_excel(xlsx_path, final_png, final_rows=None):
    """Append the finalized plate's image AND its detail table to an existing per-plate
    Excel (built during Compare with the Arrange + Machine-Cut blocks). The FINALIZED
    block sits in columns Q..W, mirroring ARRANGE (A..G) and MACHINE-CUT (I..O):
    the image on top (Q3) and the seed-detail table below (Q20). No-op if no Excel."""
    if not os.path.exists(xlsx_path) or not os.path.exists(final_png):
        return
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    wb = load_workbook(xlsx_path)
    ws = wb.active
    # Normalize the title (strip any prior suffix so re-generate doesn't duplicate it).
    base = str(ws.cell(1, 1).value or "")
    for suffix in (" — Arrange · Machine-Cut · Finalized", " — Arrange vs Machine-Cut"):
        base = base.replace(suffix, "")
    ws.cell(1, 1).value = base + " — Arrange · Machine-Cut · Finalized"
    ws.cell(2, 17, "FINALIZED").font = Font(bold=True, color="1F4E78")   # Q2 header
    for existing in list(getattr(ws, "_images", [])):                    # avoid duplicate on re-generate
        try:
            if existing.anchor._from.col == 16:  # column Q
                ws._images.remove(existing)
        except Exception:
            pass
    img = XLImage(final_png)
    img.width = img.height = 300
    ws.add_image(img, "Q3")

    # FINALIZED detail table at Q20 (cols Q..W = 17..23), same format as the other two.
    if final_rows is not None:
        for r in range(20, ws.max_row + 1):          # clear any prior finalized table (re-generate)
            for c in range(17, 24):
                cell = ws.cell(r, c)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.font = Font()
        _write_table(ws, 20, 17, f"FINALIZED — {len(final_rows)} seeds", final_rows)
    for col in "QRSTUVW":
        ws.column_dimensions[col].width = 13
    wb.save(xlsx_path)
